"""Pyodide 진입점.

브라우저(Web Worker)에서 호출되어 입력 엑셀 bytes를 받아 결과 엑셀 bytes를 반환한다.
기존 4개 모듈(parse_input, schedule, review, render_bracket)을 함수로 import해서 직접 호출.
"""
from __future__ import annotations

import json
from io import BytesIO

from openpyxl import load_workbook

from parse_input import (
    parse_players,
    parse_courts,
    build_schedule_slots,
    attach_available_slots,
    clamp_mixed_wish,
    min_to_hhmm,
    parse_member_settings,
    parse_seed,
    _max_games_streak,
    COUPLES_DEFAULT,
)
from schedule import solve
from review import compute_scores
from render_bracket import render
from build_template import build_template, build_member_settings
from history import extract_pairs_from_bracket_xlsx, DEFAULT_W1, DEFAULT_W2


def _parse_bytes(xlsx_bytes: bytes) -> dict:
    wb = load_workbook(BytesIO(xlsx_bytes), data_only=True)
    if "참가자" not in wb.sheetnames:
        raise ValueError("입력 엑셀에 '참가자' 시트가 없습니다.")
    if "코트" not in wb.sheetnames:
        raise ValueError("입력 엑셀에 '코트' 시트가 없습니다.")

    players, perr = parse_players(wb["참가자"])
    courts, cerr = parse_courts(wb["코트"])
    errs = perr + cerr
    if errs:
        raise ValueError("입력 오류:\n  - " + "\n  - ".join(errs))

    if len(players) < 4:
        raise ValueError(f"참가자가 {len(players)}명입니다. 최소 4명 필요.")
    if not courts:
        raise ValueError("사용 가능한 코트가 없습니다.")

    schedule_slots = build_schedule_slots(courts)
    attach_available_slots(players, schedule_slots)

    warnings = []

    # 씨드대진(선택) — 사용자가 직접 고정한 자리. 시트가 없거나 비어 있으면 pins=[]이고
    # 이후 동작은 종전과 완전히 같다.
    pins = []
    if "씨드대진" in wb.sheetnames:
        pins, seed_errs, seed_warns = parse_seed(wb["씨드대진"], players, schedule_slots)
        if seed_errs:
            raise ValueError("씨드대진 시트 오류:\n  - " + "\n  - ".join(seed_errs))
        warnings.extend(seed_warns)

    males = [p for p in players if p["gender"] == "M"]
    females = [p for p in players if p["gender"] == "F"]
    if len(males) < 4:
        warnings.append(f"남자가 {len(males)}명이라 남자복식 불가. 혼복만 가능.")
    if len(females) < 4:
        warnings.append(f"여자가 {len(females)}명이라 여자복식 불가. 혼복만 가능.")

    for sl in schedule_slots:
        avail = [p for p in players if sl["slot_start"] in p["available_slots"]]
        if len(avail) < 4:
            t = f"{min_to_hhmm(sl['slot_start'])}~{min_to_hhmm(sl['slot_end'])}"
            warnings.append(f"슬롯 {t}: 가용 인원 {len(avail)}명 — 일부 코트 공석 가능")

    for p in players:
        if len(p["available_slots"]) == 0:
            warnings.append(f"'{p['name']}': 가용 슬롯 없음")
        if p["max_games"] is not None and p["max_games"] > len(p["available_slots"]):
            warnings.append(
                f"'{p['name']}': 최대게임수({p['max_games']}) > 가용 슬롯수({len(p['available_slots'])}) — 자연 제한됨"
            )
        # 보장 한도는 '가용 슬롯 수'가 아니라 '본인 연속게임 설정 하에 뛸 수 있는 최대'다.
        # (CLI parse_input.main과 같은 기준 — 여기가 어긋나면 웹만 경고를 놓친다)
        streak = p.get("streak") or ""
        cap = _max_games_streak(p["available_slots"], streak)
        if p.get("min_games") and p["min_games"] > cap:
            if streak == "no2":
                cap_label = "연속금지 하에 가능한 최대"
            elif streak == "ok3":
                cap_label = "가용 슬롯수"
            else:
                cap_label = "3연속 없이 가능한 최대"
            warnings.append(
                f"'{p['name']}': 최소게임수({p['min_games']}) > {cap_label}({cap}게임) — {cap}게임까지만 보장됨"
            )

    # 혼복희망 클램프 — CLI(parse_input.main)와 같은 방어를 웹에도 건다.
    warnings.extend(clamp_mixed_wish(players))

    # 최소게임수 합계가 전체 자리(코트×슬롯×4)보다 많으면 다 지킬 수 없다 — 미리 알림
    total_seats = 0
    for sl in schedule_slots:
        n_avail = sum(1 for p in players if sl["slot_start"] in p["available_slots"])
        total_seats += min(len(sl["courts"]), n_avail // 4) * 4
    min_sum = sum(
        min(p["min_games"], _max_games_streak(p["available_slots"], p.get("streak") or ""))
        for p in players if p.get("min_games"))
    if min_sum > total_seats:
        warnings.append(f"최소게임수 합계({min_sum})가 전체 배정 가능 자리({total_seats})를 넘습니다 — 일부는 보장 못 할 수 있음")

    return {
        "courts": courts,
        "players": players,
        "schedule_slots": schedule_slots,
        "pins": pins,
        "warnings": warnings,
    }


def _to_bytes(x):
    """Pyodide에서 넘어온 JS Uint8Array/버퍼 또는 bytes를 파이썬 bytes로 정규화. None은 그대로."""
    if x is None:
        return None
    if hasattr(x, "to_py"):
        x = x.to_py()
    if isinstance(x, (bytes, bytearray)):
        return bytes(x)
    try:
        return bytes(x)
    except TypeError:
        return None


def _build_hist_pairs(prev_specs) -> list:
    """prev_specs = [(xlsx_bytes|None, weight), ...] (우선순위 높은 것 먼저).

    각 이전 대진표 엑셀에서 같은 팀 페어를 뽑아 가중치 합산한 [[a, b, w], ...]를 만든다.
    """
    weighted: dict = {}
    for raw, weight in prev_specs:
        b = _to_bytes(raw)
        if not b:
            continue
        try:
            for k in extract_pairs_from_bracket_xlsx(BytesIO(b)):
                weighted[k] = weighted.get(k, 0.0) + weight
        except Exception:
            continue
    return [[a, b, round(w, 4)] for (a, b), w in weighted.items()]


def generate_bracket(
    xlsx_bytes: bytes,
    date_str: str = "",
    seed: int = 7,
    iters: int = 20,
    title: str = "우리 테니스 클럽 대진표",
    prev1_bytes=None,
    prev2_bytes=None,
    refine: int = 4,
    kicks: int = 25,
    members_bytes=None,
) -> dict:
    """입력 엑셀 bytes → {xlsx: bytes, review: dict, summary: dict}.

    prev1_bytes(1주전)/prev2_bytes(2주전)를 주면 그 대진표들과 겹치는 페어를 최대한 피한다.
    (우리 멤버끼리일 때만 실제 반영 — 교류전이면 자동 무시)
    members_bytes(멤버 설정 엑셀)를 주면 그 파일의 부부 시트를, 없으면 내장 기본값을 쓴다.
    브라우저에서 호출 후 xlsx 필드를 Blob으로 만들어 다운로드.

    iters=초안 생성 횟수, refine/kicks=공백·대기 줄이는 로컬 개선 강도.
    브라우저(Pyodide)는 네이티브보다 느리므로 기본값을 CLI보다 낮게 잡는다.
    """
    parsed = _parse_bytes(xlsx_bytes)
    hist_pairs = _build_hist_pairs([(prev1_bytes, DEFAULT_W1), (prev2_bytes, DEFAULT_W2)])

    # 부부 페어: 업로드한 멤버 설정 > 내장 기본값
    couples = None
    members_uploaded = False
    mb = _to_bytes(members_bytes)
    if mb:
        try:
            couples = parse_member_settings(BytesIO(mb))
            members_uploaded = True
        except Exception:
            couples = None
    if couples is None:
        couples = [list(c) for c in COUPLES_DEFAULT]
    parsed["couples"] = couples
    names = {p["name"] for p in parsed["players"]}
    couples_present = sum(1 for c in couples if c[0] in names and c[1] in names)

    bracket = solve(
        parsed["players"], parsed["schedule_slots"],
        seed=seed, iters=iters, hist_pairs=hist_pairs, refine=refine, kicks=kicks,
        couples=couples, pins=parsed.get("pins"),
    )
    review = compute_scores(parsed, bracket, hist_pairs)

    out_buf = BytesIO()
    render(parsed, bracket, out_buf, date_str, title)
    out_buf.seek(0)

    distinct_clubs = {p.get("club", "") for p in parsed["players"] if p.get("club", "")}
    is_exchange = len(distinct_clubs) > 1

    return {
        "xlsx_bytes": out_buf.getvalue(),
        "review": review,
        "summary": {
            "players": len(parsed["players"]),
            "courts": len(parsed["courts"]),
            "slots": len(parsed["schedule_slots"]),
            "matches": len(bracket["matches"]),
            "warnings": parsed["warnings"],
            "history_pairs": len(hist_pairs),
            "history_repeat_pairs": review["scores"].get("history_repeat_pairs", 0),
            "history_ignored_exchange": bool(hist_pairs) and is_exchange,
            "members_uploaded": members_uploaded,
            "couples_total": len(couples),
            "couples_present": couples_present,
            "seed_seats": review["scores"].get("seed_seats", 0),
            "seed_seats_kept": review["scores"].get("seed_seats_kept", 0),
            "seed_matches": review["scores"].get("seed_matches", 0),
        },
    }


def build_empty_template_bytes(prefill: str = "") -> bytes:
    """빈 입력 양식 엑셀을 bytes로 반환. (prefill="image"면 이미지 기반 사전채움)"""
    buf = BytesIO()
    build_template(buf, prefill=prefill)
    buf.seek(0)
    return buf.getvalue()


def build_member_settings_bytes() -> bytes:
    """멤버 설정 엑셀(멤버·부부 시트, 내장 기본값)을 bytes로 반환."""
    buf = BytesIO()
    build_member_settings(buf)
    buf.seek(0)
    return buf.getvalue()


def generate_bracket_json_result(xlsx_bytes_bin, date_str="", seed=7, iters=20,
                                 title="우리 테니스 클럽 대진표", prev1_bytes=None, prev2_bytes=None,
                                 refine=4, kicks=25, members_bytes=None):
    """Pyodide JS 호출용 wrapper. JS의 Uint8Array를 받아 dict 반환.

    xlsx 결과는 별도 함수로 가져가도록 분리하지 않고, 결과 dict에 bytes 그대로 포함.
    prev1_bytes/prev2_bytes(있으면)로 지난주/2주전 페어를 회피.
    members_bytes(있으면)로 부부 페어 설정을 덮어쓴다(없으면 내장 기본값).
    """
    main_bytes = _to_bytes(xlsx_bytes_bin)
    return generate_bracket(
        main_bytes, date_str=date_str, seed=seed, iters=iters, title=title,
        prev1_bytes=prev1_bytes, prev2_bytes=prev2_bytes, refine=refine, kicks=kicks,
        members_bytes=members_bytes,
    )
