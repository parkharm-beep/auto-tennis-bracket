"""Render a finalized tennis bracket to an Excel file matching the handwritten layout.

Usage:
    python render_bracket.py --parsed <01_parsed.json> --bracket <02_bracket.json>
                             --out <bracket.xlsx> [--date "26.5.30"] [--title "..."]
"""
from __future__ import annotations

import argparse
import json
import os
from collections import defaultdict
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


COURT_COLORS = [
    "FFFF99",  # yellow
    "FFCC99",  # orange
    "99CCFF",  # sky blue
    "C5E0B4",  # light green
    "FFB6C1",  # pink
    "D5A6E0",  # light purple
]
EMPTY_FILL = PatternFill("solid", fgColor="E0E0E0")
HEADER_FILL = PatternFill("solid", fgColor="BDD7EE")
TITLE_FILL = PatternFill("solid", fgColor="FFFFFF")
VS_FILL = PatternFill("solid", fgColor="F2F2F2")
SUMMARY_NUM_FILL = PatternFill("solid", fgColor="FFF2CC")
SUMMARY_GAME_FILL = PatternFill("solid", fgColor="FFFF99")

THIN = Side(border_style="thin", color="333333")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)

FONT_TITLE = Font(name="맑은 고딕", size=18, bold=True)
FONT_HEADER = Font(name="맑은 고딕", size=11, bold=True)
FONT_NAME = Font(name="맑은 고딕", size=11, bold=True)
FONT_TIME = Font(name="맑은 고딕", size=10)
FONT_SMALL = Font(name="맑은 고딕", size=9, color="666666")
FONT_VS = Font(name="맑은 고딕", size=10, bold=True, color="999999")


def min_to_hhmm(v: int) -> str:
    return f"{v // 60:02d}:{v % 60:02d}"


FONT_STAT = Font(name="맑은 고딕", size=10)
FONT_STAT_WARN = Font(name="맑은 고딕", size=10, bold=True, color="C00000")


def _collect_person_stats(player_stats: list[dict], matches: list[dict]) -> dict:
    """통계 시트용 사람별 지표 계산 (생성 시점의 bracket 기준).

    최대 연속 게임수 / 최대 연속 휴식(경기 사이 최장 공백) / 1시간+ 공백 횟수 /
    총 대기시간 / 복식 종류별 게임수 / 파트너 수·같은 짝 반복.
    """
    type_cnt = {s["id"]: {"M": 0, "F": 0, "X": 0} for s in player_stats}
    partners = {s["id"]: defaultdict(int) for s in player_stats}
    for m in matches:
        for team in (m["team1"], m["team2"]):
            a, b = team
            type_cnt[a][m["type"]] += 1
            type_cnt[b][m["type"]] += 1
            partners[a][b] += 1
            partners[b][a] += 1

    out = {}
    for s in player_stats:
        pid = s["id"]
        slots = sorted(s["slots_played"])
        gaps = [slots[i] - slots[i - 1] - 30 for i in range(1, len(slots))]
        best = run = 0
        prev = None
        for sl in slots:
            run = run + 1 if (prev is not None and sl - prev == 30) else 1
            best = max(best, run)
            prev = sl
        start_delay = s.get("start_delay")
        if start_delay is None:
            start_delay = max(0, slots[0] - s.get("eff_in", s["in_min"])) if slots else 0
        pd = partners[pid]
        out[pid] = dict(
            slots=slots,
            max_streak=best,
            max_rest=max((g for g in gaps if g > 0), default=0),
            long_gaps=sum(1 for g in gaps if g >= 60),
            start_delay=start_delay,
            total_idle=start_delay + sum(g for g in gaps if g > 0),
            types=type_cnt[pid],
            partner_n=len(pd),
            partner_rep=sum(c - 1 for c in pd.values() if c > 1),
        )
    return out


def _build_stats_sheet(wb, bracket: dict, is_exchange: bool, clubs_ordered: list) -> None:
    """'통계' 시트: 사람별 통계 테이블 + 전체 요약."""
    matches = bracket["matches"]
    player_stats = bracket["player_stats"]
    per = _collect_person_stats(player_stats, matches)

    ws = wb.create_sheet("통계")

    cols = [("번호", 5), ("이름", 12)]
    if is_exchange:
        cols.append(("클럽", 11))
    cols += [
        ("성별", 5), ("구력", 5), ("참석", 12),
        ("게임수", 7), ("남복", 6), ("여복", 6), ("혼복", 6),
        ("첫 경기", 8), ("마지막 경기", 10),
        ("첫 대기(분)", 10), ("최대 연속 게임", 9), ("최대 연속 휴식(분)", 10),
        ("1시간+ 공백(회)", 9), ("총 대기(분)", 10),
        ("파트너 수", 8), ("같은 짝 반복", 8),
    ]
    for i, (_, w) in enumerate(cols, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    ws.cell(row=1, column=1, value="사람별 통계 — 생성 시점 기준 (대진표 시트를 손으로 고쳐도 자동 갱신되지 않음)")
    ws.cell(row=1, column=1).font = FONT_SMALL
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(cols))

    HEADER_ROW = 2
    for i, (title, _) in enumerate(cols, start=1):
        c = ws.cell(row=HEADER_ROW, column=i, value=title)
        c.font = FONT_HEADER
        c.fill = HEADER_FILL
        c.alignment = CENTER
        c.border = BORDER
    ws.row_dimensions[HEADER_ROW].height = 28

    # 정렬: 명단 패널과 동일(가용 슬롯 많은 순). 교류전이면 클럽별로 묶는다.
    if is_exchange:
        ordered = []
        for cname in clubs_ordered:
            ordered += sorted([s for s in player_stats if s.get("club", "") == cname],
                              key=lambda s: -s["available_slots"])
        ordered += sorted([s for s in player_stats if s.get("club", "") not in clubs_ordered],
                          key=lambda s: -s["available_slots"])
    else:
        ordered = sorted(player_stats, key=lambda s: -s["available_slots"])

    for idx, s in enumerate(ordered):
        p = per[s["id"]]
        played = bool(p["slots"])
        row_vals = [idx + 1, display_name(s)]
        if is_exchange:
            row_vals.append(s.get("club", ""))
        row_vals += [
            "남" if s["gender"] == "M" else "여",
            s["exp"],
            f"{min_to_hhmm(s['in_min'])}~{min_to_hhmm(s['out_min'])}",
            s["games"],
            p["types"]["M"], p["types"]["F"], p["types"]["X"],
            min_to_hhmm(p["slots"][0]) if played else "-",
            min_to_hhmm(p["slots"][-1] + 30) if played else "-",
            p["start_delay"] if played else "-",
            p["max_streak"],
            p["max_rest"] if len(p["slots"]) >= 2 else "-",
            p["long_gaps"],
            p["total_idle"] if played else "-",
            p["partner_n"],
            p["partner_rep"],
        ]
        r = HEADER_ROW + 1 + idx
        for ci, v in enumerate(row_vals, start=1):
            cell = ws.cell(row=r, column=ci, value=v)
            cell.font = FONT_STAT
            cell.alignment = CENTER
            cell.border = BORDER
        # 눈에 띄어야 하는 값: 1시간 이상 휴식 / 1시간+ 공백 발생
        off = 1 if is_exchange else 0
        if isinstance(row_vals[13 + off], int) and row_vals[13 + off] >= 60:
            ws.cell(row=r, column=14 + off).font = FONT_STAT_WARN
        if row_vals[14 + off]:
            ws.cell(row=r, column=15 + off).font = FONT_STAT_WARN
        if s["membership"] == "게스트":
            ws.cell(row=r, column=2).font = Font(name="맑은 고딕", size=10, italic=True)

    ws.freeze_panes = f"A{HEADER_ROW + 1}"

    # ── 전체 요약 ──
    base = HEADER_ROW + len(ordered) + 2
    games = [s["games"] for s in player_stats]
    total_idle_all = sum(per[s["id"]]["total_idle"] for s in player_stats if per[s["id"]]["slots"])
    long_gap_total = sum(per[s["id"]]["long_gaps"] for s in player_stats)
    long_gap_people = sum(1 for s in player_stats if per[s["id"]]["long_gaps"])
    finish_all = max((per[s["id"]]["slots"][-1] + 30 for s in player_stats if per[s["id"]]["slots"]),
                     default=None)
    tc = bracket.get("type_count", {})
    rows = [
        ("전체 요약", ""),
        ("총 매치 수", f"{len(matches)}  (남복 {tc.get('M', 0)} · 여복 {tc.get('F', 0)} · 혼복 {tc.get('X', 0)})"),
        ("참가자 수", len(player_stats)),
        ("게임수 평균 / 최대 / 최소",
         f"{round(sum(games) / max(1, len(games)), 2)} / {max(games, default=0)} / {min(games, default=0)}"),
        ("1시간 이상 공백", f"{long_gap_total}건 / {long_gap_people}명"),
        ("총 대기시간 합계", f"{total_idle_all}분"),
        ("마지막 경기 종료", min_to_hhmm(finish_all) if finish_all else "-"),
    ]
    if is_exchange:
        club_counts = defaultdict(int)
        for s in player_stats:
            club_counts[s.get("club", "")] += 1
        rows.append(("교류전 클럽", "  ·  ".join(f"{c} {club_counts[c]}명" for c in clubs_ordered)))
    for i, (k, v) in enumerate(rows):
        kc = ws.cell(row=base + i, column=1, value=k)
        kc.font = FONT_HEADER
        ws.merge_cells(start_row=base + i, start_column=1, end_row=base + i, end_column=3)
        vc = ws.cell(row=base + i, column=4, value=v)
        vc.font = FONT_STAT
        vc.alignment = Alignment(horizontal="left", vertical="center")
        ws.merge_cells(start_row=base + i, start_column=4, end_row=base + i, end_column=len(cols))


def display_name(p_stat: dict) -> str:
    n = p_stat["name"]
    if p_stat["membership"] == "게스트":
        return f"{n}(G)"
    return n


def render(parsed: dict, bracket: dict, out_path: str, date_str: str, title: str) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "대진표"

    courts = parsed["courts"]
    schedule_slots = parsed["schedule_slots"]
    matches = bracket["matches"]
    player_stats = bracket["player_stats"]

    clubs_present = {s.get("club", "") for s in player_stats if s.get("club", "")}
    is_exchange = len(clubs_present) > 1

    def _club_order_key(c):
        # 정회원이 많은 클럽(=홈)을 왼쪽으로
        members = sum(1 for s in player_stats
                      if s.get("club", "") == c and s.get("membership") == "정회원")
        return (-members, c)
    clubs_ordered = sorted(clubs_present, key=_club_order_key)
    two_club_mode = is_exchange and len(clubs_present) == 2

    matches_by_slot_court = defaultdict(dict)
    for m in matches:
        matches_by_slot_court[m["slot_start"]][m["court"]] = m

    # 컬럼 레이아웃
    # 1: 구분(번호+시간), 2: 성함/결과 라벨
    # 코트별 4컬럼: t1a, t1b, VS, t2a, t2b (= 5컬럼)
    # 우측 패널: 번호 | 이름 | 게임수 | (간격) | 번호 | 이름 | 게임수
    LABEL_COL_START = 1
    LABEL_COL_END = 2
    COURTS_COL_START = 3
    cols_per_court = 5
    courts_col_end = COURTS_COL_START + cols_per_court * len(courts) - 1

    summary_left_start = courts_col_end + 2
    SUM_NUM_W, SUM_NAME_W, SUM_GAME_W = 4, 12, 7
    summary_col_count = 3 + 1 + 3
    summary_right_end = summary_left_start + summary_col_count - 1

    # 컬럼 너비
    ws.column_dimensions[get_column_letter(LABEL_COL_START)].width = 9
    ws.column_dimensions[get_column_letter(LABEL_COL_END)].width = 7
    for i, _ in enumerate(courts):
        base = COURTS_COL_START + i * cols_per_court
        ws.column_dimensions[get_column_letter(base)].width = 10
        ws.column_dimensions[get_column_letter(base + 1)].width = 10
        ws.column_dimensions[get_column_letter(base + 2)].width = 4
        ws.column_dimensions[get_column_letter(base + 3)].width = 10
        ws.column_dimensions[get_column_letter(base + 4)].width = 10
    ws.column_dimensions[get_column_letter(courts_col_end + 1)].width = 1
    for off, w in zip(range(7), [SUM_NUM_W, SUM_NAME_W, SUM_GAME_W, 1, SUM_NUM_W, SUM_NAME_W, SUM_GAME_W]):
        ws.column_dimensions[get_column_letter(summary_left_start + off)].width = w

    # 행 1: 타이틀
    ws.cell(row=1, column=LABEL_COL_START, value=title)
    ws.merge_cells(start_row=1, start_column=LABEL_COL_START,
                   end_row=1, end_column=courts_col_end)
    ws.cell(row=1, column=LABEL_COL_START).font = FONT_TITLE
    ws.cell(row=1, column=LABEL_COL_START).alignment = CENTER
    ws.row_dimensions[1].height = 30

    # 날짜 (타이틀 영역 우측)
    ws.cell(row=1, column=summary_left_start, value=date_str)
    ws.merge_cells(start_row=1, start_column=summary_left_start,
                   end_row=1, end_column=summary_right_end)
    ws.cell(row=1, column=summary_left_start).font = FONT_HEADER
    ws.cell(row=1, column=summary_left_start).alignment = Alignment(horizontal="right", vertical="center")

    # 행 2-3: 헤더
    HEADER_ROW = 2
    ws.cell(row=HEADER_ROW, column=LABEL_COL_START, value="구분")
    ws.merge_cells(start_row=HEADER_ROW, start_column=LABEL_COL_START,
                   end_row=HEADER_ROW, end_column=LABEL_COL_END)
    c = ws.cell(row=HEADER_ROW, column=LABEL_COL_START)
    c.font = FONT_HEADER
    c.fill = HEADER_FILL
    c.alignment = CENTER
    c.border = BORDER

    for i, court in enumerate(courts):
        base = COURTS_COL_START + i * cols_per_court
        ws.cell(row=HEADER_ROW, column=base, value=f"{court['name']}번코트")
        ws.merge_cells(start_row=HEADER_ROW, start_column=base,
                       end_row=HEADER_ROW, end_column=base + cols_per_court - 1)
        hc = ws.cell(row=HEADER_ROW, column=base)
        hc.font = FONT_HEADER
        hc.fill = PatternFill("solid", fgColor=COURT_COLORS[i % len(COURT_COLORS)])
        hc.alignment = CENTER
        hc.border = BORDER

    # 우측 패널 헤더
    _left_hdr = clubs_ordered[0] if two_club_mode else "참가자"
    _right_hdr = clubs_ordered[1] if two_club_mode else "참가자"
    ws.cell(row=HEADER_ROW, column=summary_left_start, value="")
    ws.cell(row=HEADER_ROW, column=summary_left_start + 1, value=_left_hdr)
    ws.cell(row=HEADER_ROW, column=summary_left_start + 2, value="게임수")
    ws.cell(row=HEADER_ROW, column=summary_left_start + 4, value="")
    ws.cell(row=HEADER_ROW, column=summary_left_start + 5, value=_right_hdr)
    ws.cell(row=HEADER_ROW, column=summary_left_start + 6, value="게임수")
    for off in (0, 1, 2, 4, 5, 6):
        c = ws.cell(row=HEADER_ROW, column=summary_left_start + off)
        c.font = FONT_HEADER
        c.fill = HEADER_FILL
        c.alignment = CENTER
        c.border = BORDER

    # 데이터 행: 슬롯별로 2행 (성함, 결과) — 각 행 높이 22
    data_start_row = HEADER_ROW + 1
    for slot_idx, slot in enumerate(schedule_slots):
        name_row = data_start_row + slot_idx * 2
        result_row = name_row + 1
        ws.row_dimensions[name_row].height = 24
        ws.row_dimensions[result_row].height = 24

        ws.cell(row=name_row, column=LABEL_COL_START,
                value=f"{slot_idx + 1}번 게임\n{min_to_hhmm(slot['slot_start'])}\n~{min_to_hhmm(slot['slot_end'])}")
        ws.merge_cells(start_row=name_row, start_column=LABEL_COL_START,
                       end_row=result_row, end_column=LABEL_COL_START)
        c = ws.cell(row=name_row, column=LABEL_COL_START)
        c.font = FONT_TIME
        c.alignment = CENTER
        c.border = BORDER

        ws.cell(row=name_row, column=LABEL_COL_END, value="성함")
        ws.cell(row=result_row, column=LABEL_COL_END, value="결과")
        for r in (name_row, result_row):
            c = ws.cell(row=r, column=LABEL_COL_END)
            c.font = FONT_HEADER
            c.alignment = CENTER
            c.border = BORDER
            c.fill = HEADER_FILL

        for i, court in enumerate(courts):
            base = COURTS_COL_START + i * cols_per_court
            court_color = PatternFill("solid", fgColor=COURT_COLORS[i % len(COURT_COLORS)])

            for col_off in range(cols_per_court):
                for r in (name_row, result_row):
                    cell = ws.cell(row=r, column=base + col_off, value=None)
                    cell.border = BORDER
                    cell.alignment = CENTER

            court_active = slot["slot_start"] in court["slots"]
            if not court_active:
                continue

            m = matches_by_slot_court.get(slot["slot_start"], {}).get(court["name"])
            if m is None:
                continue

            stats_by_id = {s["id"]: s for s in player_stats}
            for k, p_id in enumerate(m["team1"]):
                cell = ws.cell(row=name_row, column=base + k, value=display_name(stats_by_id[p_id]))
                cell.font = FONT_NAME
                cell.fill = court_color
                cell.alignment = CENTER
                cell.border = BORDER
            vs_cell = ws.cell(row=name_row, column=base + 2, value="VS")
            vs_cell.font = FONT_VS
            vs_cell.fill = VS_FILL
            vs_cell.alignment = CENTER
            vs_cell.border = BORDER
            for k, p_id in enumerate(m["team2"]):
                cell = ws.cell(row=name_row, column=base + 3 + k, value=display_name(stats_by_id[p_id]))
                cell.font = FONT_NAME
                cell.fill = court_color
                cell.alignment = CENTER
                cell.border = BORDER

            if m["type"] == "X":
                vs_cell.value = "혼"

    # 우측 패널: 참가자 + 게임수
    # 게임수는 정적 숫자가 아니라 COUNTIF 수식 — 대진표 칸의 이름을 직접 고치면 자동 재계산된다.
    # 교류전(클럽 2개)일 때는 좌/우 열을 클럽별로 분리.
    panel_start = data_start_row
    last_grid_row = data_start_row + len(schedule_slots) * 2 - 1
    grid_abs = (f"${get_column_letter(COURTS_COL_START)}${data_start_row}:"
                f"${get_column_letter(courts_col_end)}${last_grid_row}")

    def write_entry(row, num_col, seq_num, p):
        c1 = ws.cell(row=row, column=num_col, value=seq_num)
        c1.font = FONT_HEADER
        c1.fill = SUMMARY_NUM_FILL
        c1.alignment = CENTER
        c1.border = BORDER
        ws.merge_cells(start_row=row, start_column=num_col, end_row=row + 1, end_column=num_col)

        name_cell = ws.cell(row=row, column=num_col + 1, value=display_name(p))
        name_cell.font = FONT_NAME
        name_cell.fill = SUMMARY_NUM_FILL
        name_cell.alignment = CENTER
        name_cell.border = BORDER
        info = f"{min_to_hhmm(p['in_min'])}~{min_to_hhmm(p['out_min'])}"
        if p.get("min_games"):
            info += f" / 최소 {p['min_games']}게임"
        if p.get("max_games") is not None:
            info += f" / 최대 {p['max_games']}게임"
        if is_exchange and p.get("club"):
            info += f" · {p['club']}"
        ci = ws.cell(row=row + 1, column=num_col + 1, value=info)
        ci.font = FONT_SMALL
        ci.alignment = CENTER
        ci.border = BORDER

        # 게임수 = 대진표 영역에서 이 사람 이름이 나오는 횟수 (수정 시 자동 갱신)
        name_ref = f"{get_column_letter(num_col + 1)}{row}"
        g = ws.cell(row=row, column=num_col + 2, value=f"=COUNTIF({grid_abs},{name_ref})")
        g.font = FONT_NAME
        g.fill = SUMMARY_GAME_FILL
        g.alignment = CENTER
        g.border = BORDER
        ws.merge_cells(start_row=row, start_column=num_col + 2, end_row=row + 1, end_column=num_col + 2)

    def write_total(row, game_col, first_row, last_row):
        rng = f"{get_column_letter(game_col)}{first_row}:{get_column_letter(game_col)}{last_row}"
        t = ws.cell(row=row, column=game_col, value=f"=SUM({rng})")
        t.font = FONT_HEADER
        t.fill = SUMMARY_GAME_FILL
        t.alignment = CENTER
        t.border = BORDER

    left_num_col = summary_left_start
    right_num_col = summary_left_start + 4

    if two_club_mode:
        # 좌=클럽1, 우=클럽2 로 분리
        col_lists = [
            (left_num_col, sorted([s for s in player_stats if s.get("club", "") == clubs_ordered[0]],
                                  key=lambda s: -s["available_slots"])),
            (right_num_col, sorted([s for s in player_stats if s.get("club", "") == clubs_ordered[1]],
                                   key=lambda s: -s["available_slots"])),
        ]
    else:
        stats_sorted = sorted(player_stats, key=lambda s: -s["available_slots"])
        half = (len(stats_sorted) + 1) // 2
        col_lists = [
            (left_num_col, stats_sorted[:half]),
            (right_num_col, stats_sorted[half:]),
        ]

    max_len = max((len(lst) for _, lst in col_lists), default=0)
    panel_end_row = panel_start + max_len * 2
    for num_col, lst in col_lists:
        for idx, p in enumerate(lst):
            write_entry(panel_start + idx * 2, num_col, idx + 1, p)
        if lst:
            write_total(panel_end_row, num_col + 2, panel_start, panel_end_row - 1)

    ws.sheet_view.showGridLines = False
    ws.print_options.horizontalCentered = True
    ws.page_setup.orientation = "landscape"

    # 두 번째 시트: 사람별 통계 + 전체 요약
    _build_stats_sheet(wb, bracket, is_exchange, clubs_ordered)

    wb.save(out_path)


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--parsed", required=True)
    ap.add_argument("--bracket", required=True)
    ap.add_argument("--out", required=True)
    ap.add_argument("--date", default="")
    ap.add_argument("--title", default="우리 테니스 클럽 대진표")
    args = ap.parse_args()

    with open(args.parsed, "r", encoding="utf-8") as f:
        parsed = json.load(f)
    with open(args.bracket, "r", encoding="utf-8") as f:
        bracket = json.load(f)

    render(parsed, bracket, args.out, args.date, args.title)
    print(f"[OK] 대진표 출력: {args.out}")


if __name__ == "__main__":
    main()
