"""Parse and validate a filled tennis bracket input Excel file.

Usage:
    python parse_input.py --in <input.xlsx> --out <parsed.json>
"""
from __future__ import annotations

import argparse
import json
import os
import sys
from openpyxl import load_workbook


# ---------------------------------------------------------------------------
# 클럽 멤버 설정 기본값 — 멤버 설정 파일(클럽멤버_설정.xlsx)이 없을 때 사용.
# 26.8.2 카톡 투표 명단(25명, 사용자 확정본) 기준.
# ---------------------------------------------------------------------------
MEMBERS_DEFAULT = [
    # (카톡아이디, 실제이름, 성별, 구력, 메모)
    # 성별·구력은 26.8.2 사용자 확정본(테니스_입력양식_사전채움_25명.xlsx) 기준.
    ("김도윤",      "김도윤", "남", 10, ""),
    ("김효순",      "김효순", "남", 5,  ""),
    ("남궁석",      "남궁석", "남", 10, ""),
    ("노남숙",      "노남숙", "여", 3,  ""),
    ("명수기❤️",    "서명숙", "여", 5,  ""),
    ("민기준",      "민기준", "남", 5,  ""),
    ("박경수",      "박경수", "남", 5,  ""),
    ("박진우",      "박진우", "남", 7,  ""),
    ("서종수",      "서종수", "남", 5,  ""),
    ("성현",        "경성현", "남", 4,  ""),
    ("원유철",      "원유철", "남", 4,  ""),
    ("이강진",      "이강진", "남", 4,  ""),
    ("이성돈",      "이성돈", "남", 10, ""),
    ("이성수",      "이성수", "남", 7,  ""),
    ("이지은",      "이지은", "여", 3,  ""),
    ("임성훈",      "임성훈", "남", 5,  ""),
    ("정재동",      "정재동", "남", 10, ""),
    ("정진락",      "정진락", "남", 10, ""),
    ("정희",        "정정희", "여", 5,  ""),
    ("최종인",      "최종인", "남", 10, ""),
    ("한병익",      "한병익", "남", 10, ""),
    ("혜선",        "전혜선", "여", 7,  ""),
    ("HJ Shin",     "신혁재", "남", 5,  ""),
    ("Joonhak Kim", "김준학", "남", 10, ""),
    ("Mira",        "방미라", "여", 3,  ""),
]

# 부부 페어: (이름1, 이름2, 혼복에서 부부페어를 원하는지)
# 원함=True → 혼복이 나올 때 부부가 같은 팀이 되는 것을 우대.
# 피함=False → 부부를 같은 팀으로 묶지 않게 회피(소프트).
COUPLES_DEFAULT = [
    ("박경수", "서명숙", False),
    ("한병익", "전혜선", True),
    ("신혁재", "방미라", False),
    ("원유철", "이지은", False),
]

_COUPLE_WANT_WORDS = ("원함", "희망", "허용", "O", "o", "예", "y", "Y")


def parse_member_roster(source) -> list:
    """멤버 설정 엑셀의 '멤버' 시트에서 명단을 읽는다.

    반환: [{"kakao": 카톡아이디, "name": 실제이름, "gender": '남'/'여'/'', "exp": int|None}, ...]
    입력 엑셀 초안 자동 생성 시 참석자 이름(카톡아이디)을 실명·성별·구력으로 변환하는 근거.
    멤버 시트가 없으면 빈 목록(호출부에서 MEMBERS_DEFAULT 사용).
    """
    wb = load_workbook(source, data_only=True)
    if "멤버" not in wb.sheetnames:
        return []
    ws = wb["멤버"]
    headers = [str(c.value).strip() if c.value else "" for c in ws[1]]
    idx = {h: i for i, h in enumerate(headers)}
    i_k = idx.get("카톡아이디", 1)
    i_n = idx.get("실제이름", 2)
    i_g = idx.get("성별")
    i_e = idx.get("구력")
    out = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or len(row) <= max(i_k, i_n) or not row[i_n]:
            continue
        gender = str(row[i_g]).strip() if i_g is not None and row[i_g] else ""
        exp = None
        if i_e is not None and row[i_e] not in (None, ""):
            try:
                exp = int(row[i_e])
            except (ValueError, TypeError):
                exp = None
        out.append({
            "kakao": str(row[i_k]).strip() if row[i_k] else "",
            "name": str(row[i_n]).strip(),
            "gender": gender if gender in ("남", "여") else "",
            "exp": exp,
        })
    return out


def parse_member_settings(source) -> list:
    """멤버 설정 엑셀의 '부부' 시트에서 부부 페어를 읽는다 → [[이름1, 이름2, want], ...].

    부부 시트가 없거나 비어 있으면 빈 목록. 셀 값이 '원함/희망/허용/O'면 want=True,
    그 외('피함' 포함, 빈칸)는 False. 이름은 실제이름 기준(참가자 시트와 동일해야 매칭).
    """
    wb = load_workbook(source, data_only=True)
    if "부부" not in wb.sheetnames:
        return []
    ws = wb["부부"]
    headers = [str(c.value).strip() if c.value else "" for c in ws[1]]
    try:
        i1 = headers.index("이름1")
        i2 = headers.index("이름2")
        i3 = headers.index("부부페어") if "부부페어" in headers else 2
    except ValueError:
        i1, i2, i3 = 0, 1, 2
    couples = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or len(row) <= max(i1, i2):
            continue
        a, b = row[i1], row[i2]
        if not a or not b:
            continue
        raw = str(row[i3]).strip() if len(row) > i3 and row[i3] not in (None, "") else ""
        want = raw in _COUPLE_WANT_WORDS
        couples.append([str(a).strip(), str(b).strip(), want])
    return couples


def hhmm_to_min(s: str) -> int:
    s = str(s).strip()
    if not s or ":" not in s:
        raise ValueError(f"시간 형식 오류 (HH:MM 필요): '{s}'")
    h, m = s.split(":")
    h, m = int(h), int(m)
    if not (0 <= h <= 23) or m not in (0, 30):
        raise ValueError(f"시간은 0~23시, 분은 0 또는 30이어야 합니다: '{s}'")
    return h * 60 + m


def min_to_hhmm(v: int) -> str:
    return f"{v // 60:02d}:{v % 60:02d}"


def parse_players(ws) -> tuple[list[dict], list[str]]:
    headers = [str(c.value).strip() if c.value else "" for c in ws[1]]
    required = ["이름", "성별", "구력", "구분", "IN시간", "OUT시간"]
    for r in required:
        if r not in headers:
            raise ValueError(f"참가자 시트에 필수 컬럼 '{r}'이(가) 없습니다.")
    col_idx = {h: i for i, h in enumerate(headers)}

    players, errors, names_seen = [], [], set()
    pid = 1
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or all(v in (None, "") for v in row):
            continue
        name = row[col_idx["이름"]]
        if not name:
            continue
        name = str(name).strip()
        if not name:
            continue
        if name in names_seen:
            errors.append(f"참가자 이름 중복: '{name}'")
            continue
        names_seen.add(name)

        try:
            gender = str(row[col_idx["성별"]]).strip()
            if gender not in ("남", "여"):
                raise ValueError(f"'{name}': 성별은 '남' 또는 '여'여야 합니다 (현재: '{gender}')")

            exp_raw = row[col_idx["구력"]]
            if exp_raw is None or exp_raw == "":
                raise ValueError(f"'{name}': 구력이 비어있습니다")
            exp = int(exp_raw)
            if exp < 0:
                raise ValueError(f"'{name}': 구력은 0 이상이어야 합니다")

            membership = str(row[col_idx["구분"]]).strip()
            if membership not in ("정회원", "게스트"):
                raise ValueError(f"'{name}': 구분은 '정회원' 또는 '게스트'여야 합니다 (현재: '{membership}')")

            # 클럽 (선택). 비우거나 컬럼이 없으면 '우리클럽'으로 간주 → 평소엔 영향 없음.
            # 교류전 때 둘 이상의 클럽명이 들어오면 알고리즘이 같은 클럽끼리만 팀을 만든다.
            club = "우리클럽"
            if "클럽" in col_idx:
                raw_club = row[col_idx["클럽"]]
                if raw_club not in (None, ""):
                    club = str(raw_club).strip() or "우리클럽"

            in_min = hhmm_to_min(row[col_idx["IN시간"]])
            out_min = hhmm_to_min(row[col_idx["OUT시간"]])
            if in_min >= out_min:
                raise ValueError(f"'{name}': IN시간({min_to_hhmm(in_min)})이 OUT시간({min_to_hhmm(out_min)})보다 같거나 늦습니다")

            max_games = None
            if "최대게임수" in col_idx:
                raw = row[col_idx["최대게임수"]]
                if raw not in (None, ""):
                    try:
                        max_games = int(raw)
                    except (ValueError, TypeError):
                        raise ValueError(f"'{name}': 최대게임수는 정수여야 합니다 (현재: '{raw}')")
                    if max_games < 1:
                        raise ValueError(f"'{name}': 최대게임수는 1 이상이어야 합니다")

            min_games = None
            if "최소게임수" in col_idx:
                raw = row[col_idx["최소게임수"]]
                if raw not in (None, ""):
                    try:
                        min_games = int(raw)
                    except (ValueError, TypeError):
                        raise ValueError(f"'{name}': 최소게임수는 정수여야 합니다 (현재: '{raw}')")
                    if min_games < 1:
                        raise ValueError(f"'{name}': 최소게임수는 1 이상이어야 합니다")
            if min_games is not None and max_games is not None and min_games > max_games:
                raise ValueError(f"'{name}': 최소게임수({min_games})가 최대게임수({max_games})보다 큽니다")
        except (ValueError, TypeError) as e:
            errors.append(str(e))
            continue

        players.append({
            "id": f"P{pid:02d}",
            "name": name,
            "gender": "M" if gender == "남" else "F",
            "exp": exp,
            "membership": membership,
            "club": club,
            "in_min": in_min,
            "out_min": out_min,
            "max_games": max_games,
            "min_games": min_games,
        })
        pid += 1

    return players, errors


def parse_courts(ws) -> tuple[list[dict], list[str]]:
    headers = [str(c.value).strip() if c.value else "" for c in ws[1]]
    required = ["코트명", "시작시간", "종료시간"]
    for r in required:
        if r not in headers:
            raise ValueError(f"코트 시트에 필수 컬럼 '{r}'이(가) 없습니다.")
    col_idx = {h: i for i, h in enumerate(headers)}

    courts, errors, names_seen = [], [], set()
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or all(v in (None, "") for v in row):
            continue
        name = row[col_idx["코트명"]]
        if not name:
            continue
        name = str(name).strip()
        if not name:
            continue
        if name in names_seen:
            errors.append(f"코트명 중복: '{name}'")
            continue
        names_seen.add(name)

        try:
            start_min = hhmm_to_min(row[col_idx["시작시간"]])
            end_min = hhmm_to_min(row[col_idx["종료시간"]])
            if start_min >= end_min:
                raise ValueError(f"코트 '{name}': 시작시간이 종료시간보다 같거나 늦습니다")
        except (ValueError, TypeError) as e:
            errors.append(str(e))
            continue

        slots = list(range(start_min, end_min, 30))
        courts.append({
            "name": name,
            "start_min": start_min,
            "end_min": end_min,
            "slots": slots,
        })

    return courts, errors


def build_schedule_slots(courts: list[dict]) -> list[dict]:
    slot_set = set()
    for c in courts:
        for s in c["slots"]:
            slot_set.add(s)
    sorted_slots = sorted(slot_set)
    schedule = []
    for s in sorted_slots:
        available_courts = [c["name"] for c in courts if s in c["slots"]]
        schedule.append({
            "slot_start": s,
            "slot_end": s + 30,
            "courts": available_courts,
        })
    return schedule


def attach_available_slots(players: list[dict], schedule_slots: list[dict]) -> None:
    for p in players:
        avail = []
        for sl in schedule_slots:
            if p["in_min"] <= sl["slot_start"] and p["out_min"] >= sl["slot_end"]:
                avail.append(sl["slot_start"])
        p["available_slots"] = avail


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--in", dest="inp", required=True)
    ap.add_argument("--out", required=True)
    ap.add_argument("--members", default="",
                    help="멤버 설정 엑셀(부부 시트) 경로. 생략 시 내장 기본값 사용.")
    ap.add_argument("--no-members", action="store_true",
                    help="부부 페어 규칙을 끔 (기본값도 사용 안 함)")
    args = ap.parse_args()

    if not os.path.exists(args.inp):
        print(f"[에러] 입력 파일을 찾을 수 없음: {args.inp}", file=sys.stderr)
        sys.exit(1)

    wb = load_workbook(args.inp, data_only=True)

    if "참가자" not in wb.sheetnames:
        print("[에러] '참가자' 시트가 없습니다.", file=sys.stderr)
        sys.exit(1)
    if "코트" not in wb.sheetnames:
        print("[에러] '코트' 시트가 없습니다.", file=sys.stderr)
        sys.exit(1)

    all_errors = []
    try:
        players, perr = parse_players(wb["참가자"])
        all_errors.extend(perr)
    except ValueError as e:
        print(f"[에러] {e}", file=sys.stderr)
        sys.exit(1)

    try:
        courts, cerr = parse_courts(wb["코트"])
        all_errors.extend(cerr)
    except ValueError as e:
        print(f"[에러] {e}", file=sys.stderr)
        sys.exit(1)

    if all_errors:
        for e in all_errors:
            print(f"[에러] {e}", file=sys.stderr)
        sys.exit(1)

    if len(players) < 4:
        print(f"[에러] 참가자가 {len(players)}명입니다. 최소 4명 필요.", file=sys.stderr)
        sys.exit(1)
    if not courts:
        print("[에러] 사용 가능한 코트가 없습니다.", file=sys.stderr)
        sys.exit(1)

    schedule_slots = build_schedule_slots(courts)
    attach_available_slots(players, schedule_slots)

    # 경고 수집
    warnings = []
    males = [p for p in players if p["gender"] == "M"]
    females = [p for p in players if p["gender"] == "F"]
    if len(males) < 4:
        warnings.append(f"남자가 {len(males)}명이라 남자복식 불가. 혼복만 가능.")
    if len(females) < 4:
        warnings.append(f"여자가 {len(females)}명이라 여자복식 불가. 혼복만 가능.")

    for sl in schedule_slots:
        avail = [p for p in players if sl["slot_start"] in p["available_slots"]]
        if len(avail) < 4:
            t = f"{min_to_hhmm(sl['slot_start'])}~{min_to_hhmm(sl['slot_end'])}"
            warnings.append(f"슬롯 {t}: 가용 인원 {len(avail)}명 — 일부 코트 공석 가능")

    for p in players:
        if len(p["available_slots"]) == 0:
            warnings.append(f"'{p['name']}': 가용 슬롯 없음 — 코트 운영 시간과 IN/OUT 범위 확인 필요")
        if p["max_games"] is not None and p["max_games"] > len(p["available_slots"]):
            warnings.append(f"'{p['name']}': 최대게임수({p['max_games']}) > 가용 슬롯수({len(p['available_slots'])}) — 가용 슬롯 한도로 자연 제한됨")
        if p.get("min_games") and p["min_games"] > len(p["available_slots"]):
            warnings.append(f"'{p['name']}': 최소게임수({p['min_games']}) > 가용 슬롯수({len(p['available_slots'])}) — 가용 슬롯 수까지만 보장됨")

    # 최소게임수 합계가 전체 자리(코트×슬롯×4)보다 많으면 다 지킬 수 없다 — 미리 알림
    total_seats = 0
    for sl in schedule_slots:
        n_avail = sum(1 for p in players if sl["slot_start"] in p["available_slots"])
        total_seats += min(len(sl["courts"]), n_avail // 4) * 4
    min_sum = sum(min(p["min_games"], len(p["available_slots"])) for p in players if p.get("min_games"))
    if min_sum > total_seats:
        warnings.append(f"최소게임수 합계({min_sum})가 전체 배정 가능 자리({total_seats})를 넘습니다 — 일부는 보장 못 할 수 있음")

    if warnings:
        for w in warnings:
            print(f"[경고] {w}", file=sys.stderr)

    # 부부 페어 설정: 파일 지정 > 내장 기본값 > (--no-members) 없음
    couples: list = []
    if not args.no_members:
        if args.members and os.path.exists(args.members):
            try:
                couples = parse_member_settings(args.members)
                print(f"[안내] 멤버 설정 반영: {args.members} — 부부 {len(couples)}쌍"
                      f" (부부페어 원함 {sum(1 for c in couples if c[2])}쌍)", file=sys.stderr)
            except Exception as e:
                print(f"[경고] 멤버 설정 파일을 읽지 못해 기본값을 사용합니다: {e}", file=sys.stderr)
                couples = [list(c) for c in COUPLES_DEFAULT]
        else:
            if args.members:
                print(f"[경고] 멤버 설정 파일 없음({args.members}) — 내장 기본값 사용", file=sys.stderr)
            couples = [list(c) for c in COUPLES_DEFAULT]
        names = {p["name"] for p in players}
        present = [c for c in couples if c[0] in names and c[1] in names]
        if present:
            print(f"[안내] 이번 주 참가자 중 부부 {len(present)}쌍: "
                  + ", ".join(f"{a}·{b}({'원함' if w else '피함'})" for a, b, w in present),
                  file=sys.stderr)

    result = {
        "courts": courts,
        "players": players,
        "schedule_slots": schedule_slots,
        "couples": couples,
        "warnings": warnings,
    }
    os.makedirs(os.path.dirname(os.path.abspath(args.out)), exist_ok=True)
    with open(args.out, "w", encoding="utf-8") as f:
        json.dump(result, f, ensure_ascii=False, indent=2)
    print(f"[OK] 파싱 완료: {args.out}  (참가자 {len(players)}명, 코트 {len(courts)}개, 슬롯 {len(schedule_slots)}개)")


if __name__ == "__main__":
    main()
