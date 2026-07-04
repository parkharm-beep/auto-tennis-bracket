"""과거(지난주/2주전) 대진표 xlsx에서 '같은 팀 페어'를 뽑아 회피용 가중치 맵을 만든다.

핵심 아이디어
- 참가자 id(P01, P02…)는 매주 입력 순서대로 새로 매겨져서 주 사이에 안정적이지 않다.
  따라서 과거 페어는 반드시 **이름**으로 식별한다.
- 사용자는 결과 엑셀을 손으로 고치기도 하므로(게임수를 COUNTIF로 만든 이유), 실제로 함께 뛴
  페어의 가장 정확한 근거는 중간 JSON이 아니라 **결과 엑셀 그리드 자체**다. 그래서 렌더된
  대진표 xlsx의 셀을 직접 읽어 페어를 복원한다.
- 우선순위: 1주전 > 2주전. 최근 파일일수록 큰 가중치를 준다.

사용(단독 CLI):
    python history.py --prev1 지난주.xlsx --prev2 2주전.xlsx --out 00_history.json
"""
from __future__ import annotations

import argparse
import json
import os
import sys

from openpyxl import load_workbook

# 우선순위 기본 가중치 (1주전 > 2주전). schedule.py의 W["history_pair_repeat"]와 곱해져 최종 비용이 된다.
DEFAULT_W1 = 1.0   # 1주전
DEFAULT_W2 = 0.5   # 2주전


def _clean_name(v) -> str:
    """셀 값에서 표시용 장식을 제거해 순수 이름만 남긴다. ('홍길동(G)' → '홍길동')"""
    if v is None:
        return ""
    s = str(v).strip()
    if not s:
        return ""
    # 게스트 표기 (G) 접미사 제거
    if s.endswith("(G)"):
        s = s[:-3].strip()
    return s


def pair_key(a: str, b: str) -> tuple[str, str]:
    return (a, b) if a <= b else (b, a)


def extract_pairs_from_bracket_xlsx(source) -> list[tuple[str, str]]:
    """렌더된 대진표 xlsx(경로 또는 파일류 객체)에서 같은 팀 페어(이름쌍) 목록을 반환.

    렌더 레이아웃(render_bracket.py)에 맞춰 그리드를 읽는다:
      - 헤더 행: 'N번코트' 텍스트가 있는 셀의 열이 각 코트 블록의 시작(base) 열.
      - 각 코트 블록 5열: base+0=t1a, base+1=t1b, base+2=VS/혼, base+3=t2a, base+4=t2b.
      - 데이터는 헤더 다음 행부터 슬롯당 2행(성함/결과), 성함은 홀수 간격(2행 간격)의 첫 행.
    한 팀의 두 이름이 모두 있으면 페어로 본다. (상대 팀과의 조합은 페어가 아니므로 무시)
    같은 파일에서 같은 페어가 여러 번 나오면 그대로 여러 번 담는다(반복 강도 반영).
    """
    wb = load_workbook(source, data_only=True)
    ws = wb["대진표"] if "대진표" in wb.sheetnames else wb.active

    # 1) 헤더 행과 코트 base 열 찾기
    header_row = None
    bases: list[int] = []
    for r in range(1, min(ws.max_row, 8) + 1):
        cols = []
        for c in range(1, ws.max_column + 1):
            v = ws.cell(row=r, column=c).value
            if v and "번코트" in str(v):
                cols.append(c)
        if cols:
            header_row = r
            bases = cols
            break
    if header_row is None or not bases:
        return []  # 인식 불가 — 빈 결과 (호출부에서 안전 처리)

    data_start = header_row + 1

    pairs: list[tuple[str, str]] = []
    # 성함 행: data_start 부터 2행 간격. 그리드를 벗어난 행은 base 열이 비어 페어가 안 생긴다.
    row = data_start
    while row <= ws.max_row:
        for base in bases:
            t1a = _clean_name(ws.cell(row=row, column=base + 0).value)
            t1b = _clean_name(ws.cell(row=row, column=base + 1).value)
            t2a = _clean_name(ws.cell(row=row, column=base + 3).value)
            t2b = _clean_name(ws.cell(row=row, column=base + 4).value)
            if t1a and t1b and t1a != t1b:
                pairs.append(pair_key(t1a, t1b))
            if t2a and t2b and t2a != t2b:
                pairs.append(pair_key(t2a, t2b))
        row += 2
    return pairs


def build_history(prev_specs: list[tuple[object, float]]) -> dict:
    """prev_specs = [(source, weight), ...] (우선순위 높은=가중치 큰 것 먼저).

    반환: {"pairs": [[a, b, w], ...], "sources": [{"weight": w, "pairs": n}, ...]}
    같은 페어가 여러 파일/여러 번 나오면 가중치를 합산한다.
    """
    weighted: dict[tuple[str, str], float] = {}
    sources_meta = []
    for source, weight in prev_specs:
        try:
            found = extract_pairs_from_bracket_xlsx(source)
        except Exception as e:  # 손상/양식불일치 파일은 건너뛰되 진행은 계속
            sources_meta.append({"weight": weight, "pairs": 0, "error": str(e)})
            continue
        for k in found:
            weighted[k] = weighted.get(k, 0.0) + weight
        sources_meta.append({"weight": weight, "pairs": len(found)})

    pairs = [[a, b, round(w, 4)] for (a, b), w in sorted(weighted.items())]
    return {"pairs": pairs, "sources": sources_meta}


def main():
    ap = argparse.ArgumentParser(description="이전 대진표 xlsx에서 회피용 페어 히스토리 생성")
    ap.add_argument("--prev1", help="1주전 대진표 xlsx (우선순위 높음)")
    ap.add_argument("--prev2", help="2주전 대진표 xlsx")
    ap.add_argument("--w1", type=float, default=DEFAULT_W1)
    ap.add_argument("--w2", type=float, default=DEFAULT_W2)
    ap.add_argument("--out", required=True)
    args = ap.parse_args()

    specs: list[tuple[object, float]] = []
    if args.prev1 and os.path.exists(args.prev1):
        specs.append((args.prev1, args.w1))
    elif args.prev1:
        print(f"[경고] 1주전 파일 없음: {args.prev1}", file=sys.stderr)
    if args.prev2 and os.path.exists(args.prev2):
        specs.append((args.prev2, args.w2))
    elif args.prev2:
        print(f"[경고] 2주전 파일 없음: {args.prev2}", file=sys.stderr)

    history = build_history(specs)

    os.makedirs(os.path.dirname(os.path.abspath(args.out)), exist_ok=True)
    with open(args.out, "w", encoding="utf-8") as f:
        json.dump(history, f, ensure_ascii=False, indent=2)

    n_pairs = len(history["pairs"])
    print(f"[OK] 히스토리 생성: {args.out}  (회피 대상 페어 {n_pairs}쌍, 소스 {len(specs)}개)")
    for i, meta in enumerate(history["sources"], 1):
        tag = "1주전" if i == 1 else ("2주전" if i == 2 else f"소스{i}")
        if meta.get("error"):
            print(f"  - {tag}: 읽기 실패 ({meta['error']})")
        else:
            print(f"  - {tag}: 페어 {meta['pairs']}건 (가중치 {meta['weight']})")


if __name__ == "__main__":
    main()
