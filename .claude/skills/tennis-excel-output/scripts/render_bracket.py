"""Render a finalized tennis bracket to an Excel file matching the handwritten layout.

Usage:
    python render_bracket.py --parsed <01_parsed.json> --bracket <02_bracket.json>
                             --out <bracket.xlsx> [--date "26.5.30"] [--title "..."]
"""
from __future__ import annotations

import argparse
import json
import os
from collections import defaultdict
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


COURT_COLORS = [
    "FFFF99",  # yellow
    "FFCC99",  # orange
    "99CCFF",  # sky blue
    "C5E0B4",  # light green
    "FFB6C1",  # pink
    "D5A6E0",  # light purple
]
EMPTY_FILL = PatternFill("solid", fgColor="E0E0E0")
HEADER_FILL = PatternFill("solid", fgColor="BDD7EE")
TITLE_FILL = PatternFill("solid", fgColor="FFFFFF")
VS_FILL = PatternFill("solid", fgColor="F2F2F2")
SUMMARY_NUM_FILL = PatternFill("solid", fgColor="FFF2CC")
SUMMARY_GAME_FILL = PatternFill("solid", fgColor="FFFF99")

THIN = Side(border_style="thin", color="333333")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)

FONT_TITLE = Font(name="맑은 고딕", size=18, bold=True)
FONT_HEADER = Font(name="맑은 고딕", size=11, bold=True)
FONT_NAME = Font(name="맑은 고딕", size=11, bold=True)
FONT_TIME = Font(name="맑은 고딕", size=10)
FONT_SMALL = Font(name="맑은 고딕", size=9, color="666666")
FONT_VS = Font(name="맑은 고딕", size=10, bold=True, color="999999")


def min_to_hhmm(v: int) -> str:
    return f"{v // 60:02d}:{v % 60:02d}"


def display_name(p_stat: dict) -> str:
    n = p_stat["name"]
    if p_stat["membership"] == "게스트":
        return f"{n}(G)"
    return n


def render(parsed: dict, bracket: dict, out_path: str, date_str: str, title: str) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "대진표"

    courts = parsed["courts"]
    schedule_slots = parsed["schedule_slots"]
    matches = bracket["matches"]
    player_stats = bracket["player_stats"]

    clubs_present = {s.get("club", "") for s in player_stats if s.get("club", "")}
    is_exchange = len(clubs_present) > 1

    def _club_order_key(c):
        # 정회원이 많은 클럽(=홈)을 왼쪽으로
        members = sum(1 for s in player_stats
                      if s.get("club", "") == c and s.get("membership") == "정회원")
        return (-members, c)
    clubs_ordered = sorted(clubs_present, key=_club_order_key)
    two_club_mode = is_exchange and len(clubs_present) == 2

    matches_by_slot_court = defaultdict(dict)
    for m in matches:
        matches_by_slot_court[m["slot_start"]][m["court"]] = m

    # 컬럼 레이아웃
    # 1: 구분(번호+시간), 2: 성함/결과 라벨
    # 코트별 4컬럼: t1a, t1b, VS, t2a, t2b (= 5컬럼)
    # 우측 패널: 번호 | 이름 | 게임수 | (간격) | 번호 | 이름 | 게임수
    LABEL_COL_START = 1
    LABEL_COL_END = 2
    COURTS_COL_START = 3
    cols_per_court = 5
    courts_col_end = COURTS_COL_START + cols_per_court * len(courts) - 1

    summary_left_start = courts_col_end + 2
    SUM_NUM_W, SUM_NAME_W, SUM_GAME_W = 4, 12, 7
    summary_col_count = 3 + 1 + 3
    summary_right_end = summary_left_start + summary_col_count - 1

    # 컬럼 너비
    ws.column_dimensions[get_column_letter(LABEL_COL_START)].width = 9
    ws.column_dimensions[get_column_letter(LABEL_COL_END)].width = 7
    for i, _ in enumerate(courts):
        base = COURTS_COL_START + i * cols_per_court
        ws.column_dimensions[get_column_letter(base)].width = 10
        ws.column_dimensions[get_column_letter(base + 1)].width = 10
        ws.column_dimensions[get_column_letter(base + 2)].width = 4
        ws.column_dimensions[get_column_letter(base + 3)].width = 10
        ws.column_dimensions[get_column_letter(base + 4)].width = 10
    ws.column_dimensions[get_column_letter(courts_col_end + 1)].width = 1
    for off, w in zip(range(7), [SUM_NUM_W, SUM_NAME_W, SUM_GAME_W, 1, SUM_NUM_W, SUM_NAME_W, SUM_GAME_W]):
        ws.column_dimensions[get_column_letter(summary_left_start + off)].width = w

    # 행 1: 타이틀
    ws.cell(row=1, column=LABEL_COL_START, value=title)
    ws.merge_cells(start_row=1, start_column=LABEL_COL_START,
                   end_row=1, end_column=courts_col_end)
    ws.cell(row=1, column=LABEL_COL_START).font = FONT_TITLE
    ws.cell(row=1, column=LABEL_COL_START).alignment = CENTER
    ws.row_dimensions[1].height = 30

    # 날짜 (타이틀 영역 우측)
    ws.cell(row=1, column=summary_left_start, value=date_str)
    ws.merge_cells(start_row=1, start_column=summary_left_start,
                   end_row=1, end_column=summary_right_end)
    ws.cell(row=1, column=summary_left_start).font = FONT_HEADER
    ws.cell(row=1, column=summary_left_start).alignment = Alignment(horizontal="right", vertical="center")

    # 행 2-3: 헤더
    HEADER_ROW = 2
    ws.cell(row=HEADER_ROW, column=LABEL_COL_START, value="구분")
    ws.merge_cells(start_row=HEADER_ROW, start_column=LABEL_COL_START,
                   end_row=HEADER_ROW, end_column=LABEL_COL_END)
    c = ws.cell(row=HEADER_ROW, column=LABEL_COL_START)
    c.font = FONT_HEADER
    c.fill = HEADER_FILL
    c.alignment = CENTER
    c.border = BORDER

    for i, court in enumerate(courts):
        base = COURTS_COL_START + i * cols_per_court
        ws.cell(row=HEADER_ROW, column=base, value=f"{court['name']}번코트")
        ws.merge_cells(start_row=HEADER_ROW, start_column=base,
                       end_row=HEADER_ROW, end_column=base + cols_per_court - 1)
        hc = ws.cell(row=HEADER_ROW, column=base)
        hc.font = FONT_HEADER
        hc.fill = PatternFill("solid", fgColor=COURT_COLORS[i % len(COURT_COLORS)])
        hc.alignment = CENTER
        hc.border = BORDER

    # 우측 패널 헤더
    _left_hdr = clubs_ordered[0] if two_club_mode else "참가자"
    _right_hdr = clubs_ordered[1] if two_club_mode else "참가자"
    ws.cell(row=HEADER_ROW, column=summary_left_start, value="")
    ws.cell(row=HEADER_ROW, column=summary_left_start + 1, value=_left_hdr)
    ws.cell(row=HEADER_ROW, column=summary_left_start + 2, value="게임수")
    ws.cell(row=HEADER_ROW, column=summary_left_start + 4, value="")
    ws.cell(row=HEADER_ROW, column=summary_left_start + 5, value=_right_hdr)
    ws.cell(row=HEADER_ROW, column=summary_left_start + 6, value="게임수")
    for off in (0, 1, 2, 4, 5, 6):
        c = ws.cell(row=HEADER_ROW, column=summary_left_start + off)
        c.font = FONT_HEADER
        c.fill = HEADER_FILL
        c.alignment = CENTER
        c.border = BORDER

    # 데이터 행: 슬롯별로 2행 (성함, 결과) — 각 행 높이 22
    data_start_row = HEADER_ROW + 1
    for slot_idx, slot in enumerate(schedule_slots):
        name_row = data_start_row + slot_idx * 2
        result_row = name_row + 1
        ws.row_dimensions[name_row].height = 24
        ws.row_dimensions[result_row].height = 24

        ws.cell(row=name_row, column=LABEL_COL_START,
                value=f"{slot_idx + 1}번 게임\n{min_to_hhmm(slot['slot_start'])}\n~{min_to_hhmm(slot['slot_end'])}")
        ws.merge_cells(start_row=name_row, start_column=LABEL_COL_START,
                       end_row=result_row, end_column=LABEL_COL_START)
        c = ws.cell(row=name_row, column=LABEL_COL_START)
        c.font = FONT_TIME
        c.alignment = CENTER
        c.border = BORDER

        ws.cell(row=name_row, column=LABEL_COL_END, value="성함")
        ws.cell(row=result_row, column=LABEL_COL_END, value="결과")
        for r in (name_row, result_row):
            c = ws.cell(row=r, column=LABEL_COL_END)
            c.font = FONT_HEADER
            c.alignment = CENTER
            c.border = BORDER
            c.fill = HEADER_FILL

        for i, court in enumerate(courts):
            base = COURTS_COL_START + i * cols_per_court
            court_color = PatternFill("solid", fgColor=COURT_COLORS[i % len(COURT_COLORS)])

            for col_off in range(cols_per_court):
                for r in (name_row, result_row):
                    cell = ws.cell(row=r, column=base + col_off, value=None)
                    cell.border = BORDER
                    cell.alignment = CENTER

            court_active = slot["slot_start"] in court["slots"]
            if not court_active:
                continue

            m = matches_by_slot_court.get(slot["slot_start"], {}).get(court["name"])
            if m is None:
                continue

            stats_by_id = {s["id"]: s for s in player_stats}
            for k, p_id in enumerate(m["team1"]):
                cell = ws.cell(row=name_row, column=base + k, value=display_name(stats_by_id[p_id]))
                cell.font = FONT_NAME
                cell.fill = court_color
                cell.alignment = CENTER
                cell.border = BORDER
            vs_cell = ws.cell(row=name_row, column=base + 2, value="VS")
            vs_cell.font = FONT_VS
            vs_cell.fill = VS_FILL
            vs_cell.alignment = CENTER
            vs_cell.border = BORDER
            for k, p_id in enumerate(m["team2"]):
                cell = ws.cell(row=name_row, column=base + 3 + k, value=display_name(stats_by_id[p_id]))
                cell.font = FONT_NAME
                cell.fill = court_color
                cell.alignment = CENTER
                cell.border = BORDER

            if m["type"] == "X":
                vs_cell.value = "혼"

    # 우측 패널: 참가자 + 게임수
    # 게임수는 정적 숫자가 아니라 COUNTIF 수식 — 대진표 칸의 이름을 직접 고치면 자동 재계산된다.
    # 교류전(클럽 2개)일 때는 좌/우 열을 클럽별로 분리.
    panel_start = data_start_row
    last_grid_row = data_start_row + len(schedule_slots) * 2 - 1
    grid_abs = (f"${get_column_letter(COURTS_COL_START)}${data_start_row}:"
                f"${get_column_letter(courts_col_end)}${last_grid_row}")

    def write_entry(row, num_col, seq_num, p):
        c1 = ws.cell(row=row, column=num_col, value=seq_num)
        c1.font = FONT_HEADER
        c1.fill = SUMMARY_NUM_FILL
        c1.alignment = CENTER
        c1.border = BORDER
        ws.merge_cells(start_row=row, start_column=num_col, end_row=row + 1, end_column=num_col)

        name_cell = ws.cell(row=row, column=num_col + 1, value=display_name(p))
        name_cell.font = FONT_NAME
        name_cell.fill = SUMMARY_NUM_FILL
        name_cell.alignment = CENTER
        name_cell.border = BORDER
        info = f"{min_to_hhmm(p['in_min'])}~{min_to_hhmm(p['out_min'])}"
        if p.get("max_games") is not None:
            info += f" / 최대 {p['max_games']}게임"
        if is_exchange and p.get("club"):
            info += f" · {p['club']}"
        ci = ws.cell(row=row + 1, column=num_col + 1, value=info)
        ci.font = FONT_SMALL
        ci.alignment = CENTER
        ci.border = BORDER

        # 게임수 = 대진표 영역에서 이 사람 이름이 나오는 횟수 (수정 시 자동 갱신)
        name_ref = f"{get_column_letter(num_col + 1)}{row}"
        g = ws.cell(row=row, column=num_col + 2, value=f"=COUNTIF({grid_abs},{name_ref})")
        g.font = FONT_NAME
        g.fill = SUMMARY_GAME_FILL
        g.alignment = CENTER
        g.border = BORDER
        ws.merge_cells(start_row=row, start_column=num_col + 2, end_row=row + 1, end_column=num_col + 2)

    def write_total(row, game_col, first_row, last_row):
        rng = f"{get_column_letter(game_col)}{first_row}:{get_column_letter(game_col)}{last_row}"
        t = ws.cell(row=row, column=game_col, value=f"=SUM({rng})")
        t.font = FONT_HEADER
        t.fill = SUMMARY_GAME_FILL
        t.alignment = CENTER
        t.border = BORDER

    left_num_col = summary_left_start
    right_num_col = summary_left_start + 4

    if two_club_mode:
        # 좌=클럽1, 우=클럽2 로 분리
        col_lists = [
            (left_num_col, sorted([s for s in player_stats if s.get("club", "") == clubs_ordered[0]],
                                  key=lambda s: -s["available_slots"])),
            (right_num_col, sorted([s for s in player_stats if s.get("club", "") == clubs_ordered[1]],
                                   key=lambda s: -s["available_slots"])),
        ]
    else:
        stats_sorted = sorted(player_stats, key=lambda s: -s["available_slots"])
        half = (len(stats_sorted) + 1) // 2
        col_lists = [
            (left_num_col, stats_sorted[:half]),
            (right_num_col, stats_sorted[half:]),
        ]

    max_len = max((len(lst) for _, lst in col_lists), default=0)
    panel_end_row = panel_start + max_len * 2
    for num_col, lst in col_lists:
        for idx, p in enumerate(lst):
            write_entry(panel_start + idx * 2, num_col, idx + 1, p)
        if lst:
            write_total(panel_end_row, num_col + 2, panel_start, panel_end_row - 1)

    ws.sheet_view.showGridLines = False
    ws.print_options.horizontalCentered = True
    ws.page_setup.orientation = "landscape"

    # 두 번째 시트: 통계 요약
    ws2 = wb.create_sheet("통계")
    ws2.column_dimensions["A"].width = 22
    ws2.column_dimensions["B"].width = 18
    rows = [
        ("총 매치 수", len(matches)),
        ("남자복식", bracket["type_count"].get("M", 0)),
        ("여자복식", bracket["type_count"].get("F", 0)),
        ("혼합복식", bracket["type_count"].get("X", 0)),
        ("", ""),
        ("참가자 수", len(player_stats)),
        ("게임수 평균", round(sum(p["games"] for p in player_stats) / max(1, len(player_stats)), 2)),
        ("게임수 최대", max((p["games"] for p in player_stats), default=0)),
        ("게임수 최소", min((p["games"] for p in player_stats), default=0)),
    ]
    if is_exchange:
        rows.append(("", ""))
        rows.append(("교류전 클럽 수", len(clubs_present)))
        club_counts = defaultdict(int)
        for s in player_stats:
            club_counts[s.get("club", "")] += 1
        for cname in sorted(clubs_present):
            rows.append((f"  · {cname}", f"{club_counts[cname]}명"))
    for i, (k, v) in enumerate(rows, start=1):
        ws2.cell(row=i, column=1, value=k).font = FONT_HEADER
        ws2.cell(row=i, column=2, value=v).font = FONT_NAME

    wb.save(out_path)


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--parsed", required=True)
    ap.add_argument("--bracket", required=True)
    ap.add_argument("--out", required=True)
    ap.add_argument("--date", default="")
    ap.add_argument("--title", default="우리 테니스 클럽 대진표")
    args = ap.parse_args()

    with open(args.parsed, "r", encoding="utf-8") as f:
        parsed = json.load(f)
    with open(args.bracket, "r", encoding="utf-8") as f:
        bracket = json.load(f)

    render(parsed, bracket, args.out, args.date, args.title)
    print(f"[OK] 대진표 출력: {args.out}")


if __name__ == "__main__":
    main()
