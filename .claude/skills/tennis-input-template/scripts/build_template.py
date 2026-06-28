"""Build an empty tennis bracket input Excel template.

Usage:
    python build_template.py --out <output.xlsx>
"""
from __future__ import annotations

import argparse
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation


HEADER_FILL = PatternFill("solid", fgColor="D9E1F2")
GUIDE_FILL = PatternFill("solid", fgColor="FFF2CC")
HEADER_FONT = Font(name="맑은 고딕", size=11, bold=True)
BODY_FONT = Font(name="맑은 고딕", size=11)
THIN = Side(border_style="thin", color="888888")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
CENTER = Alignment(horizontal="center", vertical="center")
LEFT = Alignment(horizontal="left", vertical="center", wrap_text=True)


def _style_header(cell):
    cell.font = HEADER_FONT
    cell.fill = HEADER_FILL
    cell.alignment = CENTER
    cell.border = BORDER


def _style_body(cell):
    cell.font = BODY_FONT
    cell.alignment = CENTER
    cell.border = BORDER


PREFILL_FROM_IMAGE = [
    # 우리클럽 멤버 명단 (사전채움) — (이름, 성별, 구력, 구분, IN, OUT, 최대게임수, 메모)
    # 평소엔 우리클럽 내에서만 대진. 클럽 칸은 비워 둠(교류전 때만 입력).
    ("경성현", "남", 3, "정회원", "08:00", "12:00", 4, ""),
    ("김준학", "남", 10, "정회원", "08:00", "12:00", 4, ""),
    ("남궁석", "남", 10, "정회원", "08:00", "12:00", 4, ""),
    ("노남숙", "여", 3, "정회원", "08:00", "12:00", 4, ""),
    ("전혜선", "여", 7, "정회원", "08:00", "12:00", 4, ""),
    ("서종수", "남", 5, "정회원", "08:00", "12:00", 4, ""),
    ("원유철", "남", 4, "정회원", "08:00", "12:00", 4, ""),
    ("이강진", "남", 3, "정회원", "08:00", "12:00", 4, ""),
    ("이성돈", "남", 10, "정회원", "08:00", "12:00", 4, ""),
    ("이성수", "남", 5, "정회원", "08:00", "12:00", 4, ""),
    ("정진락", "남", 10, "정회원", "08:00", "12:00", 4, ""),
    ("최종인", "남", 10, "정회원", "10:00", "12:00", 3, ""),
    ("한병익", "남", 10, "정회원", "08:00", "12:00", 4, ""),
    ("김도윤", "남", 5, "정회원", "08:00", "12:00", 4, ""),
]

PREFILL_NOTE_FILL = PatternFill("solid", fgColor="FFF8DC")


def _build_players_sheet(ws, prefill: str = ""):
    headers = ["번호", "이름", "성별", "구력", "구분", "클럽", "IN시간", "OUT시간", "최대게임수", "메모"]
    widths = [6, 12, 8, 8, 12, 12, 10, 10, 12, 28]
    for col_idx, (title, w) in enumerate(zip(headers, widths), start=1):
        cell = ws.cell(row=1, column=col_idx, value=title)
        _style_header(cell)
        ws.column_dimensions[get_column_letter(col_idx)].width = w

    prefill_rows = PREFILL_FROM_IMAGE if prefill == "image" else []

    for r in range(2, 32):
        ws.cell(row=r, column=1, value=r - 1)
        for c in range(1, 11):
            _style_body(ws.cell(row=r, column=c))
        ws.cell(row=r, column=7).number_format = "@"  # IN시간
        ws.cell(row=r, column=8).number_format = "@"  # OUT시간

        idx = r - 2
        if idx < len(prefill_rows):
            name, gender, exp, mem, in_t, out_t, max_g, memo = prefill_rows[idx]
            if name: ws.cell(row=r, column=2, value=name)
            if gender: ws.cell(row=r, column=3, value=gender)
            if exp != "" and exp is not None: ws.cell(row=r, column=4, value=exp)
            if mem: ws.cell(row=r, column=5, value=mem)
            # 클럽(col 6)은 사전채움 비움 — 교류전 때만 직접 입력
            if in_t: ws.cell(row=r, column=7, value=in_t)
            if out_t: ws.cell(row=r, column=8, value=out_t)
            if max_g != "" and max_g is not None:
                ws.cell(row=r, column=9, value=max_g)
                ws.cell(row=r, column=9).fill = PREFILL_NOTE_FILL
            if memo:
                ws.cell(row=r, column=10, value=memo)
                ws.cell(row=r, column=10).fill = PREFILL_NOTE_FILL
            for c in (3, 4, 5, 6, 7, 8):
                if ws.cell(row=r, column=c).value in (None, ""):
                    ws.cell(row=r, column=c).fill = PREFILL_NOTE_FILL

    dv_gender = DataValidation(type="list", formula1='"남,여"', allow_blank=True)
    dv_gender.add("C2:C31")
    ws.add_data_validation(dv_gender)

    dv_member = DataValidation(type="list", formula1='"정회원,게스트"', allow_blank=True)
    dv_member.add("E2:E31")
    ws.add_data_validation(dv_member)

    ws.freeze_panes = "A2"


def _build_courts_sheet(ws):
    headers = ["코트명", "시작시간", "종료시간"]
    widths = [12, 12, 12]
    for col_idx, (title, w) in enumerate(zip(headers, widths), start=1):
        cell = ws.cell(row=1, column=col_idx, value=title)
        _style_header(cell)
        ws.column_dimensions[get_column_letter(col_idx)].width = w

    defaults = [
        ("A", "08:00", "12:00"),
        ("B", "08:00", "12:00"),
        ("C", "07:00", "09:00"),
    ]
    for r_idx, (name, start, end) in enumerate(defaults, start=2):
        ws.cell(row=r_idx, column=1, value=name)
        ws.cell(row=r_idx, column=2, value=start).number_format = "@"
        ws.cell(row=r_idx, column=3, value=end).number_format = "@"
        for c in range(1, 4):
            _style_body(ws.cell(row=r_idx, column=c))

    # 추가 빈 행
    for r in range(5, 12):
        for c in range(1, 4):
            _style_body(ws.cell(row=r, column=c))
            if c >= 2:
                ws.cell(row=r, column=c).number_format = "@"

    ws.freeze_panes = "A2"


def _build_guide_sheet(ws):
    ws.column_dimensions["A"].width = 100
    lines = [
        ("우리 테니스 클럽 대진표 — 입력 양식 작성 안내", True),
        ("", False),
        ("■ 1. 참가자 시트 작성법", True),
        ("• 번호: 자동 채워져 있음 (수정 불필요)", False),
        ("• 이름: 클럽 내 중복 없게 입력", False),
        ("• 성별: 드롭다운에서 '남' / '여' 선택 (필수)", False),
        ("• 구력: 테니스 경력 년수, 정수로 입력 — 예: 3, 10 (필수)", False),
        ("• 구분: '정회원' / '게스트' 선택 (필수)", False),
        ("• 클럽: 평소엔 비워두세요(자동으로 '우리클럽'). 교류전 때만 클럽명 입력 (선택)", False),
        ("    - 둘 이상의 클럽명이 들어오면 '교류전 모드'가 켜져 같은 클럽끼리만 한 팀이 됩니다", False),
        ("    - 예: 우리 회원은 '우리클럽', 방문팀은 '○○클럽' 식으로 구분", False),
        ("• IN시간 / OUT시간: HH:MM 형식, 30분 단위 (예: 08:00, 08:30) (필수)", False),
        ("    - IN은 코트장에 들어올 수 있는 가장 빠른 시각", False),
        ("    - OUT은 코트장을 떠나야 하는 시각", False),
        ("    - 반드시 IN < OUT, 둘 다 30분 단위 (00분 또는 30분)", False),
        ("• 최대게임수: 이 사람이 출전할 수 있는 최대 게임 수 (정수, 선택)", False),
        ("    - 비우면 무제한 (IN~OUT 범위 안에서 알고리즘이 자동 분배)", False),
        ("    - 예: '3게임만 하고 갈게요' → 3 입력", False),
        ("• 메모: 자유 기재 (선택, 알고리즘에 영향 없음)", False),
        ("", False),
        ("■ 2. 코트 시트 작성법", True),
        ("• 기본값: A코트(08:00-12:00), B코트(08:00-12:00), C코트(07:00-09:00)", False),
        ("• 코트가 더 있거나 운영시간이 다르면 행을 수정/추가하세요", False),
        ("• 시간은 30분 단위, 시작 < 종료", False),
        ("• 사용하지 않는 행은 빈 칸으로 두면 됩니다", False),
        ("", False),
        ("■ 3. 자동 배정 규칙 (알고리즘 동작 원리)", True),
        ("", False),
        ("[A] 기본 단위", True),
        ("• 1게임 = 30분", False),
        ("• 한 사람은 같은 시간 슬롯에 두 개 코트 동시 출전 불가", False),
        ("• 각자 본인 IN~OUT 범위 안의 슬롯에만 배정됨", False),
        ("", False),
        ("[B] 복식 종류 우선순위", True),
        ("• 남자복식, 여자복식이 우선 (가용 인원이 4명 이상일 때)", False),
        ("• 단성 복식이 안 되거나 인원이 너무 한쪽에 몰릴 때 혼합복식", False),
        ("• 혼합복식 규칙: 같은 팀의 남자 구력 ≥ 여자 구력", False),
        ("    (남자가 같은 팀 여자보다 경력이 같거나 더 많아야 함)", False),
        ("", False),
        ("[C] 코트별 우선 배정", True),
        ("• A코트: 여자복식 + 혼합복식 우선", False),
        ("• B코트: 남자복식 우선", False),
        ("• C코트: 무관 (균등 배분)", False),
        ("  ※ 단, 인원 부족 시 위 우선순위는 양보될 수 있음", False),
        ("", False),
        ("[D] 시간 제약", True),
        ("• 여성 참가자는 가급적 07:30 이후 슬롯에 배정", False),
        ("  (IN시간이 더 우선 — 7시 IN이면 7시부터 가능)", False),
        ("", False),
        ("[E] 팀 매칭 (재미와 균형)", True),
        ("• 두 팀의 합산 구력이 비슷하도록 매칭 (예: 5+7=12 vs 6+6=12)", False),
        ("• 한 번 같은 팀이었던 페어는 가능한 한 다시 같은 팀이 안 되게", False),
        ("• 한 팀에 정회원+게스트 혼합 약하게 권장 (강제 아님)", False),
        ("", False),
        ("[F] 게임수 균형", True),
        ("• 가용 시간이 비슷한 사람들끼리 게임수 격차 ±1~2 이내 목표", False),
        ("• 일찍 와서 늦게 가는 사람은 자연스럽게 더 많이 배정", False),
        ("• 늦게 오거나 일찍 가는 사람은 그만큼 적게 배정", False),
        ("• 전체 격차는 4 이하 권장 (가용 인원이 빠듯하면 더 클 수 있음)", False),
        ("", False),
        ("[G] 연속 출전 회피", True),
        ("• 한 사람이 3슬롯(1.5시간) 연속 출전은 가능한 한 회피", False),
        ("  (가용 인원이 코트수×4와 비슷하면 불가피하게 발생할 수 있음)", False),
        ("", False),
        ("[H] 개인별 최대 게임수", True),
        ("• 최대게임수 칸에 정수 입력 시 정확히 그 수 이하로 배정 (hard 제약)", False),
        ("• 빈 칸이면 IN~OUT 범위 안에서 균형 배정", False),
        ("", False),
        ("[I] 교류전 (클럽 대항)", True),
        ("• '클럽' 칸에 둘 이상의 클럽명이 있으면 교류전 모드 자동 작동", False),
        ("• 한 팀(복식 2명)은 반드시 같은 클럽끼리 구성 (hard 제약)", False),
        ("• 모든 경기는 반드시 상대 클럽과 (같은 클럽끼리는 절대 안 붙음, hard)", False),
        ("• 게임수 균형은 '각 클럽 내부'에서만 — 두 클럽 인원이 다르면 클럽 간 평균은 다를 수 있음", False),
        ("• 클럽 칸이 모두 비었거나 한 종류면 평소대로 동작 (영향 없음)", False),
        ("• 주의: 상대 클럽 인원이 부족한 슬롯에선 남는 사람이 쉬고 그 코트는 공석이 될 수 있음", False),
        ("", False),
        ("■ 4. 알고리즘이 회피 못 하는 입력 구조 (참고)", True),
        ("• 여자 인원이 4명 미만 → 여자복식 불가, 혼복만 가능", False),
        ("• 남자 인원이 4명 미만 → 남자복식 불가, 혼복만 가능", False),
        ("• 여자 최고 구력 > 남자 최고 구력 → 혼복 시 일부 규칙 위반 불가피", False),
        ("    (예: 김승회 10년 vs 남자 최고 8년)", False),
        ("• 가용 인원 = 코트수×4 → 휴식 자리 없어 연속 출전 발생", False),
        ("• 특정 슬롯에 가용 인원 4명 미만 → 그 코트는 자동 공석", False),
        ("", False),
        ("■ 5. 결과 생성 방법", True),
        ("• Windows 명령창(또는 PowerShell)에서:", False),
        ('     python 대진표_생성.py --date "26.5.30"', False),
        ("• 또는 대진표_생성.bat 더블클릭", False),
        ("• 결과 파일: 출력/테니스_대진표_<날짜>.xlsx", False),
        ("", False),
        ("■ 6. 다시 생성하고 싶을 때", True),
        ("• 같은 입력으로 다른 패턴 원하면 --seed 99 (또는 다른 숫자)", False),
        ("• 입력 양식 수정 후 다시 실행하면 새로 생성", False),
        ("", False),
        ("■ 7. 파일 위치", True),
        ("• 입력 양식: 입력/테니스_입력양식.xlsx", False),
        ("• 결과 파일: 출력/테니스_대진표_<YYMMDD>.xlsx", False),
        ("• 샘플/참고: 샘플/  (이미지·예시 데이터)", False),
    ]
    for i, (text, bold) in enumerate(lines, start=1):
        cell = ws.cell(row=i, column=1, value=text)
        cell.alignment = LEFT
        cell.font = Font(name="맑은 고딕", size=12 if bold else 11, bold=bold)
        if bold and text:
            cell.fill = GUIDE_FILL


def build_template(out_path: str, prefill: str = "") -> None:
    wb = Workbook()
    ws_players = wb.active
    ws_players.title = "참가자"
    _build_players_sheet(ws_players, prefill=prefill)

    ws_courts = wb.create_sheet("코트")
    _build_courts_sheet(ws_courts)

    ws_guide = wb.create_sheet("안내")
    _build_guide_sheet(ws_guide)

    wb.move_sheet(ws_guide, offset=-2)

    wb.save(out_path)


def main():
    p = argparse.ArgumentParser()
    p.add_argument("--out", required=True, help="출력 .xlsx 경로")
    p.add_argument("--prefill", default="", choices=["", "image"],
                   help="image: 우리클럽 멤버 명단으로 사전 채움")
    args = p.parse_args()
    build_template(args.out, prefill=args.prefill)
    msg = f"[OK] 입력 템플릿 생성: {args.out}"
    if args.prefill == "image":
        msg += "  (우리클럽 멤버 사전 채움 — 클럽 칸은 교류전 때만 입력)"
    print(msg)


if __name__ == "__main__":
    main()
