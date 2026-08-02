"""카카오 캘린더 일정(참석/불참/게스트)에서 대진표 입력 엑셀 초안을 생성.

주간 워크플로우 (Claude가 브라우저로 일정을 읽은 뒤):
    python build_draft.py --snapshot <일정 텍스트 파일> [--out 입력/테니스_입력양식.xlsx]

입력 텍스트는 gstack browse snapshot 출력이든 카톡 복사 텍스트든 아래 패턴만 있으면 된다:
    참석 <이름> [주최자]     ← 한 줄에 하나
    불참 <이름>
    ... 게스트 ...           ← 설명란: "권명숙님 게스트 7년 백용승님 게스트 금배" 식

규칙:
- 이름 매칭: 멤버 설정(카톡아이디↔실제이름)으로 변환. 이모지·공백 차이는 무시(명수기❤️=명수기).
- 게스트 구력: 'N년' 또는 '금배'(=20년). 성별은 명단에 없으므로 비워 두고 경고 —
  워크플로우에서 Claude가 이름으로 추정해 채우거나 사용자가 초안에서 수정.
- IN/OUT: 사전채움 로스터의 개인별 평소 시간. 없는 사람은 코트 운영 전체 시간.
- 최소/최대게임수: 사전채움 로스터 값 승계. 게스트는 기본 최대 4게임(--guest-max로 변경).
- 명단에 없는 참석자(탈퇴자 등)는 초안에서 제외하고 경고로 표시.
- 기존 출력 파일은 입력/…_백업_<날짜>.xlsx 로 백업 후 덮어씀(--no-backup으로 끔).
"""
from __future__ import annotations

import argparse
import datetime as _dt
import io
import os
import re
import shutil
import sys

from openpyxl import load_workbook

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
from parse_input import MEMBERS_DEFAULT, parse_member_roster  # noqa: E402
from build_template import PREFILL_FROM_IMAGE, build_template  # noqa: E402

KEUMBAE_EXP = 20   # '금배' = 구력 20년 이상으로 간주 (사용자 확정, 26.8.2)


def _norm(s: str) -> str:
    """이름 비교용 정규화 — 한글·영문·숫자만 남긴다 (이모지·공백·기호 제거)."""
    return "".join(ch for ch in str(s) if ch.isalnum()).lower()


def parse_schedule_text(text: str) -> dict:
    """일정 텍스트에서 참석/불참/게스트/날짜를 추출."""
    attend, absent = [], []
    for m in re.finditer(r"(참석|불참)\s+([^\r\n@\[\]]+)", text):
        name = m.group(2).strip()
        name = re.sub(r"\s*주최자\s*$", "", name).strip()
        if not name or name in ("여부", "목록", "응답"):   # '참석 여부'/'참석 목록' 같은 라벨 제외
            continue
        (attend if m.group(1) == "참석" else absent).append(name)

    guests = []
    for m in re.finditer(r"([가-힣A-Za-z]{2,10}?)\s*님?\s*게스트\s*(금배|\d+\s*년|\d+)?", text):
        name = re.sub(r"님$", "", m.group(1))   # '권명숙님' → '권명숙'
        raw = (m.group(2) or "").strip()
        if raw == "금배":
            exp = KEUMBAE_EXP
        elif raw:
            exp = int(re.sub(r"\D", "", raw))
        else:
            exp = None
        guests.append({"name": name, "exp": exp})

    date = None
    dm = re.search(r"(\d{4})\.\s*(\d{1,2})\.\s*(\d{1,2})", text)
    if dm:
        date = f"{int(dm.group(1)) % 100}.{int(dm.group(2))}.{int(dm.group(3))}"

    return {"attend": attend, "absent": absent, "guests": guests, "date": date}


def build_draft(snapshot_text: str, members_path: str = "", out_path: str = "",
                guest_max: int | None = 4, backup: bool = True,
                guest_genders: dict | None = None) -> dict:
    """일정 텍스트 → 입력 엑셀 초안. 요약 dict 반환."""
    parsed = parse_schedule_text(snapshot_text)

    # 멤버 명단 (성별·구력): 설정 파일 > 내장 기본값
    roster = []
    if members_path and os.path.exists(members_path):
        try:
            roster = parse_member_roster(members_path)
        except Exception:
            roster = []
    if not roster:
        roster = [{"kakao": k, "name": n, "gender": g, "exp": e} for k, n, g, e, _ in MEMBERS_DEFAULT]
    by_key = {}
    for mrow in roster:
        by_key[_norm(mrow["kakao"])] = mrow
        by_key[_norm(mrow["name"])] = mrow

    # 사전채움 로스터: 개인별 평소 IN/OUT·최소/최대게임수
    prefill = {name: dict(in_t=in_t, out_t=out_t, min_g=min_g, max_g=max_g)
               for (name, _g, _e, _m, in_t, out_t, min_g, max_g, _memo) in PREFILL_FROM_IMAGE}

    rows, unknown = [], []
    guest_names = {_norm(g["name"]) for g in parsed["guests"]}
    for disp in parsed["attend"]:
        m = by_key.get(_norm(disp))
        if m is None:
            if _norm(disp) not in guest_names:
                unknown.append(disp)
            continue
        pf = prefill.get(m["name"], {})
        rows.append({
            "name": m["name"], "gender": m["gender"], "exp": m["exp"], "mem": "정회원",
            "in_t": pf.get("in_t") or "07:00", "out_t": pf.get("out_t") or "12:00",
            "min_g": pf.get("min_g") or "", "max_g": pf.get("max_g") or "", "memo": "",
        })

    gg = guest_genders or {}
    for g in parsed["guests"]:
        rows.append({
            "name": g["name"], "gender": gg.get(g["name"], ""), "exp": g["exp"] if g["exp"] is not None else "",
            "mem": "게스트", "in_t": "07:00", "out_t": "12:00",
            "min_g": "", "max_g": guest_max if guest_max else "",
            "memo": "게스트" + (" · 구력 금배(20년 이상)" if g["exp"] == KEUMBAE_EXP else ""),
        })

    # 양식 골격 생성 후 참가자 시트를 초안 명단으로 교체
    buf = io.BytesIO()
    build_template(buf)
    buf.seek(0)
    wb = load_workbook(buf)
    ws = wb["참가자"]
    col = {c.value: i for i, c in enumerate(ws[1], 1)}
    for i, r in enumerate(rows, 2):
        ws.cell(row=i, column=col["이름"]).value = r["name"]
        ws.cell(row=i, column=col["성별"]).value = r["gender"] or None
        ws.cell(row=i, column=col["구력"]).value = r["exp"] if r["exp"] != "" else None
        ws.cell(row=i, column=col["구분"]).value = r["mem"]
        ws.cell(row=i, column=col["IN시간"]).value = r["in_t"]
        ws.cell(row=i, column=col["OUT시간"]).value = r["out_t"]
        ws.cell(row=i, column=col["최소게임수"]).value = r["min_g"] if r["min_g"] != "" else None
        ws.cell(row=i, column=col["최대게임수"]).value = r["max_g"] if r["max_g"] != "" else None
        ws.cell(row=i, column=col["메모"]).value = r["memo"] or None

    if not out_path:
        out_path = "입력/테니스_입력양식.xlsx"
    if backup and os.path.exists(out_path):
        stem, ext = os.path.splitext(out_path)
        bak = f"{stem}_백업_{_dt.date.today().strftime('%Y%m%d')}{ext}"
        shutil.copy2(out_path, bak)
    os.makedirs(os.path.dirname(os.path.abspath(out_path)), exist_ok=True)
    wb.save(out_path)

    n_f = sum(1 for r in rows if r["gender"] == "여")
    n_guest = sum(1 for r in rows if r["mem"] == "게스트")
    no_gender = [r["name"] for r in rows if not r["gender"]]
    return {
        "out": out_path, "date": parsed["date"], "total": len(rows),
        "members": len(rows) - n_guest, "guests": n_guest, "women": n_f,
        "absent": parsed["absent"], "unknown": unknown, "no_gender": no_gender,
    }


def main():
    ap = argparse.ArgumentParser(description="카카오 일정 텍스트 → 대진표 입력 엑셀 초안")
    ap.add_argument("--snapshot", required=True, help="일정 텍스트 파일 (browse snapshot 출력 또는 복사 텍스트)")
    ap.add_argument("--members", default="", help="멤버 설정 엑셀 (생략 시 내장 기본값)")
    ap.add_argument("--out", default="", help="초안 출력 경로 (기본: 입력/테니스_입력양식.xlsx)")
    ap.add_argument("--guest-max", type=int, default=4, help="게스트 기본 최대게임수 (0=없음)")
    ap.add_argument("--guest-gender", action="append", default=[],
                    help="게스트 성별 지정: 이름=남|여 (반복 가능)")
    ap.add_argument("--no-backup", action="store_true")
    args = ap.parse_args()

    with open(args.snapshot, "r", encoding="utf-8") as f:
        text = f.read()
    gg = {}
    for spec in args.guest_gender:
        if "=" in spec:
            k, v = spec.split("=", 1)
            gg[k.strip()] = v.strip()

    r = build_draft(text, members_path=args.members, out_path=args.out,
                    guest_max=args.guest_max or None, backup=not args.no_backup,
                    guest_genders=gg)

    print(f"[OK] 입력 엑셀 초안 생성: {r['out']}")
    print(f"  일정 날짜: {r['date'] or '(미검출)'}")
    print(f"  참가자 {r['total']}명 = 정회원 {r['members']} + 게스트 {r['guests']}  (여자 {r['women']}명)")
    if r["absent"]:
        print(f"  불참 {len(r['absent'])}명: {', '.join(r['absent'])}")
    if r["unknown"]:
        print(f"  [경고] 명단에 없는 참석자(초안 제외): {', '.join(r['unknown'])} — 새 멤버면 멤버 설정에 추가 필요")
    if r["no_gender"]:
        print(f"  [경고] 성별 미입력(초안에서 채워야 함): {', '.join(r['no_gender'])}")
    print("  다음 단계: python 대진표_생성.py --check  →  이상 없으면  python 대진표_생성.py --date \""
          + (r["date"] or "YY.M.D") + "\"")


if __name__ == "__main__":
    main()
