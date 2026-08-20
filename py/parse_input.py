"""Parse and validate a filled tennis bracket input Excel file.

Usage:
    python parse_input.py --in <input.xlsx> --out <parsed.json>
"""
from __future__ import annotations

import argparse
import json
import os
import re
import sys
from collections import defaultdict
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter


# ---------------------------------------------------------------------------
# 클럽 멤버 설정 기본값 — 멤버 설정 파일(클럽멤버_설정.xlsx)이 없을 때 사용.
# 26.8.2 카톡 투표 명단(25명, 사용자 확정본) 기준.
# ---------------------------------------------------------------------------
MEMBERS_DEFAULT = [
    # (카톡아이디, 실제이름, 성별, 구력, 메모)
    # 성별·구력은 26.8.2 사용자 확정본(테니스_입력양식_사전채움_25명.xlsx) 기준.
    ("김도윤",      "김도윤", "남", 10, ""),
    ("김효순",      "김효순", "남", 5,  ""),
    ("남궁석",      "남궁석", "남", 10, ""),
    ("노남숙",      "노남숙", "여", 3,  ""),
    ("명수기❤️",    "서명숙", "여", 5,  ""),
    ("민기준",      "민기준", "남", 5,  ""),
    ("박경수",      "박경수", "남", 5,  ""),
    ("박진우",      "박진우", "남", 7,  ""),
    ("서종수",      "서종수", "남", 5,  ""),
    ("성현",        "경성현", "남", 4,  ""),
    ("원유철",      "원유철", "남", 4,  ""),
    ("이강진",      "이강진", "남", 4,  ""),
    ("이성돈",      "이성돈", "남", 10, ""),
    ("이성수",      "이성수", "남", 7,  ""),
    ("이지은",      "이지은", "여", 3,  ""),
    ("임성훈",      "임성훈", "남", 5,  ""),
    ("정재동",      "정재동", "남", 10, ""),
    ("정진락",      "정진락", "남", 10, ""),
    ("정희",        "정정희", "여", 5,  ""),
    ("최종인",      "최종인", "남", 10, ""),
    ("한병익",      "한병익", "남", 10, ""),
    ("혜선",        "전혜선", "여", 7,  ""),
    ("HJ Shin",     "신혁재", "남", 5,  ""),
    ("Joonhak Kim", "김준학", "남", 10, ""),
    ("Mira",        "방미라", "여", 3,  ""),
]

# 부부 페어: (이름1, 이름2, 혼복에서 부부페어를 원하는지, 종료시간차)
# 원함=True → 혼복이 나올 때 부부가 같은 팀이 되는 것을 우대.
# 피함=False → 부부를 같은 팀으로 묶지 않게 회피(소프트).
# 종료시간차: None=같이 끝남(30분 이내 목표) / 30=마지막 경기 종료가 반드시 정확히 30분 차이
#   (신혁재·방미라 — 26.8.6 사용자 확정 '기본 반영 사항').
COUPLES_DEFAULT = [
    ("박경수", "서명숙", False, None),
    ("한병익", "전혜선", True,  None),
    ("신혁재", "방미라", False, 30),
    ("원유철", "이지은", False, None),
    # 게스트 부부 (26.8.6 '피함' 확정) — 설정 파일에만 있고 내장 기본값에 없어서
    # 멤버 설정을 안 올리면 웹에서 부부 규칙이 통째로 미적용이던 것을 26.8.14 등록.
    # 부부 규칙은 두 사람이 모두 참가자 명단에 있을 때만 발동하므로 상시 등록해도 무해.
    ("이선우", "김희진", False, None),
]

_COUPLE_WANT_WORDS = ("원함", "희망", "허용", "O", "o", "예", "y", "Y")


def parse_member_roster(source) -> list:
    """멤버 설정 엑셀의 '멤버' 시트에서 명단을 읽는다.

    반환: [{"kakao": 카톡아이디, "name": 실제이름, "gender": '남'/'여'/'', "exp": int|None}, ...]
    입력 엑셀 초안 자동 생성 시 참석자 이름(카톡아이디)을 실명·성별·구력으로 변환하는 근거.
    멤버 시트가 없으면 빈 목록(호출부에서 MEMBERS_DEFAULT 사용).
    """
    wb = load_workbook(source, data_only=True)
    if "멤버" not in wb.sheetnames:
        return []
    ws = wb["멤버"]
    headers = [str(c.value).strip() if c.value else "" for c in ws[1]]
    idx = {h: i for i, h in enumerate(headers)}
    i_k = idx.get("카톡아이디", 1)
    i_n = idx.get("실제이름", 2)
    i_g = idx.get("성별")
    i_e = idx.get("구력")
    out = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or len(row) <= max(i_k, i_n) or not row[i_n]:
            continue
        gender = str(row[i_g]).strip() if i_g is not None and row[i_g] else ""
        exp = None
        if i_e is not None and row[i_e] not in (None, ""):
            try:
                exp = int(row[i_e])
            except (ValueError, TypeError):
                exp = None
        out.append({
            "kakao": str(row[i_k]).strip() if row[i_k] else "",
            "name": str(row[i_n]).strip(),
            "gender": gender if gender in ("남", "여") else "",
            "exp": exp,
        })
    return out


def parse_member_settings(source) -> list:
    """멤버 설정 엑셀의 '부부' 시트에서 부부 페어를 읽는다 → [[이름1, 이름2, want, 종료시간차], ...].

    부부 시트가 없거나 비어 있으면 빈 목록. 셀 값이 '원함/희망/허용/O'면 want=True,
    그 외('피함' 포함, 빈칸)는 False. 이름은 실제이름 기준(참가자 시트와 동일해야 매칭).
    종료시간차: 빈칸/'같이'=같이 끝남(30분 이내 목표) / '30분차이'=반드시 정확히 30분 차이(→30).
    구양식(컬럼 없음)도 그대로 읽힌다(전부 기본값).
    """
    wb = load_workbook(source, data_only=True)
    if "부부" not in wb.sheetnames:
        return []
    ws = wb["부부"]
    headers = [str(c.value).strip() if c.value else "" for c in ws[1]]
    try:
        i1 = headers.index("이름1")
        i2 = headers.index("이름2")
        i3 = headers.index("부부페어") if "부부페어" in headers else 2
    except ValueError:
        i1, i2, i3 = 0, 1, 2
    i4 = headers.index("종료시간차") if "종료시간차" in headers else None
    couples = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or len(row) <= max(i1, i2):
            continue
        a, b = row[i1], row[i2]
        if not a or not b:
            continue
        raw = str(row[i3]).strip() if len(row) > i3 and row[i3] not in (None, "") else ""
        want = raw in _COUPLE_WANT_WORDS
        gap = None
        if i4 is not None and len(row) > i4 and row[i4] not in (None, ""):
            raw4 = str(row[i4]).replace(" ", "")
            if "30" in raw4 and "이내" not in raw4:   # '30분차이', '정확히30분' 등
                gap = 30
        couples.append([str(a).strip(), str(b).strip(), want, gap])
    return couples


def hhmm_to_min(s: str) -> int:
    s = str(s).strip()
    if not s or ":" not in s:
        raise ValueError(f"시간 형식 오류 (HH:MM 필요): '{s}'")
    h, m = s.split(":")
    h, m = int(h), int(m)
    if not (0 <= h <= 23) or m not in (0, 30):
        raise ValueError(f"시간은 0~23시, 분은 0 또는 30이어야 합니다: '{s}'")
    return h * 60 + m


# 개인 '연속게임' 칸 → 연속으로 뛸 수 있는 최대 게임수.
# 금지=1(한 게임 뛰면 반드시 쉼) / 빈칸=2(현행 기본) / 허용=3.
# '허용'은 무제한이 아니다 — 6슬롯에 5게임 같은 보장을 위해 3연속까지만 연다.
STREAK_RUN_LIMIT = {"": 2, "no2": 1, "ok3": 3}


def _max_games_streak(slots: list, streak: str = "") -> int:
    """개인 연속게임 모드를 지키며 이 슬롯들에서 뛸 수 있는 최대 게임수.

    모드는 '연속으로 몇 게임까지 뛸 수 있는가'(run limit)로 환원된다 —
    금지(no2)=1 / 빈칸(기본)=2 / 허용(ok3)=3.
    '허용'도 무제한이 아니라 **3연속까지**다(안내 문구와 동일한 의미).
    """
    limit = STREAK_RUN_LIMIT.get(streak, STREAK_RUN_LIMIT[""])
    # best[k] = 지금까지 중 '마지막 슬롯이 k연속째'인 상태의 최대 게임수 (k=0은 그 슬롯을 안 뜀)
    best = [0] + [None] * limit
    prev = None
    for s in sorted(slots):
        adj = prev is not None and s - prev == 30
        rest = max(x for x in best if x is not None)
        nxt = [rest] + [None] * limit
        for k in range(1, limit + 1):
            if k == 1:
                src = best[0] if adj else rest   # 인접이면 직전은 쉬었어야 새 연속이 시작된다
            else:
                src = best[k - 1] if adj else None
            if src is not None:
                nxt[k] = src + 1
        best, prev = nxt, s
    return max(x for x in best if x is not None)


def _max_games_no3(slots: list) -> int:
    """기존 호출부 호환용: 기본 3연속 금지 모드의 최대 게임수."""
    return _max_games_streak(slots, "")


def min_to_hhmm(v: int) -> str:
    return f"{v // 60:02d}:{v % 60:02d}"


def parse_players(ws) -> tuple[list[dict], list[str]]:
    headers = [str(c.value).strip() if c.value else "" for c in ws[1]]
    required = ["이름", "성별", "구력", "구분", "IN시간", "OUT시간"]
    for r in required:
        if r not in headers:
            raise ValueError(f"참가자 시트에 필수 컬럼 '{r}'이(가) 없습니다.")
    col_idx = {h: i for i, h in enumerate(headers)}

    players, errors, names_seen = [], [], set()
    pid = 1
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or all(v in (None, "") for v in row):
            continue
        name = row[col_idx["이름"]]
        if not name:
            continue
        name = str(name).strip()
        if not name:
            continue
        if name in names_seen:
            errors.append(f"참가자 이름 중복: '{name}'")
            continue
        names_seen.add(name)

        try:
            gender = str(row[col_idx["성별"]]).strip()
            if gender not in ("남", "여"):
                raise ValueError(f"'{name}': 성별은 '남' 또는 '여'여야 합니다 (현재: '{gender}')")

            exp_raw = row[col_idx["구력"]]
            if exp_raw is None or exp_raw == "":
                raise ValueError(f"'{name}': 구력이 비어있습니다")
            exp = int(exp_raw)
            if exp < 0:
                raise ValueError(f"'{name}': 구력은 0 이상이어야 합니다")

            membership = str(row[col_idx["구분"]]).strip()
            if membership not in ("정회원", "게스트"):
                raise ValueError(f"'{name}': 구분은 '정회원' 또는 '게스트'여야 합니다 (현재: '{membership}')")

            # 클럽 (선택). 비우거나 컬럼이 없으면 '우리클럽'으로 간주 → 평소엔 영향 없음.
            # 교류전 때 둘 이상의 클럽명이 들어오면 알고리즘이 같은 클럽끼리만 팀을 만든다.
            club = "우리클럽"
            if "클럽" in col_idx:
                raw_club = row[col_idx["클럽"]]
                if raw_club not in (None, ""):
                    club = str(raw_club).strip() or "우리클럽"

            in_min = hhmm_to_min(row[col_idx["IN시간"]])
            out_min = hhmm_to_min(row[col_idx["OUT시간"]])
            if in_min >= out_min:
                raise ValueError(f"'{name}': IN시간({min_to_hhmm(in_min)})이 OUT시간({min_to_hhmm(out_min)})보다 같거나 늦습니다")

            max_games = None
            if "최대게임수" in col_idx:
                raw = row[col_idx["최대게임수"]]
                if raw not in (None, ""):
                    try:
                        max_games = int(raw)
                    except (ValueError, TypeError):
                        raise ValueError(f"'{name}': 최대게임수는 정수여야 합니다 (현재: '{raw}')")
                    if max_games < 1:
                        raise ValueError(f"'{name}': 최대게임수는 1 이상이어야 합니다")

            min_games = None
            if "최소게임수" in col_idx:
                raw = row[col_idx["최소게임수"]]
                if raw not in (None, ""):
                    try:
                        min_games = int(raw)
                    except (ValueError, TypeError):
                        raise ValueError(f"'{name}': 최소게임수는 정수여야 합니다 (현재: '{raw}')")
                    if min_games < 1:
                        raise ValueError(f"'{name}': 최소게임수는 1 이상이어야 합니다")
            if min_games is not None and max_games is not None and min_games > max_games:
                raise ValueError(f"'{name}': 최소게임수({min_games})가 최대게임수({max_games})보다 큽니다")

            # 혼복희망 (선택): "그중 혼복으로 N게임 뛰고 싶다".
            # 비워 두면 지금까지와 완전히 동일 — 적은 사람이 있을 때만 혼복 하한이 열린다.
            mixed_wish = None
            if "혼복희망" in col_idx:
                raw = row[col_idx["혼복희망"]]
                if raw not in (None, ""):
                    try:
                        mixed_wish = int(raw)
                    except (ValueError, TypeError):
                        raise ValueError(f"'{name}': 혼복희망은 정수여야 합니다 (현재: '{raw}')")
                    if mixed_wish < 1:
                        raise ValueError(f"'{name}': 혼복희망은 1 이상이어야 합니다")
                    if max_games is not None and mixed_wish > max_games:
                        raise ValueError(
                            f"'{name}': 혼복희망({mixed_wish})이 최대게임수({max_games})보다 큽니다")

            streak = ""
            if "연속게임" in col_idx:
                raw = row[col_idx["연속게임"]]
                normalized = "".join(str(raw).split()) if raw is not None else ""
                if normalized:   # 공백만 든 셀(드롭다운을 스페이스로 지운 경우)은 빈칸 취급
                    if normalized in {"금지", "연속금지", "불가", "안함"}:
                        streak = "no2"
                    elif normalized in {"허용", "연속허용", "가능", "연속게임가능"}:
                        streak = "ok3"
                    else:
                        raise ValueError(
                            f"'{name}': 연속게임은 '금지' 또는 '허용'만 입력하세요 (현재: '{raw}')")
        except (ValueError, TypeError) as e:
            errors.append(str(e))
            continue

        players.append({
            "id": f"P{pid:02d}",
            "name": name,
            "gender": "M" if gender == "남" else "F",
            "exp": exp,
            "membership": membership,
            "club": club,
            "in_min": in_min,
            "out_min": out_min,
            "max_games": max_games,
            "min_games": min_games,
            "mixed_wish": mixed_wish,
            "streak": streak,
        })
        pid += 1

    return players, errors


def parse_courts(ws) -> tuple[list[dict], list[str]]:
    headers = [str(c.value).strip() if c.value else "" for c in ws[1]]
    required = ["코트명", "시작시간", "종료시간"]
    for r in required:
        if r not in headers:
            raise ValueError(f"코트 시트에 필수 컬럼 '{r}'이(가) 없습니다.")
    col_idx = {h: i for i, h in enumerate(headers)}

    courts, errors, names_seen = [], [], set()
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or all(v in (None, "") for v in row):
            continue
        name = row[col_idx["코트명"]]
        if not name:
            continue
        name = str(name).strip()
        if not name:
            continue
        if name in names_seen:
            errors.append(f"코트명 중복: '{name}'")
            continue
        names_seen.add(name)

        try:
            start_min = hhmm_to_min(row[col_idx["시작시간"]])
            end_min = hhmm_to_min(row[col_idx["종료시간"]])
            if start_min >= end_min:
                raise ValueError(f"코트 '{name}': 시작시간이 종료시간보다 같거나 늦습니다")
        except (ValueError, TypeError) as e:
            errors.append(str(e))
            continue

        slots = list(range(start_min, end_min, 30))
        courts.append({
            "name": name,
            "start_min": start_min,
            "end_min": end_min,
            "slots": slots,
        })

    return courts, errors


def build_schedule_slots(courts: list[dict]) -> list[dict]:
    slot_set = set()
    for c in courts:
        for s in c["slots"]:
            slot_set.add(s)
    sorted_slots = sorted(slot_set)
    schedule = []
    for s in sorted_slots:
        available_courts = [c["name"] for c in courts if s in c["slots"]]
        schedule.append({
            "slot_start": s,
            "slot_end": s + 30,
            "courts": available_courts,
        })
    return schedule


def clamp_mixed_wish(players: list[dict]) -> list[str]:
    """혼복희망을 '본인이 실제로 뛸 수 있는 최대 게임수'로 자르고 경고를 돌려준다.

    자르지 않으면 실현 불가능한 값(예: 30) 하나가 `mixed_limits()`의 혼복 하한을
    그만큼 밀어 올려 대진표 전체가 혼복이 된다(실측: 여복 3판 → 1판, 혼복 4판 → 9판,
    그런데 review는 PASS). 최소게임수가 `_max_games_no3` 클램프 + 경고로 막혀 있는 것과
    같은 방어를 대칭으로 건다.

    CLI(`main()`)와 웹(`run.py`)이 각자 경고 루프를 갖고 있으므로 **양쪽 모두** 이 함수를
    부른다 — 한쪽에만 넣으면 웹이 무방비가 된다(사용자는 웹만 쓴다).
    반드시 `attach_available_slots()` 뒤에 호출할 것.
    """
    warns = []
    for p in players:
        w = p.get("mixed_wish")
        if not w:
            continue
        cap = _max_games_streak(p.get("available_slots") or [], p.get("streak") or "")
        if p.get("max_games") is not None:
            cap = min(cap, p["max_games"])
        if w > cap:
            warns.append(f"'{p['name']}': 혼복희망({w}) > 본인이 뛸 수 있는 최대({cap}게임)"
                         f" — {cap}게임까지만 반영됨")
            p["mixed_wish"] = cap if cap > 0 else None
    return warns


def attach_available_slots(players: list[dict], schedule_slots: list[dict]) -> None:
    for p in players:
        avail = []
        for sl in schedule_slots:
            if p["in_min"] <= sl["slot_start"] and p["out_min"] >= sl["slot_end"]:
                avail.append(sl["slot_start"])
        p["available_slots"] = avail


_SEED_TIME_RE = re.compile(r"(\d{1,2}):(\d{2})")


def _clean_seed_name(v) -> str:
    """셀 값에서 표시용 장식을 제거해 순수 이름만 남긴다. history.py의 _clean_name과 동일 규칙 —

    씨드대진은 지난주 결과 엑셀을 그대로 복사해 붙여넣는 용도라 게스트 표기 '(G)'가
    섞여 들어오는 게 정상이다.
    """
    if v is None:
        return ""
    s = str(v).strip()
    if not s:
        return ""
    if s.endswith("(G)"):
        s = s[:-3].strip()
    return s


def _max_consecutive_run(slots: list[int]) -> int:
    """정렬된 슬롯 시작시각 목록에서 30분 간격으로 이어진 최장 연속 게임 수."""
    if not slots:
        return 0
    slots = sorted(slots)
    best = cur = 1
    for i in range(1, len(slots)):
        cur = cur + 1 if slots[i] - slots[i - 1] == 30 else 1
        best = max(best, cur)
    return best


def _seed_types_possible(t1: list, t2: list, players_by_id: dict) -> set:
    """이 씨드 배치를 완성할 수 있는 경기 종류 집합 ({"M","F","X"} 중).

    경기는 남복(전원 남)·여복(전원 여)·혼복(각 팀 남1+여1) 셋뿐이다. 셋 다 불가능한
    조합(예: 한 팀에 남자 둘 + 상대 팀에 여자 하나)을 그대로 두면 알고리즘이 그 코트를
    조용히 비우고, 사용자는 결과를 열어 보고서야 알게 된다.
    """
    g1 = [players_by_id[i]["gender"] for i in t1 if i]
    g2 = [players_by_id[i]["gender"] for i in t2 if i]
    allg = g1 + g2
    if not allg:
        return {"M", "F", "X"}
    out = set()
    if all(g == "M" for g in allg):
        out.add("M")
    if all(g == "F" for g in allg):
        out.add("F")
    # 혼복은 각 팀이 남1·여1 — 팀별로 같은 성별이 둘이면 불가능
    if all(side.count("M") <= 1 and side.count("F") <= 1 for side in (g1, g2)):
        out.add("X")
    return out


def _seed_club_error(t1: list, t2: list, players_by_id: dict) -> str:
    """교류전(클럽 2개 이상)에서 이 고정이 클럽 규칙을 어기는가. 어기면 사유 문자열.

    교류전은 '같은 팀은 같은 클럽, 상대는 반드시 다른 클럽'이 하드 규칙이라, 이를 어긴
    고정은 어떤 후보로도 완성되지 않는다(알고리즘이 그 코트를 비운다).
    """
    c1 = {players_by_id[i].get("club", "") for i in t1 if i}
    c2 = {players_by_id[i].get("club", "") for i in t2 if i}
    if len(c1) > 1 or len(c2) > 1:
        return "교류전에서는 같은 팀이 같은 클럽이어야 합니다"
    if c1 and c2 and c1 == c2:
        return "교류전에서는 상대 팀이 다른 클럽이어야 합니다"
    return ""


def parse_seed(ws, players: list[dict], schedule_slots: list[dict]) -> tuple[list[dict], list[str], list[str]]:
    """'씨드대진' 시트에서 고정 배치를 읽는다 → (pins, errors, warnings).

    pins = [{"slot_start": int, "court": str,
             "team1": [pid|None, pid|None], "team2": [pid|None, pid|None]}, ...]
    이름이 하나도 없는 (슬롯,코트)는 pins에 넣지 않는다.

    시트 좌표를 하드코딩하지 않고 history.py의 extract_pairs_from_bracket_xlsx()와 같은
    방식으로 스스로 찾는다 — build_template.py가 만든 그리드와 셀 위치가 같다는 전제는
    유지하되(성능상 이점), 헤더가 지워지거나 코트 순서가 바뀌어도 안전하게 동작한다.
    반드시 attach_available_slots() 호출 뒤에 불러야 p["available_slots"]가 채워져 있다.
    """
    pins: list[dict] = []
    errors: list[str] = []
    warnings: list[str] = []

    # 1) 헤더 행과 코트 base 열 찾기 (history.py와 동일한 방식)
    header_row = None
    bases: list[int] = []
    for r in range(1, min(ws.max_row, 8) + 1):
        cols = []
        for c in range(1, ws.max_column + 1):
            v = ws.cell(row=r, column=c).value
            if v and "번코트" in str(v):
                cols.append(c)
        if cols:
            header_row = r
            bases = cols
            break

    if header_row is None:
        # 1행(사용법 안내 타이틀)을 제외하고 뭔가 남아 있으면 헤더가 지워진 손상 상태로 본다.
        for row in ws.iter_rows(min_row=2, values_only=True):
            if any(v not in (None, "") and str(v).strip() for v in row):
                errors.append("씨드대진 시트: 헤더 행('N번코트')을 찾을 수 없습니다 — 시트 형식이 손상되었을 수 있습니다")
                break
        return pins, errors, warnings

    # 코트 이름이 겹치면 뒤 블록이 앞 블록을 덮어써 앞에 적은 고정이 조용히 사라진다
    # (그런데 그 사람의 '예약'은 남아 다른 코트에서도 못 쓰게 된다). 시트가 손상된 상태이므로 막는다.
    court_names = [str(ws.cell(row=header_row, column=c).value or "").replace("번코트", "").strip()
                   for c in bases]
    dup = sorted({c for c in court_names if court_names.count(c) > 1})
    if dup:
        errors.append(f"씨드대진 시트: 코트 헤더가 중복됩니다 ({', '.join(dup)}) — 시트 형식이 손상되었을 수 있습니다")
        return pins, errors, warnings

    name_to_pid = {p["name"]: p["id"] for p in players}
    players_by_id = {p["id"]: p for p in players}
    schedule_by_start = {s["slot_start"]: s for s in schedule_slots}
    # 교류전이면 팀 구성에 클럽 제약이 붙는다
    multi_club = len({p.get("club", "") for p in players if p.get("club", "")}) > 1
    # 그 슬롯에 실제로 뛸 수 있는 인원(성별별) — 인원이 모자라 완성 불가능한 고정을 잡는 데 쓴다
    avail_by_slot = {}
    for sl in schedule_slots:
        s0 = sl["slot_start"]
        avail_by_slot[s0] = (
            sum(1 for p in players if p["gender"] == "M" and s0 in p.get("available_slots", [])),
            sum(1 for p in players if p["gender"] == "F" and s0 in p.get("available_slots", [])),
        )

    # 2) 성함 행(헤더 다음 행부터 2행 간격) 목록 확보 — col1이 비면 데이터 그리드가 끝난 것.
    #    슬롯 수만큼만 읽는다: 경계가 없으면 시트 아래 메모처럼 '시간처럼 보이는 글자'가 든 행을
    #    슬롯으로 오인해 엉뚱한 고정을 만들어 낸다.
    name_rows = []
    row = header_row + 1
    while row <= ws.max_row and len(name_rows) < len(schedule_slots):
        col1 = ws.cell(row=row, column=1).value
        if col1 is None or not str(col1).strip():
            break
        name_rows.append((row, col1))
        row += 2

    seen_slots: set = set()
    slot_needs: dict = {}          # 슬롯 → 그 슬롯 고정들의 '가능한 경기 종류' 목록
    for name_row, col1_val in name_rows:
        # 이 행(=하나의 시간 슬롯)에 적힌 이름을 코트별로 먼저 걷어온다
        seats_by_base = {}
        row_has_names = False
        for base in bases:
            vals = [ws.cell(row=name_row, column=base + off).value for off in (0, 1, 3, 4)]
            cleaned = [_clean_seed_name(v) for v in vals]
            seats_by_base[base] = cleaned
            if any(cleaned):
                row_has_names = True
            # 한 칸 어긋나게 붙여넣으면 이름이 VS칸으로 밀린다. 그대로 두면 그 사람은
            # 조용히 사라지고 나머지는 옆자리로 밀려 '다른 대진'이 확정된다 —
            # 실패가 아니라 틀린 성공이라 더 위험하다.
            mid = _clean_seed_name(ws.cell(row=name_row, column=base + 2).value)
            if mid and mid not in ("VS", "vs", "혼"):
                errors.append(f"씨드대진 {name_row}행: 가운데(VS) 칸에 '{mid}'이(가) 있습니다 — "
                              f"열이 한 칸 어긋나게 붙여넣어졌는지 확인하세요")
                row_has_names = True
        if not row_has_names:
            continue

        m = _SEED_TIME_RE.search(str(col1_val))
        if not m:
            errors.append(f"씨드대진 {name_row}행: 이름이 있는데 시간을 인식할 수 없습니다 (칸 내용: '{col1_val}')")
            continue
        slot_start = int(m.group(1)) * 60 + int(m.group(2))
        if slot_start in seen_slots:
            errors.append(f"씨드대진 {name_row}행: 시각 {min_to_hhmm(slot_start)}이(가) 두 번 나옵니다"
                          f" — 시트 형식이 손상되었을 수 있습니다")
            continue
        seen_slots.add(slot_start)
        slot = schedule_by_start.get(slot_start)
        if slot is None:
            errors.append(f"씨드대진 {min_to_hhmm(slot_start)}: 현재 코트 시간표에 없는 시각입니다"
                           f" (코트 운영시간이 바뀌어 없어진 슬롯인데 씨드가 남아 있을 수 있습니다)")
            continue

        slot_seen_pids: set = set()
        for base, cleaned in seats_by_base.items():
            if not any(cleaned):
                continue
            hdr_val = ws.cell(row=header_row, column=base).value
            court_name = str(hdr_val).replace("번코트", "").strip() if hdr_val else ""
            tag = f"씨드대진 {min_to_hhmm(slot_start)} {court_name}코트"

            if court_name not in slot["courts"]:
                errors.append(f"{tag}: 이 시간에는 {court_name}코트를 운영하지 않습니다")
                continue

            pids: list = [None, None, None, None]
            resolved = []
            court_has_error = False
            for i, nm in enumerate(cleaned):
                if not nm:
                    continue
                pid = name_to_pid.get(nm)
                if pid is None:
                    errors.append(f"{tag}: '{nm}'은 참가자 시트에 없는 이름입니다")
                    court_has_error = True
                    continue
                p = players_by_id[pid]
                if slot_start not in p.get("available_slots", []):
                    errors.append(f"{tag}: '{nm}'은 이 시간(IN~OUT) 밖입니다")
                    court_has_error = True
                    continue
                if pid in slot_seen_pids:
                    errors.append(f"{tag}: '{nm}'이 같은 시간대에 두 번 배치되었습니다 (다른 코트 포함 중복 불가)")
                    court_has_error = True
                    continue
                slot_seen_pids.add(pid)
                pids[i] = pid
                resolved.append(p)

            if court_has_error:
                continue
            if all(x is None for x in pids):
                continue

            types = _seed_types_possible(pids[0:2], pids[2:4], players_by_id)
            if not types:
                errors.append(f"{tag}: 이 조합으로는 경기를 만들 수 없습니다 — "
                              f"남복(전원 남)·여복(전원 여)·혼복(각 팀 남1+여1) 중 하나여야 합니다")
                continue

            # 그 시간에 인원이 모자라면 어떤 종류로도 완성할 수 없다. 여기서 막지 않으면
            # 알고리즘이 그 코트를 조용히 비우고 사용자는 결과를 열어 보고서야 알게 된다.
            n_m, n_f = avail_by_slot.get(slot_start, (0, 0))
            feasible = ("M" in types and n_m >= 4) or ("F" in types and n_f >= 4) \
                       or ("X" in types and n_m >= 2 and n_f >= 2)
            if not feasible:
                errors.append(f"{tag}: 이 시간에 뛸 수 있는 인원이 남 {n_m}명·여 {n_f}명뿐이라 "
                              f"이 고정을 완성할 수 없습니다")
                continue

            if multi_club:
                why = _seed_club_error(pids[0:2], pids[2:4], players_by_id)
                if why:
                    errors.append(f"{tag}: {why}")
                    continue

            genders = {p["gender"] for p in resolved}
            if genders == {"M", "F"}:
                warnings.append(f"{tag}: 남녀가 섞여 있어 이 경기는 혼복이 됩니다")

            slot_needs.setdefault(slot_start, []).append(types)
            pins.append({"slot_start": slot_start, "court": court_name,
                         "team1": pids[0:2], "team2": pids[2:4]})

    # 3) 같은 슬롯의 고정들이 인원을 나눠 쓸 수 있는지 (경기 하나씩 보면 통과하지만
    #    합쳐 보면 인원이 모자란 경우가 있다 — 예: 여자 5명인 슬롯에서 A코트 여복 4명 +
    #    B코트 혼복 1자리 → 여자만 6명 필요. 이건 시드를 바꿔도 절대 안 채워진다)
    _MEN_REQ = {"M": 4, "F": 0, "X": 2}
    _WOMEN_REQ = {"M": 0, "F": 4, "X": 2}
    for s0, type_list in sorted(slot_needs.items()):
        n_m, n_f = avail_by_slot.get(s0, (0, 0))
        # 종류가 아직 안 정해진 고정은 '가장 적게 쓰는 종류' 기준으로 센다 → 확실한 하한
        need_m = sum(min(_MEN_REQ[t] for t in ts) for ts in type_list)
        need_f = sum(min(_WOMEN_REQ[t] for t in ts) for ts in type_list)
        need_all = 4 * len(type_list)
        if need_m > n_m or need_f > n_f or need_all > n_m + n_f:
            errors.append(
                f"씨드대진 {min_to_hhmm(s0)}: 이 시각의 고정 {len(type_list)}경기를 함께 채울 인원이"
                f" 모자랍니다 (필요 최소 남 {need_m}명·여 {need_f}명, 합 {need_all}명 /"
                f" 가능 남 {n_m}명·여 {n_f}명)")

    # 4) 격자 안에서 '읽지 않은 자리'에 참가자 이름이 있으면 알린다.
    #    결과 행에 붙여넣거나 시간 라벨이 지워지면 지금까지는 씨드가 통째로 증발하면서도
    #    아무 말이 없었다(웹은 씨드 표시 자체를 안 해 사용자가 알 길이 없다).
    seat_cols = {b + off for b in bases for off in (0, 1, 3, 4)}
    name_row_set = {r for r, _ in name_rows}
    stray = []
    for r in range(header_row + 1, ws.max_row + 1):
        for c in range(min(bases), ws.max_column + 1):
            if r in name_row_set and c in seat_cols:
                continue        # 정상적으로 읽은 자리
            nm = _clean_seed_name(ws.cell(row=r, column=c).value)
            if nm and nm in name_to_pid:
                stray.append(f"{r}행 {get_column_letter(c)}열 '{nm}'")
    if stray:
        errors.append("씨드대진: 인식되지 않는 위치에 이름이 있습니다 — "
                      + ", ".join(stray[:6]) + (" 외" if len(stray) > 6 else "")
                      + " (성함 행이 아닌 '결과' 행이거나 열이 어긋났을 수 있습니다)")

    # 5) 씨드 배치만으로 개인별 최대게임수·연속게임 한도를 넘는지 경고
    pid_slots: dict = defaultdict(list)
    for pin in pins:
        for pid in pin["team1"] + pin["team2"]:
            if pid:
                pid_slots[pid].append(pin["slot_start"])

    for pid, slots_here in pid_slots.items():
        p = players_by_id[pid]
        count = len(slots_here)
        if p.get("max_games") is not None and count > p["max_games"]:
            warnings.append(f"'{p['name']}': 씨드에서 이미 {count}게임 배치됨 — 최대게임수({p['max_games']})를 넘습니다")
        limit = STREAK_RUN_LIMIT.get(p.get("streak") or "", STREAK_RUN_LIMIT[""])
        run = _max_consecutive_run(slots_here)
        if run > limit:
            warnings.append(f"'{p['name']}': 씨드에서 {run}연속 배치 — 연속게임 규칙(최대 {limit}연속)을"
                             f" 넘습니다. 씨드가 우선하므로 그대로 진행됩니다")

    return pins, errors, warnings


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--in", dest="inp", required=True)
    ap.add_argument("--out", required=True)
    ap.add_argument("--members", default="",
                    help="멤버 설정 엑셀(부부 시트) 경로. 생략 시 내장 기본값 사용.")
    ap.add_argument("--no-members", action="store_true",
                    help="부부 페어 규칙을 끔 (기본값도 사용 안 함)")
    args = ap.parse_args()

    if not os.path.exists(args.inp):
        print(f"[에러] 입력 파일을 찾을 수 없음: {args.inp}", file=sys.stderr)
        sys.exit(1)

    wb = load_workbook(args.inp, data_only=True)

    if "참가자" not in wb.sheetnames:
        print("[에러] '참가자' 시트가 없습니다.", file=sys.stderr)
        sys.exit(1)
    if "코트" not in wb.sheetnames:
        print("[에러] '코트' 시트가 없습니다.", file=sys.stderr)
        sys.exit(1)

    all_errors = []
    try:
        players, perr = parse_players(wb["참가자"])
        all_errors.extend(perr)
    except ValueError as e:
        print(f"[에러] {e}", file=sys.stderr)
        sys.exit(1)

    try:
        courts, cerr = parse_courts(wb["코트"])
        all_errors.extend(cerr)
    except ValueError as e:
        print(f"[에러] {e}", file=sys.stderr)
        sys.exit(1)

    if all_errors:
        for e in all_errors:
            print(f"[에러] {e}", file=sys.stderr)
        sys.exit(1)

    if len(players) < 4:
        print(f"[에러] 참가자가 {len(players)}명입니다. 최소 4명 필요.", file=sys.stderr)
        sys.exit(1)
    if not courts:
        print("[에러] 사용 가능한 코트가 없습니다.", file=sys.stderr)
        sys.exit(1)

    schedule_slots = build_schedule_slots(courts)
    attach_available_slots(players, schedule_slots)

    # 경고 수집
    warnings = []
    males = [p for p in players if p["gender"] == "M"]
    females = [p for p in players if p["gender"] == "F"]
    if len(males) < 4:
        warnings.append(f"남자가 {len(males)}명이라 남자복식 불가. 혼복만 가능.")
    if len(females) < 4:
        warnings.append(f"여자가 {len(females)}명이라 여자복식 불가. 혼복만 가능.")

    for sl in schedule_slots:
        avail = [p for p in players if sl["slot_start"] in p["available_slots"]]
        if len(avail) < 4:
            t = f"{min_to_hhmm(sl['slot_start'])}~{min_to_hhmm(sl['slot_end'])}"
            warnings.append(f"슬롯 {t}: 가용 인원 {len(avail)}명 — 일부 코트 공석 가능")

    for p in players:
        if len(p["available_slots"]) == 0:
            warnings.append(f"'{p['name']}': 가용 슬롯 없음 — 코트 운영 시간과 IN/OUT 범위 확인 필요")
        if p["max_games"] is not None and p["max_games"] > len(p["available_slots"]):
            warnings.append(f"'{p['name']}': 최대게임수({p['max_games']}) > 가용 슬롯수({len(p['available_slots'])}) — 가용 슬롯 한도로 자연 제한됨")
        streak = p.get("streak") or ""
        cap = _max_games_streak(p["available_slots"], streak)
        if p.get("min_games") and p["min_games"] > cap:
            if streak == "no2":
                cap_label = "연속금지 하에 가능한 최대"
            elif streak == "ok3":
                cap_label = "3연속까지 허용 하에 가능한 최대"
            else:
                cap_label = "3연속 없이 가능한 최대"
            warnings.append(f"'{p['name']}': 최소게임수({p['min_games']}) > {cap_label}({cap}게임) — {cap}게임까지만 보장됨")
    warnings.extend(clamp_mixed_wish(players))

    # 씨드대진: 참가자 명단이 있어야 이름 대조가 되므로 반드시 위 참가자·코트 파싱 이후,
    # available_slots가 채워진 뒤에 부른다. 에러는 여기서 바로 중단(다른 에러들과 동일 방식).
    pins: list = []
    if "씨드대진" in wb.sheetnames:
        pins, seed_errors, seed_warnings = parse_seed(wb["씨드대진"], players, schedule_slots)
        if seed_errors:
            for e in seed_errors:
                print(f"[에러] {e}", file=sys.stderr)
            sys.exit(1)
        warnings.extend(seed_warnings)

    # 최소게임수 합계가 전체 자리(코트×슬롯×4)보다 많으면 다 지킬 수 없다 — 미리 알림
    total_seats = 0
    for sl in schedule_slots:
        n_avail = sum(1 for p in players if sl["slot_start"] in p["available_slots"])
        total_seats += min(len(sl["courts"]), n_avail // 4) * 4
    min_sum = sum(
        min(p["min_games"], _max_games_streak(p["available_slots"], p.get("streak") or ""))
        for p in players if p.get("min_games"))
    if min_sum > total_seats:
        warnings.append(f"최소게임수 합계({min_sum})가 전체 배정 가능 자리({total_seats})를 넘습니다 — 일부는 보장 못 할 수 있음")

    if warnings:
        for w in warnings:
            print(f"[경고] {w}", file=sys.stderr)

    # 부부 페어 설정: 파일 지정 > 내장 기본값 > (--no-members) 없음
    couples: list = []
    if not args.no_members:
        if args.members and os.path.exists(args.members):
            try:
                couples = parse_member_settings(args.members)
                print(f"[안내] 멤버 설정 반영: {args.members} — 부부 {len(couples)}쌍"
                      f" (부부페어 원함 {sum(1 for c in couples if c[2])}쌍)", file=sys.stderr)
            except Exception as e:
                print(f"[경고] 멤버 설정 파일을 읽지 못해 기본값을 사용합니다: {e}", file=sys.stderr)
                couples = [list(c) for c in COUPLES_DEFAULT]
        else:
            if args.members:
                print(f"[경고] 멤버 설정 파일 없음({args.members}) — 내장 기본값 사용", file=sys.stderr)
            couples = [list(c) for c in COUPLES_DEFAULT]
        names = {p["name"] for p in players}
        present = [c for c in couples if c[0] in names and c[1] in names]
        if present:
            def _ctag(c):
                tag = "원함" if c[2] else "피함"
                if len(c) > 3 and c[3]:
                    tag += "·종료 30분차이"
                return f"{c[0]}·{c[1]}({tag})"
            print(f"[안내] 이번 주 참가자 중 부부 {len(present)}쌍: "
                  + ", ".join(_ctag(c) for c in present),
                  file=sys.stderr)

    result = {
        "courts": courts,
        "players": players,
        "schedule_slots": schedule_slots,
        "couples": couples,
        "pins": pins,
        "warnings": warnings,
    }
    os.makedirs(os.path.dirname(os.path.abspath(args.out)), exist_ok=True)
    with open(args.out, "w", encoding="utf-8") as f:
        json.dump(result, f, ensure_ascii=False, indent=2)
    ok_msg = f"[OK] 파싱 완료: {args.out}  (참가자 {len(players)}명, 코트 {len(courts)}개, 슬롯 {len(schedule_slots)}개"
    if pins:
        seats = sum(1 for pin in pins for pid in pin["team1"] + pin["team2"] if pid)
        ok_msg += f", 씨드 {seats}자리"
    ok_msg += ")"
    print(ok_msg)


if __name__ == "__main__":
    main()
