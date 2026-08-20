"""Render a finalized tennis bracket to an Excel file matching the handwritten layout.

Usage:
    python render_bracket.py --parsed <01_parsed.json> --bracket <02_bracket.json>
                             --out <bracket.xlsx> [--date "26.5.30"] [--title "..."]
"""
from __future__ import annotations

import argparse
import json
import os
from collections import defaultdict
from openpyxl import Workbook
from openpyxl.formatting.rule import CellIsRule, FormulaRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


COURT_COLORS = [
    "FFFF99",  # yellow
    "FFCC99",  # orange
    "99CCFF",  # sky blue
    "C5E0B4",  # light green
    "FFB6C1",  # pink
    "D5A6E0",  # light purple
]
EMPTY_FILL = PatternFill("solid", fgColor="E0E0E0")
HEADER_FILL = PatternFill("solid", fgColor="BDD7EE")
TITLE_FILL = PatternFill("solid", fgColor="FFFFFF")
VS_FILL = PatternFill("solid", fgColor="F2F2F2")
SUMMARY_NUM_FILL = PatternFill("solid", fgColor="FFF2CC")
SUMMARY_GAME_FILL = PatternFill("solid", fgColor="FFFF99")

THIN = Side(border_style="thin", color="333333")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)

FONT_TITLE = Font(name="맑은 고딕", size=18, bold=True)
FONT_HEADER = Font(name="맑은 고딕", size=11, bold=True)
FONT_NAME = Font(name="맑은 고딕", size=11, bold=True)
# 씨드('씨드대진' 시트)로 사용자가 직접 고정한 이름 — 내가 정한 자리가 그대로 남았는지
# 결과 파일에서 바로 알아볼 수 있게 진한 파란색으로 구분한다.
FONT_NAME_SEED = Font(name="맑은 고딕", size=11, bold=True, color="1F4E79")
FONT_TIME = Font(name="맑은 고딕", size=10)
FONT_SMALL = Font(name="맑은 고딕", size=9, color="666666")
FONT_VS = Font(name="맑은 고딕", size=10, bold=True, color="999999")


def min_to_hhmm(v: int) -> str:
    return f"{v // 60:02d}:{v % 60:02d}"


FONT_STAT = Font(name="맑은 고딕", size=10)


GRID_SHEET = "대진표"
CALC_SHEET = "통계_계산"


def _build_calc_sheet(wb, ordered: list[dict], grid: dict, stats_header_row: int):
    """숨김 시트 '통계_계산': 대진표 '성함' 칸을 수식으로 읽는 사람×슬롯 매트릭스.

    블록 구성 (사람=행, 슬롯=열):
      P=출전(0/1) · S=연속게임 run · R=휴식연속(첫 경기 전 0 게이트) ·
      G=경기 직전 공백 슬롯수 · T=같은 팀 파트너 이름
    끝에 첫 슬롯 시작(분)·마지막 종료(분) 열. 통계 시트의 파생 지표가 전부
    이 매트릭스를 참조하므로 대진표를 손으로 고치면 통계도 자동 재계산된다.
    """
    slots = grid["schedule_slots"]
    n = len(slots)
    ws = wb.create_sheet(CALC_SHEET)
    ws.sheet_state = "hidden"

    L = get_column_letter
    c0 = 4  # A=이름, B=실질도착(분), C=여백
    blocks = {k: c0 + i * n for i, k in enumerate(("P", "S", "R", "G", "T"))}
    col_first = c0 + 5 * n
    col_last = col_first + 1
    grid_c0 = L(grid["courts_col_start"])
    grid_c1 = L(grid["courts_col_end"])

    ws.cell(row=3, column=1, value="이름")
    ws.cell(row=3, column=2, value="실질도착(분)")
    for k, label in (("P", "출전(슬롯별)"), ("S", "연속게임"), ("R", "휴식연속"),
                     ("G", "직전공백(슬롯)"), ("T", "파트너")):
        ws.cell(row=3, column=blocks[k], value=label)
    ws.cell(row=3, column=col_first, value="첫슬롯(분)")
    ws.cell(row=3, column=col_last, value="종료(분)")
    for j, sl in enumerate(slots):  # 행1=슬롯 시작(분)·행2=종료(분), P 블록 컬럼에 정렬
        ws.cell(row=1, column=blocks["P"] + j, value=sl["slot_start"])
        ws.cell(row=2, column=blocks["P"] + j, value=sl["slot_end"])

    refs = []
    for idx, s in enumerate(ordered):
        r = 4 + idx
        stats_row = stats_header_row + 1 + idx
        ws.cell(row=r, column=1, value=f"='통계'!$B${stats_row}")
        ws.cell(row=r, column=2, value=s.get("eff_in", s["in_min"]))

        p_cols = [L(blocks["P"] + j) for j in range(n)]
        s_cols = [L(blocks["S"] + j) for j in range(n)]
        r_cols = [L(blocks["R"] + j) for j in range(n)]
        g_cols = [L(blocks["G"] + j) for j in range(n)]
        t_cols = [L(blocks["T"] + j) for j in range(n)]

        for j in range(n):
            nr = grid["data_start_row"] + 2 * j  # 이 슬롯의 '성함' 행
            ws.cell(row=r, column=blocks["P"] + j,
                    value=f"=IF(COUNTIF('{GRID_SHEET}'!${grid_c0}${nr}:${grid_c1}${nr},$A{r})>0,1,0)")
            if j == 0:
                ws.cell(row=r, column=blocks["S"], value=f"={p_cols[0]}{r}")
                ws.cell(row=r, column=blocks["R"], value=0)
                ws.cell(row=r, column=blocks["G"], value=0)
            else:
                ws.cell(row=r, column=blocks["S"] + j,
                        value=f"=IF({p_cols[j]}{r}=1,{s_cols[j - 1]}{r}+1,0)")
                ws.cell(row=r, column=blocks["R"] + j,
                        value=(f"=IF({p_cols[j]}{r}=1,0,"
                               f"IF(SUM(${p_cols[0]}{r}:{p_cols[j]}{r})=0,0,{r_cols[j - 1]}{r}+1))"))
                ws.cell(row=r, column=blocks["G"] + j,
                        value=f"=IF({p_cols[j]}{r}=1,{r_cols[j - 1]}{r},0)")
            # T: 자기가 낀 팀 2칸 중 자기 아닌 쪽 (코트별 팀1·팀2 전부 검사)
            expr = '""'
            for ci in range(len(grid["courts"]) - 1, -1, -1):
                base = grid["courts_col_start"] + ci * grid["cols_per_court"]
                for a, b in ((base + 3, base + 4), (base, base + 1)):
                    for x, y in ((b, a), (a, b)):
                        expr = (f"IF($A{r}='{GRID_SHEET}'!${L(x)}${nr},"
                                f"'{GRID_SHEET}'!${L(y)}${nr},{expr})")
            ws.cell(row=r, column=blocks["T"] + j, value="=" + expr)

        p_rng = f"{p_cols[0]}{r}:{p_cols[-1]}{r}"
        times1 = f"${p_cols[0]}$1:${p_cols[-1]}$1"
        times2 = f"${p_cols[0]}$2:${p_cols[-1]}$2"
        ws.cell(row=r, column=col_first,
                value=f"=IF(SUM({p_rng})=0,-1,INDEX({times1},MATCH(1,{p_rng},0)))")
        ws.cell(row=r, column=col_last,
                value=f"=IF(SUM({p_rng})=0,-1,LOOKUP(2,1/({p_rng}=1),{times2}))")

        refs.append(dict(
            p_rng=f"'{CALC_SHEET}'!{p_rng}",
            s_rng=f"'{CALC_SHEET}'!{s_cols[0]}{r}:{s_cols[-1]}{r}",
            g_rng=f"'{CALC_SHEET}'!{g_cols[0]}{r}:{g_cols[-1]}{r}",
            t_rng=f"'{CALC_SHEET}'!{t_cols[0]}{r}:{t_cols[-1]}{r}",
            first=f"'{CALC_SHEET}'!${L(col_first)}${r}",
            last=f"'{CALC_SHEET}'!${L(col_last)}${r}",
            eff=f"'{CALC_SHEET}'!$B${r}",
        ))
    last_end_rng = f"'{CALC_SHEET}'!${L(col_last)}$4:${L(col_last)}${3 + len(ordered)}"
    return refs, last_end_rng


def _mixed_count_formula(grid: dict, name_cell: str) -> str:
    """이 사람의 혼복 게임수: 팀 사이 칸이 "혼"인 행에 이름이 있는 횟수 (코트별 합)."""
    L = get_column_letter
    r0, r1 = grid["data_start_row"], grid["last_grid_row"]
    terms = []
    for ci in range(len(grid["courts"])):
        base = grid["courts_col_start"] + ci * grid["cols_per_court"]
        rng = lambda c: f"'{GRID_SHEET}'!${L(c)}${r0}:${L(c)}${r1}"
        eq = "+".join(f"({rng(c)}={name_cell})" for c in (base, base + 1, base + 3, base + 4))
        terms.append(f'SUMPRODUCT(({eq})*({rng(base + 2)}="혼"))')
    return "=" + "+".join(terms)


def _build_stats_sheet(wb, bracket: dict, is_exchange: bool, clubs_ordered: list, grid: dict) -> None:
    """'통계' 시트: 사람별 통계 테이블 + 전체 요약 — 전부 수식 기반(손 수정 시 자동 재계산)."""
    player_stats = bracket["player_stats"]

    ws = wb.create_sheet("통계")

    cols = [("번호", 5), ("이름", 12)]
    if is_exchange:
        cols.append(("클럽", 11))
    cols += [
        ("성별", 5), ("구력", 5), ("참석", 12),
        ("게임수", 7), ("남복", 6), ("여복", 6), ("혼복", 6),
        ("첫 경기", 8), ("마지막 경기", 10),
        ("첫 대기(분)", 10), ("최대 연속 게임", 9), ("최대 연속 휴식(분)", 10),
        ("1시간+ 공백(회)", 9), ("총 대기(분)", 10),
        ("파트너 수", 8), ("같은 짝 반복", 8),
    ]
    for i, (_, w) in enumerate(cols, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    ws.cell(row=1, column=1, value="사람별 통계 — 수식 기반: 대진표 시트 '성함' 칸을 고치면 자동 재계산됩니다"
                                   " (혼복 판정은 팀 사이 칸의 \"혼\" 표기 기준 · 계산 상세는 숨김 시트 '통계_계산')")
    ws.cell(row=1, column=1).font = FONT_SMALL
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(cols))

    HEADER_ROW = 2
    for i, (title, _) in enumerate(cols, start=1):
        c = ws.cell(row=HEADER_ROW, column=i, value=title)
        c.font = FONT_HEADER
        c.fill = HEADER_FILL
        c.alignment = CENTER
        c.border = BORDER
    ws.row_dimensions[HEADER_ROW].height = 28

    # 정렬: 명단 패널과 동일(가용 슬롯 많은 순). 교류전이면 클럽별로 묶는다.
    if is_exchange:
        ordered = []
        for cname in clubs_ordered:
            ordered += sorted([s for s in player_stats if s.get("club", "") == cname],
                              key=lambda s: -s["available_slots"])
        ordered += sorted([s for s in player_stats if s.get("club", "") not in clubs_ordered],
                          key=lambda s: -s["available_slots"])
    else:
        ordered = sorted(player_stats, key=lambda s: -s["available_slots"])

    # 숨김 계산 시트: 대진표를 읽는 사람×슬롯 매트릭스 (통계 수식들의 근거)
    refs, last_end_rng = _build_calc_sheet(wb, ordered, grid, HEADER_ROW)

    L = get_column_letter
    col = {title: i for i, (title, _) in enumerate(cols, start=1)}  # 헤더 제목 → 컬럼 번호

    for idx, s in enumerate(ordered):
        r = HEADER_ROW + 1 + idx
        ref = refs[idx]
        g_cell = f"{L(col['게임수'])}{r}"
        x_cell = f"{L(col['혼복'])}{r}"
        gen_cell = f"{L(col['성별'])}{r}"
        pn_cell = f"{L(col['파트너 수'])}{r}"
        wait_expr = f"MAX(0,{ref['first']}-{ref['eff']})"
        row_vals = [idx + 1, display_name(s)]
        if is_exchange:
            row_vals.append(s.get("club", ""))
        row_vals += [
            "남" if s["gender"] == "M" else "여",
            s["exp"],
            f"{min_to_hhmm(s['in_min'])}~{min_to_hhmm(s['out_min'])}",
            f"=SUM({ref['p_rng']})",
            f'=IF({gen_cell}="남",{g_cell}-{x_cell},0)',
            f'=IF({gen_cell}="여",{g_cell}-{x_cell},0)',
            _mixed_count_formula(grid, f"$B{r}"),
            f'=IF({g_cell}=0,"-",TEXT({ref["first"]}/1440,"HH:MM"))',
            f'=IF({g_cell}=0,"-",TEXT({ref["last"]}/1440,"HH:MM"))',
            f'=IF({g_cell}=0,"-",{wait_expr})',
            f"=MAX({ref['s_rng']})",
            f'=IF({g_cell}<2,"-",MAX({ref["g_rng"]})*30)',
            f'=COUNTIF({ref["g_rng"]},">=2")',
            f'=IF({g_cell}=0,"-",{wait_expr}+SUM({ref["g_rng"]})*30)',
            f'=IF({g_cell}=0,0,SUMPRODUCT(({ref["t_rng"]}<>"")/COUNTIF({ref["t_rng"]},{ref["t_rng"]}&"")))',
            f'=IF({g_cell}=0,0,SUMPRODUCT(({ref["t_rng"]}<>"")*1)-{pn_cell})',
        ]
        for ci, v in enumerate(row_vals, start=1):
            cell = ws.cell(row=r, column=ci, value=v)
            cell.font = FONT_STAT
            cell.alignment = CENTER
            cell.border = BORDER
        if s["membership"] == "게스트":
            ws.cell(row=r, column=2).font = Font(name="맑은 고딕", size=10, italic=True)

    # 눈에 띄어야 하는 값(값이 수식이라 조건부 서식으로): 1시간 이상 휴식 / 1시간+ 공백 발생
    first_data = HEADER_ROW + 1
    last_data = HEADER_ROW + len(ordered)
    rest_l = L(col["최대 연속 휴식(분)"])
    lg_l = L(col["1시간+ 공백(회)"])
    warn_font = Font(name="맑은 고딕", size=10, bold=True, color="C00000")
    if ordered:
        ws.conditional_formatting.add(
            f"{rest_l}{first_data}:{rest_l}{last_data}",
            FormulaRule(formula=[f"AND(ISNUMBER({rest_l}{first_data}),{rest_l}{first_data}>=60)"],
                        font=warn_font))
        ws.conditional_formatting.add(
            f"{lg_l}{first_data}:{lg_l}{last_data}",
            CellIsRule(operator="greaterThan", formula=["0"], font=warn_font))

    ws.freeze_panes = f"A{HEADER_ROW + 1}"

    # ── 전체 요약 (역시 수식 — 손 수정 시 자동 재계산) ──
    base = HEADER_ROW + len(ordered) + 2

    def col_rng(title):
        cl = L(col[title])
        return f"{cl}{first_data}:{cl}{last_data}"

    g_rng, m_rng, f_rng, x_rng = (col_rng(t) for t in ("게임수", "남복", "여복", "혼복"))
    lg_rng, ti_rng = col_rng("1시간+ 공백(회)"), col_rng("총 대기(분)")
    rows = [
        ("전체 요약", ""),
        ("총 매치 수", f'=ROUND(SUM({g_rng})/4,2)&"  (남복 "&ROUND(SUM({m_rng})/4,2)'
                      f'&" · 여복 "&ROUND(SUM({f_rng})/4,2)&" · 혼복 "&ROUND(SUM({x_rng})/4,2)&")"'),
        ("참가자 수", len(player_stats)),
        ("게임수 평균 / 최대 / 최소",
         f'=ROUND(AVERAGE({g_rng}),2)&" / "&MAX({g_rng})&" / "&MIN({g_rng})'),
        ("1시간 이상 공백", f'=SUM({lg_rng})&"건 / "&COUNTIF({lg_rng},">0")&"명"'),
        ("총 대기시간 합계", f'=SUM({ti_rng})&"분"'),
        ("마지막 경기 종료",
         f'=IF(MAX({last_end_rng})<=0,"-",TEXT(MAX({last_end_rng})/1440,"HH:MM"))'),
    ]
    if is_exchange:
        club_counts = defaultdict(int)
        for s in player_stats:
            club_counts[s.get("club", "")] += 1
        rows.append(("교류전 클럽", "  ·  ".join(f"{c} {club_counts[c]}명" for c in clubs_ordered)))
    for i, (k, v) in enumerate(rows):
        kc = ws.cell(row=base + i, column=1, value=k)
        kc.font = FONT_HEADER
        ws.merge_cells(start_row=base + i, start_column=1, end_row=base + i, end_column=3)
        vc = ws.cell(row=base + i, column=4, value=v)
        vc.font = FONT_STAT
        vc.alignment = Alignment(horizontal="left", vertical="center")
        ws.merge_cells(start_row=base + i, start_column=4, end_row=base + i, end_column=len(cols))


def display_name(p_stat: dict) -> str:
    n = p_stat["name"]
    if p_stat["membership"] == "게스트":
        return f"{n}(G)"
    return n


def render(parsed: dict, bracket: dict, out_path: str, date_str: str, title: str) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "대진표"

    courts = parsed["courts"]
    schedule_slots = parsed["schedule_slots"]
    matches = bracket["matches"]
    player_stats = bracket["player_stats"]

    clubs_present = {s.get("club", "") for s in player_stats if s.get("club", "")}
    is_exchange = len(clubs_present) > 1

    def _club_order_key(c):
        # 정회원이 많은 클럽(=홈)을 왼쪽으로
        members = sum(1 for s in player_stats
                      if s.get("club", "") == c and s.get("membership") == "정회원")
        return (-members, c)
    clubs_ordered = sorted(clubs_present, key=_club_order_key)
    two_club_mode = is_exchange and len(clubs_present) == 2

    matches_by_slot_court = defaultdict(dict)
    for m in matches:
        matches_by_slot_court[m["slot_start"]][m["court"]] = m

    # 컬럼 레이아웃
    # 1: 구분(번호+시간), 2: 성함/결과 라벨
    # 코트별 4컬럼: t1a, t1b, VS, t2a, t2b (= 5컬럼)
    # 우측 패널: 번호 | 이름 | 게임수 | (간격) | 번호 | 이름 | 게임수
    LABEL_COL_START = 1
    LABEL_COL_END = 2
    COURTS_COL_START = 3
    cols_per_court = 5
    courts_col_end = COURTS_COL_START + cols_per_court * len(courts) - 1

    summary_left_start = courts_col_end + 2
    SUM_NUM_W, SUM_NAME_W, SUM_GAME_W = 4, 12, 7
    summary_col_count = 3 + 1 + 3
    summary_right_end = summary_left_start + summary_col_count - 1

    # 컬럼 너비
    ws.column_dimensions[get_column_letter(LABEL_COL_START)].width = 9
    ws.column_dimensions[get_column_letter(LABEL_COL_END)].width = 7
    for i, _ in enumerate(courts):
        base = COURTS_COL_START + i * cols_per_court
        ws.column_dimensions[get_column_letter(base)].width = 10
        ws.column_dimensions[get_column_letter(base + 1)].width = 10
        ws.column_dimensions[get_column_letter(base + 2)].width = 4
        ws.column_dimensions[get_column_letter(base + 3)].width = 10
        ws.column_dimensions[get_column_letter(base + 4)].width = 10
    ws.column_dimensions[get_column_letter(courts_col_end + 1)].width = 1
    for off, w in zip(range(7), [SUM_NUM_W, SUM_NAME_W, SUM_GAME_W, 1, SUM_NUM_W, SUM_NAME_W, SUM_GAME_W]):
        ws.column_dimensions[get_column_letter(summary_left_start + off)].width = w

    # 행 1: 타이틀
    ws.cell(row=1, column=LABEL_COL_START, value=title)
    ws.merge_cells(start_row=1, start_column=LABEL_COL_START,
                   end_row=1, end_column=courts_col_end)
    ws.cell(row=1, column=LABEL_COL_START).font = FONT_TITLE
    ws.cell(row=1, column=LABEL_COL_START).alignment = CENTER
    ws.row_dimensions[1].height = 30

    # 날짜 (타이틀 영역 우측)
    ws.cell(row=1, column=summary_left_start, value=date_str)
    ws.merge_cells(start_row=1, start_column=summary_left_start,
                   end_row=1, end_column=summary_right_end)
    ws.cell(row=1, column=summary_left_start).font = FONT_HEADER
    ws.cell(row=1, column=summary_left_start).alignment = Alignment(horizontal="right", vertical="center")

    # 행 2-3: 헤더
    HEADER_ROW = 2
    ws.cell(row=HEADER_ROW, column=LABEL_COL_START, value="구분")
    ws.merge_cells(start_row=HEADER_ROW, start_column=LABEL_COL_START,
                   end_row=HEADER_ROW, end_column=LABEL_COL_END)
    c = ws.cell(row=HEADER_ROW, column=LABEL_COL_START)
    c.font = FONT_HEADER
    c.fill = HEADER_FILL
    c.alignment = CENTER
    c.border = BORDER

    for i, court in enumerate(courts):
        base = COURTS_COL_START + i * cols_per_court
        ws.cell(row=HEADER_ROW, column=base, value=f"{court['name']}번코트")
        ws.merge_cells(start_row=HEADER_ROW, start_column=base,
                       end_row=HEADER_ROW, end_column=base + cols_per_court - 1)
        hc = ws.cell(row=HEADER_ROW, column=base)
        hc.font = FONT_HEADER
        hc.fill = PatternFill("solid", fgColor=COURT_COLORS[i % len(COURT_COLORS)])
        hc.alignment = CENTER
        hc.border = BORDER

    # 우측 패널 헤더
    _left_hdr = clubs_ordered[0] if two_club_mode else "참가자"
    _right_hdr = clubs_ordered[1] if two_club_mode else "참가자"
    ws.cell(row=HEADER_ROW, column=summary_left_start, value="")
    ws.cell(row=HEADER_ROW, column=summary_left_start + 1, value=_left_hdr)
    ws.cell(row=HEADER_ROW, column=summary_left_start + 2, value="게임수")
    ws.cell(row=HEADER_ROW, column=summary_left_start + 4, value="")
    ws.cell(row=HEADER_ROW, column=summary_left_start + 5, value=_right_hdr)
    ws.cell(row=HEADER_ROW, column=summary_left_start + 6, value="게임수")
    for off in (0, 1, 2, 4, 5, 6):
        c = ws.cell(row=HEADER_ROW, column=summary_left_start + off)
        c.font = FONT_HEADER
        c.fill = HEADER_FILL
        c.alignment = CENTER
        c.border = BORDER

    # 데이터 행: 슬롯별로 2행 (성함, 결과) — 각 행 높이 22
    data_start_row = HEADER_ROW + 1
    for slot_idx, slot in enumerate(schedule_slots):
        name_row = data_start_row + slot_idx * 2
        result_row = name_row + 1
        ws.row_dimensions[name_row].height = 24
        ws.row_dimensions[result_row].height = 24

        ws.cell(row=name_row, column=LABEL_COL_START,
                value=f"{slot_idx + 1}번 게임\n{min_to_hhmm(slot['slot_start'])}\n~{min_to_hhmm(slot['slot_end'])}")
        ws.merge_cells(start_row=name_row, start_column=LABEL_COL_START,
                       end_row=result_row, end_column=LABEL_COL_START)
        c = ws.cell(row=name_row, column=LABEL_COL_START)
        c.font = FONT_TIME
        c.alignment = CENTER
        c.border = BORDER

        ws.cell(row=name_row, column=LABEL_COL_END, value="성함")
        ws.cell(row=result_row, column=LABEL_COL_END, value="결과")
        for r in (name_row, result_row):
            c = ws.cell(row=r, column=LABEL_COL_END)
            c.font = FONT_HEADER
            c.alignment = CENTER
            c.border = BORDER
            c.fill = HEADER_FILL

        for i, court in enumerate(courts):
            base = COURTS_COL_START + i * cols_per_court
            court_color = PatternFill("solid", fgColor=COURT_COLORS[i % len(COURT_COLORS)])

            for col_off in range(cols_per_court):
                for r in (name_row, result_row):
                    cell = ws.cell(row=r, column=base + col_off, value=None)
                    cell.border = BORDER
                    cell.alignment = CENTER

            court_active = slot["slot_start"] in court["slots"]
            if not court_active:
                continue

            m = matches_by_slot_court.get(slot["slot_start"], {}).get(court["name"])
            if m is None:
                continue

            stats_by_id = {s["id"]: s for s in player_stats}
            seeded = set(m.get("pinned") or ())
            for k, p_id in enumerate(m["team1"]):
                cell = ws.cell(row=name_row, column=base + k, value=display_name(stats_by_id[p_id]))
                cell.font = FONT_NAME_SEED if p_id in seeded else FONT_NAME
                cell.fill = court_color
                cell.alignment = CENTER
                cell.border = BORDER
            vs_cell = ws.cell(row=name_row, column=base + 2, value="VS")
            vs_cell.font = FONT_VS
            vs_cell.fill = VS_FILL
            vs_cell.alignment = CENTER
            vs_cell.border = BORDER
            for k, p_id in enumerate(m["team2"]):
                cell = ws.cell(row=name_row, column=base + 3 + k, value=display_name(stats_by_id[p_id]))
                cell.font = FONT_NAME_SEED if p_id in seeded else FONT_NAME
                cell.fill = court_color
                cell.alignment = CENTER
                cell.border = BORDER

            if m["type"] == "X":
                vs_cell.value = "혼"

    # 우측 패널: 참가자 + 게임수
    # 게임수는 정적 숫자가 아니라 COUNTIF 수식 — 대진표 칸의 이름을 직접 고치면 자동 재계산된다.
    # 교류전(클럽 2개)일 때는 좌/우 열을 클럽별로 분리.
    panel_start = data_start_row
    last_grid_row = data_start_row + len(schedule_slots) * 2 - 1
    grid_abs = (f"${get_column_letter(COURTS_COL_START)}${data_start_row}:"
                f"${get_column_letter(courts_col_end)}${last_grid_row}")

    def write_entry(row, num_col, seq_num, p):
        c1 = ws.cell(row=row, column=num_col, value=seq_num)
        c1.font = FONT_HEADER
        c1.fill = SUMMARY_NUM_FILL
        c1.alignment = CENTER
        c1.border = BORDER
        ws.merge_cells(start_row=row, start_column=num_col, end_row=row + 1, end_column=num_col)

        name_cell = ws.cell(row=row, column=num_col + 1, value=display_name(p))
        name_cell.font = FONT_NAME
        name_cell.fill = SUMMARY_NUM_FILL
        name_cell.alignment = CENTER
        name_cell.border = BORDER
        info = f"{min_to_hhmm(p['in_min'])}~{min_to_hhmm(p['out_min'])}"
        if p.get("min_games"):
            info += f" / 최소 {p['min_games']}게임"
        if p.get("streak") == "no2":
            info += " / 연속 금지"
        elif p.get("streak") == "ok3":
            info += " / 3연속 허용"
        if p.get("max_games") is not None:
            info += f" / 최대 {p['max_games']}게임"
        if is_exchange and p.get("club"):
            info += f" · {p['club']}"
        ci = ws.cell(row=row + 1, column=num_col + 1, value=info)
        ci.font = FONT_SMALL
        ci.alignment = CENTER
        ci.border = BORDER

        # 게임수 = 대진표 영역에서 이 사람 이름이 나오는 횟수 (수정 시 자동 갱신)
        name_ref = f"{get_column_letter(num_col + 1)}{row}"
        g = ws.cell(row=row, column=num_col + 2, value=f"=COUNTIF({grid_abs},{name_ref})")
        g.font = FONT_NAME
        g.fill = SUMMARY_GAME_FILL
        g.alignment = CENTER
        g.border = BORDER
        ws.merge_cells(start_row=row, start_column=num_col + 2, end_row=row + 1, end_column=num_col + 2)

    def write_total(row, game_col, first_row, last_row):
        rng = f"{get_column_letter(game_col)}{first_row}:{get_column_letter(game_col)}{last_row}"
        t = ws.cell(row=row, column=game_col, value=f"=SUM({rng})")
        t.font = FONT_HEADER
        t.fill = SUMMARY_GAME_FILL
        t.alignment = CENTER
        t.border = BORDER

    left_num_col = summary_left_start
    right_num_col = summary_left_start + 4

    if two_club_mode:
        # 좌=클럽1, 우=클럽2 로 분리
        col_lists = [
            (left_num_col, sorted([s for s in player_stats if s.get("club", "") == clubs_ordered[0]],
                                  key=lambda s: -s["available_slots"])),
            (right_num_col, sorted([s for s in player_stats if s.get("club", "") == clubs_ordered[1]],
                                   key=lambda s: -s["available_slots"])),
        ]
    else:
        stats_sorted = sorted(player_stats, key=lambda s: -s["available_slots"])
        half = (len(stats_sorted) + 1) // 2
        col_lists = [
            (left_num_col, stats_sorted[:half]),
            (right_num_col, stats_sorted[half:]),
        ]

    max_len = max((len(lst) for _, lst in col_lists), default=0)
    panel_end_row = panel_start + max_len * 2
    for num_col, lst in col_lists:
        for idx, p in enumerate(lst):
            write_entry(panel_start + idx * 2, num_col, idx + 1, p)
        if lst:
            write_total(panel_end_row, num_col + 2, panel_start, panel_end_row - 1)

    ws.sheet_view.showGridLines = False
    ws.print_options.horizontalCentered = True
    ws.page_setup.orientation = "landscape"

    # 두 번째 시트: 사람별 통계 + 전체 요약 (수식 기반) — 대진표 그리드 좌표를 넘긴다
    grid = dict(
        data_start_row=data_start_row,
        last_grid_row=last_grid_row,
        courts_col_start=COURTS_COL_START,
        courts_col_end=courts_col_end,
        cols_per_court=cols_per_court,
        schedule_slots=schedule_slots,
        courts=courts,
    )
    _build_stats_sheet(wb, bracket, is_exchange, clubs_ordered, grid)

    wb.save(out_path)


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--parsed", required=True)
    ap.add_argument("--bracket", required=True)
    ap.add_argument("--out", required=True)
    ap.add_argument("--date", default="")
    ap.add_argument("--title", default="우리 테니스 클럽 대진표")
    args = ap.parse_args()

    with open(args.parsed, "r", encoding="utf-8") as f:
        parsed = json.load(f)
    with open(args.bracket, "r", encoding="utf-8") as f:
        bracket = json.load(f)

    render(parsed, bracket, args.out, args.date, args.title)
    print(f"[OK] 대진표 출력: {args.out}")


if __name__ == "__main__":
    main()
