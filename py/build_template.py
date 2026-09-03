"""Build an empty tennis bracket input Excel template.

Usage:
    python build_template.py --out <output.xlsx>
"""
from __future__ import annotations

import argparse
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation


HEADER_FILL = PatternFill("solid", fgColor="D9E1F2")
GUIDE_FILL = PatternFill("solid", fgColor="FFF2CC")
INACTIVE_FILL = PatternFill("solid", fgColor="D9D9D9")   # 씨드대진: 운영 안 하는 코트 칸
SEED_VS_FILL = PatternFill("solid", fgColor="F2F2F2")
HEADER_FONT = Font(name="맑은 고딕", size=11, bold=True)
BODY_FONT = Font(name="맑은 고딕", size=11)
THIN = Side(border_style="thin", color="888888")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
CENTER = Alignment(horizontal="center", vertical="center")
LEFT = Alignment(horizontal="left", vertical="center", wrap_text=True)

# 코트 기본값 — '코트' 시트와 '씨드대진' 시트가 같은 슬롯 집합을 써야 하므로 공유 상수로 뺐다.
# 여기 하드코딩이 두 곳으로 갈라지면 코트 시간이 바뀔 때 씨드대진 레이아웃만 조용히 어긋난다.
# 26.9.5~ 코트 운영시간: 1·2번(A·B) 08:00~12:00, 3번(C) 07:00~09:00
#   (26.8.8~ A·B 07:00 시작이었으나 26.9.5부터 08:00 시작으로 변경. C는 그대로)
COURTS_DEFAULT = [
    ("A", "08:00", "12:00"),
    ("B", "08:00", "12:00"),
    ("C", "07:00", "09:00"),
]


def _style_header(cell):
    cell.font = HEADER_FONT
    cell.fill = HEADER_FILL
    cell.alignment = CENTER
    cell.border = BORDER


def _style_body(cell):
    cell.font = BODY_FONT
    cell.alignment = CENTER
    cell.border = BORDER


PREFILL_FROM_IMAGE = [
    # 정회원 사전채움 로스터 — 26.8.2 사용자 확정본(테니스_입력양식_사전채움_25명.xlsx)
    #   + 26.8.6 사용자 확정 '기본 반영 사항'(개인별 평소 IN/OUT 희망) 반영.
    # 정회원 25명 전원. 게스트는 매주 달라지므로 담지 않는다.
    # IN/OUT은 전원 07:00~12:00 기본. 26.9.5~ A·B 코트가 08:00 시작으로 바뀌었으나
    #   개인 IN은 07:00 그대로 둔다(사용자 확정 26.9.3) — C코트가 07:00~09:00로 남아 있고,
    #   코트가 안 열린 슬롯은 어차피 배정되지 않는다.
    # 예외(26.8.6): 이성돈(09:00~12:00, 최대 3게임), 원유철·이지은(08:00~11:00, 본업 수업),
    #   이강진(~11:00 종료), 김효순·박진우(~10:00 종료), 남궁석(08:30~ 시작),
    #   최종인·경성현(09:00~12:00 고정, 26.8.20~), 이성돈 연속게임=금지(26.8.20~).
    # 신혁재·방미라 부부 '종료시각 반드시 30분 차이'는 멤버 설정(COUPLES_DEFAULT 종료시간차=30)이 담당.
    # (이름, 성별, 구력, 구분, IN, OUT, 최소게임수, 최대게임수, 연속게임, 메모)  단일 클럽이라 클럽 칸은 비워 둠.
    ("경성현", "남", 4,  "정회원", "09:00", "12:00", "", "", "", "09:00~12:00 고정"),
    ("김준학", "남", 10, "정회원", "07:00", "12:00", "", "", "", ""),
    ("김효순", "남", 5,  "정회원", "07:00", "10:00", "", "", "", "10:00 종료 희망"),
    ("노남숙", "여", 3,  "정회원", "07:00", "12:00", "", "", "", ""),
    ("서명숙", "여", 5,  "정회원", "07:00", "12:00", "", "", "", ""),
    ("이지은", "여", 3,  "정회원", "08:00", "11:00", "", "", "", "본업 수업 — 08:00~11:00 참석"),
    ("정정희", "여", 5,  "정회원", "07:00", "12:00", "", "", "", ""),
    ("박경수", "남", 5,  "정회원", "07:00", "12:00", "", "", "", ""),
    ("박진우", "남", 7,  "정회원", "07:00", "10:00", "", "", "", "10:00 종료 희망"),
    ("서종수", "남", 5,  "정회원", "07:00", "12:00", "", "", "", ""),
    ("최종인", "남", 10, "정회원", "09:00", "12:00", "", "", "", "09:00~12:00 고정"),
    ("원유철", "남", 4,  "정회원", "08:00", "11:00", "", "", "", "본업 수업 — 08:00~11:00 참석"),
    ("이강진", "남", 4,  "정회원", "07:00", "11:00", "", "", "", "11:00 종료 희망"),
    ("이성돈", "남", 10, "정회원", "09:00", "12:00", "", 3, "금지", "연속 게임 어려움 — 한 게임 뛰고 한 게임 쉼"),
    ("이성수", "남", 7,  "정회원", "07:00", "12:00", "", "", "", ""),
    ("정진락", "남", 10, "정회원", "07:00", "12:00", "", "", "", ""),
    ("민기준", "남", 5,  "정회원", "07:00", "12:00", "", "", "", ""),
    ("임성훈", "남", 5,  "정회원", "07:00", "12:00", "", "", "", ""),
    ("김도윤", "남", 10, "정회원", "07:00", "12:00", "", "", "", ""),
    ("남궁석", "남", 10, "정회원", "08:30", "12:00", "", "", "", "08:30 시작 희망"),
    ("정재동", "남", 10, "정회원", "07:00", "12:00", "", "", "", ""),
    ("한병익", "남", 10, "정회원", "07:00", "12:00", "", "", "", ""),
    ("전혜선", "여", 7,  "정회원", "07:00", "12:00", "", "", "", ""),
    ("신혁재", "남", 5,  "정회원", "07:00", "12:00", "", "", "", "방미라와 종료 30분 차이(자동 반영)"),
    ("방미라", "여", 3,  "정회원", "07:00", "12:00", "", "", "", "신혁재와 종료 30분 차이(자동 반영)"),
]

PREFILL_NOTE_FILL = PatternFill("solid", fgColor="FFF8DC")


def _build_players_sheet(ws, prefill: str = ""):
    headers = ["번호", "이름", "성별", "구력", "구분", "클럽", "IN시간", "OUT시간",
               "최소게임수", "최대게임수", "혼복희망", "연속게임", "메모"]
    widths = [6, 12, 8, 8, 12, 12, 10, 10, 12, 12, 10, 10, 28]
    for col_idx, (title, w) in enumerate(zip(headers, widths), start=1):
        cell = ws.cell(row=1, column=col_idx, value=title)
        _style_header(cell)
        ws.column_dimensions[get_column_letter(col_idx)].width = w

    prefill_rows = PREFILL_FROM_IMAGE if prefill == "image" else []

    for r in range(2, 32):
        ws.cell(row=r, column=1, value=r - 1)
        for c in range(1, 14):
            _style_body(ws.cell(row=r, column=c))
        ws.cell(row=r, column=7).number_format = "@"  # IN시간
        ws.cell(row=r, column=8).number_format = "@"  # OUT시간

        idx = r - 2
        if idx < len(prefill_rows):
            name, gender, exp, mem, in_t, out_t, min_g, max_g, streak, memo = prefill_rows[idx]
            if name: ws.cell(row=r, column=2, value=name)
            if gender: ws.cell(row=r, column=3, value=gender)
            if exp != "" and exp is not None: ws.cell(row=r, column=4, value=exp)
            if mem: ws.cell(row=r, column=5, value=mem)
            # 클럽(col 6)은 사전채움 비움 — 교류전 때만 직접 입력
            if in_t: ws.cell(row=r, column=7, value=in_t)
            if out_t: ws.cell(row=r, column=8, value=out_t)
            if min_g != "" and min_g is not None:
                ws.cell(row=r, column=9, value=min_g)
                ws.cell(row=r, column=9).fill = PREFILL_NOTE_FILL
            if max_g != "" and max_g is not None:
                ws.cell(row=r, column=10, value=max_g)
                ws.cell(row=r, column=10).fill = PREFILL_NOTE_FILL
            # 혼복희망(col 11)은 사전채움 비움 — 그 주에 요청한 사람만 직접 입력
            if streak:
                ws.cell(row=r, column=12, value=streak)
                ws.cell(row=r, column=12).fill = PREFILL_NOTE_FILL
            if memo:
                ws.cell(row=r, column=13, value=memo)
                ws.cell(row=r, column=13).fill = PREFILL_NOTE_FILL
            for c in (3, 4, 5, 6, 7, 8):
                if ws.cell(row=r, column=c).value in (None, ""):
                    ws.cell(row=r, column=c).fill = PREFILL_NOTE_FILL

    dv_gender = DataValidation(type="list", formula1='"남,여"', allow_blank=True)
    dv_gender.add("C2:C31")
    ws.add_data_validation(dv_gender)

    dv_member = DataValidation(type="list", formula1='"정회원,게스트"', allow_blank=True)
    dv_member.add("E2:E31")
    ws.add_data_validation(dv_member)

    dv_streak = DataValidation(type="list", formula1='"금지,허용"', allow_blank=True)
    dv_streak.add("L2:L31")
    ws.add_data_validation(dv_streak)

    ws.freeze_panes = "A2"


def _build_courts_sheet(ws):
    headers = ["코트명", "시작시간", "종료시간"]
    widths = [12, 12, 12]
    for col_idx, (title, w) in enumerate(zip(headers, widths), start=1):
        cell = ws.cell(row=1, column=col_idx, value=title)
        _style_header(cell)
        ws.column_dimensions[get_column_letter(col_idx)].width = w

    for r_idx, (name, start, end) in enumerate(COURTS_DEFAULT, start=2):
        ws.cell(row=r_idx, column=1, value=name)
        ws.cell(row=r_idx, column=2, value=start).number_format = "@"
        ws.cell(row=r_idx, column=3, value=end).number_format = "@"
        for c in range(1, 4):
            _style_body(ws.cell(row=r_idx, column=c))

    # 추가 빈 행
    for r in range(5, 12):
        for c in range(1, 4):
            _style_body(ws.cell(row=r, column=c))
            if c >= 2:
                ws.cell(row=r, column=c).number_format = "@"

    ws.freeze_panes = "A2"


def _seed_schedule_from_courts(courts_default: list[tuple[str, str, str]]) -> tuple[list[dict], list[dict]]:
    """COURTS_DEFAULT 튜플에서 (코트별 슬롯, 전체 스케줄 슬롯) 을 만든다.

    씨드 시트는 참가자 데이터가 아직 없는 '빈 템플릿 생성' 시점에 만들어지므로 parse_input의
    파싱 결과가 아니라 코트 기본값에서 직접 계산한다. parse_input.build_schedule_slots와
    같은 규칙(코트별 30분 슬롯의 합집합)을 그대로 따라야 결과 대진표(render_bracket.py)와
    슬롯 개수·순서가 어긋나지 않는다.
    """
    courts = []
    for name, start, end in courts_default:
        sh, sm = start.split(":")
        eh, em = end.split(":")
        s_min, e_min = int(sh) * 60 + int(sm), int(eh) * 60 + int(em)
        courts.append({"name": name, "slots": list(range(s_min, e_min, 30))})
    slot_set = sorted({s for c in courts for s in c["slots"]})
    schedule_slots = [{"slot_start": s, "slot_end": s + 30} for s in slot_set]
    return courts, schedule_slots


def _min_to_hhmm(v: int) -> str:
    return f"{v // 60:02d}:{v % 60:02d}"


def _build_seed_sheet(ws, courts_default: list[tuple[str, str, str]]):
    """'씨드대진' 시트 — 결과 대진표(render_bracket.py)와 셀 좌표를 1:1로 맞춘 빈 입력 그리드.

    사용자가 지난주 결과 엑셀에서 팀 블록을 그대로 복사해 붙여넣을 수 있는 것이 이 설계의
    핵심이므로, 컬럼 레이아웃(1=구분, 2=성함/결과, 3부터 코트당 5컬럼=[t1a,t1b,VS,t2a,t2b])과
    행 구조(슬롯당 2행: 성함/결과)를 render_bracket.py 336~479행과 그대로 맞춘다.
    여기서 어긋나면 붙여넣은 이름이 엉뚱한 슬롯·코트로 읽힌다(parse_seed는 헤더 행의
    'N번코트' 텍스트로 코트 위치를 스스로 찾긴 하지만, 슬롯 순서·간격은 이 구조를 전제한다).
    """
    courts, schedule_slots = _seed_schedule_from_courts(courts_default)

    LABEL_COL_START = 1
    LABEL_COL_END = 2
    COURTS_COL_START = 3
    cols_per_court = 5
    courts_col_end = COURTS_COL_START + cols_per_court * len(courts) - 1

    ws.column_dimensions[get_column_letter(LABEL_COL_START)].width = 9
    ws.column_dimensions[get_column_letter(LABEL_COL_END)].width = 7
    for i in range(len(courts)):
        base = COURTS_COL_START + i * cols_per_court
        for off, w in zip(range(cols_per_court), (10, 10, 4, 10, 10)):
            ws.column_dimensions[get_column_letter(base + off)].width = w

    # 행 1: 결과 대진표라면 타이틀이 들어갈 자리에 사용법 안내를 적는다
    usage = ("[선택] 씨드 대진 — 미리 정해두고 싶은 자리에만 이름을 적으세요. "
             "빈칸은 알고리즘이 알아서 채웁니다. 시트를 통째로 비워두면 기존과 똑같이 동작합니다. "
             "※ 이 격자는 '코트' 시트 기본값(A·B 08:00~12:00, C 07:00~09:00) 기준입니다 — "
             "코트 시간을 바꾸셨다면 회색/노란색 표시가 실제와 다를 수 있고, 판정은 '코트' 시트가 기준입니다.")
    ws.cell(row=1, column=LABEL_COL_START, value=usage)
    ws.merge_cells(start_row=1, start_column=LABEL_COL_START, end_row=1, end_column=courts_col_end)
    tc = ws.cell(row=1, column=LABEL_COL_START)
    tc.font = Font(name="맑은 고딕", size=11, bold=True)
    tc.alignment = LEFT
    tc.fill = GUIDE_FILL
    ws.row_dimensions[1].height = 30

    # 행 2: 헤더 — parse_seed는 이 'N번코트' 텍스트로 코트 위치(base 컬럼)를 스스로 찾는다
    HEADER_ROW = 2
    hc = ws.cell(row=HEADER_ROW, column=LABEL_COL_START, value="구분")
    ws.merge_cells(start_row=HEADER_ROW, start_column=LABEL_COL_START,
                    end_row=HEADER_ROW, end_column=LABEL_COL_END)
    _style_header(hc)

    for i, court in enumerate(courts):
        base = COURTS_COL_START + i * cols_per_court
        hc2 = ws.cell(row=HEADER_ROW, column=base, value=f"{court['name']}번코트")
        ws.merge_cells(start_row=HEADER_ROW, start_column=base,
                        end_row=HEADER_ROW, end_column=base + cols_per_court - 1)
        _style_header(hc2)

    # 데이터 행: 슬롯당 2행(성함/결과). 결과 행은 씨드에서 쓰지 않지만, 결과 엑셀과
    # 행 구조(2행씩)를 맞춰야 복붙했을 때 슬롯이 밀리지 않는다.
    data_start_row = HEADER_ROW + 1
    for slot_idx, slot in enumerate(schedule_slots):
        name_row = data_start_row + slot_idx * 2
        result_row = name_row + 1
        ws.row_dimensions[name_row].height = 24
        ws.row_dimensions[result_row].height = 24

        lc = ws.cell(row=name_row, column=LABEL_COL_START,
                      value=f"{slot_idx + 1}번 게임\n{_min_to_hhmm(slot['slot_start'])}\n~{_min_to_hhmm(slot['slot_end'])}")
        ws.merge_cells(start_row=name_row, start_column=LABEL_COL_START,
                        end_row=result_row, end_column=LABEL_COL_START)
        lc.font = BODY_FONT
        lc.alignment = CENTER
        lc.border = BORDER

        nc = ws.cell(row=name_row, column=LABEL_COL_END, value="성함")
        rc = ws.cell(row=result_row, column=LABEL_COL_END, value="결과")
        for cell in (nc, rc):
            _style_header(cell)

        for i, court in enumerate(courts):
            base = COURTS_COL_START + i * cols_per_court
            court_active = slot["slot_start"] in court["slots"]

            for col_off in range(cols_per_court):
                for r in (name_row, result_row):
                    cell = ws.cell(row=r, column=base + col_off)
                    cell.border = BORDER
                    cell.alignment = CENTER
                    cell.font = BODY_FONT
                    if not court_active:
                        cell.fill = INACTIVE_FILL   # 이 슬롯엔 이 코트가 운영하지 않음 — 입력칸 아님

            if not court_active:
                continue

            vs_cell = ws.cell(row=name_row, column=base + 2, value="VS")
            vs_cell.font = Font(name="맑은 고딕", size=10, bold=True, color="999999")
            vs_cell.fill = SEED_VS_FILL
            for off in (0, 1, 3, 4):   # 입력칸: t1a, t1b, t2a, t2b
                ws.cell(row=name_row, column=base + off).fill = GUIDE_FILL

    # 주의사항 (마지막 데이터 행 + 2)
    note_row = data_start_row + len(schedule_slots) * 2 + 1
    notes = [
        "· 이름은 '참가자' 시트와 똑같이 적어야 합니다. (게스트는 '홍길동(G)'처럼 적어도 인식됩니다)",
        "· 왼쪽 두 칸이 한 팀, 오른쪽 두 칸이 상대 팀입니다. 적은 자리 그대로 고정됩니다.",
        "· 남자와 여자를 같은 경기에 적으면 그 경기는 혼합복식이 됩니다.",
    ]
    for i, t in enumerate(notes):
        nc2 = ws.cell(row=note_row + i, column=1, value=t)
        nc2.font = Font(name="맑은 고딕", size=9, color="666666")
        nc2.alignment = LEFT

    ws.freeze_panes = f"C{data_start_row}"


def _build_guide_sheet(ws):
    ws.column_dimensions["A"].width = 100
    lines = [
        ("우리 테니스 클럽 대진표 — 입력 양식 작성 안내", True),
        ("", False),
        ("■ 1. 참가자 시트 작성법", True),
        ("• 번호: 자동 채워져 있음 (수정 불필요)", False),
        ("• 이름: 클럽 내 중복 없게 입력", False),
        ("• 성별: 드롭다운에서 '남' / '여' 선택 (필수)", False),
        ("• 구력: 테니스 경력 년수, 정수로 입력 — 예: 3, 10 (필수)", False),
        ("• 구분: '정회원' / '게스트' 선택 (필수)", False),
        ("• 클럽: 평소엔 비워두세요(자동으로 '우리클럽'). 교류전 때만 클럽명 입력 (선택)", False),
        ("    - 둘 이상의 클럽명이 들어오면 '교류전 모드'가 켜져 같은 클럽끼리만 한 팀이 됩니다", False),
        ("    - 예: 우리 회원은 '우리클럽', 방문팀은 '○○클럽' 식으로 구분", False),
        ("• IN시간 / OUT시간: HH:MM 형식, 30분 단위 (예: 08:00, 08:30) (필수)", False),
        ("    - IN은 코트장에 들어올 수 있는 가장 빠른 시각", False),
        ("    - OUT은 코트장을 떠나야 하는 시각", False),
        ("    - 반드시 IN < OUT, 둘 다 30분 단위 (00분 또는 30분)", False),
        ("• 최소게임수: 이 사람에게 반드시 보장할 최소 게임 수 (정수, 선택)", False),
        ("    - 넣으면 적어도 그 게임수 이상 배정됩니다 (어기지 않는 규칙)", False),
        ("    - 예: '4게임은 꼭 뛰고 싶어요' → 4 입력", False),
        ("    - 단, 본인 IN~OUT 안의 슬롯 수보다 크게 적으면 슬롯 수까지만 보장", False),
        ("• 최대게임수: 이 사람이 출전할 수 있는 최대 게임 수 (정수, 선택)", False),
        ("    - 비우면 무제한 (IN~OUT 범위 안에서 알고리즘이 자동 분배)", False),
        ("    - 예: '3게임만 하고 갈게요' → 3 입력", False),
        ("    - 최소게임수와 같이 적으면 최소 ≤ 최대여야 합니다", False),
        ("• 혼복희망: 그중 혼복(남녀 섞인 복식)으로 뛰고 싶은 게임 수 (정수, 선택)", False),
        ("    - 예: '혼복으로 2게임은 하고 싶어요' → 2 입력", False),
        ("    - 비워 두는 것이 기본입니다. 원칙은 남복/여복이라 혼복은 0판일 수 있습니다", False),
        ("    - 적은 사람이 있으면 그만큼 혼복 판을 만들고 그 자리에 그 사람을 넣습니다", False),
        ("• 연속게임: 빈칸=기본(2연속까지 가능·3연속 금지), '금지'=한 게임 뒤 반드시 쉼, '허용'=3연속까지 가능(4연속은 그래도 금지)", False),
        ("    - '허용'은 짧은 시간에 많은 게임을 보장해야 할 때만 사용하세요", False),
        ("• 메모: 자유 기재 (선택, 알고리즘에 영향 없음)", False),
        ("", False),
        ("■ 2. 코트 시트 작성법", True),
        ("• 기본값: A코트=1번(08:00-12:00), B코트=2번(08:00-12:00), C코트=3번(07:00-09:00)", False),
        ("    - 2026.9.5부터 1·2번 코트는 08:00 시작 (3번 코트는 07:00~09:00 그대로)", False),
        ("• 코트가 더 있거나 운영시간이 다르면 행을 수정/추가하세요", False),
        ("• 시간은 30분 단위, 시작 < 종료", False),
        ("• 사용하지 않는 행은 빈 칸으로 두면 됩니다", False),
        ("", False),
        ("■ 3. 씨드대진 시트 작성법 (선택)", True),
        ("• 대진의 일부를 미리 정해두고 싶을 때만 사용 — 완전히 비워두면 기존과 똑같이 동작합니다", False),
        ("• 결과 대진표와 셀 위치가 똑같습니다 — 지난주 결과 엑셀에서 팀 블록을 복사해 그대로", False),
        ("  붙여넣을 수 있습니다", False),
        ("• 이름은 '참가자' 시트와 똑같이 적어야 인식됩니다 (게스트는 '홍길동(G)'도 인식)", False),
        ("• 왼쪽 두 칸이 한 팀, 오른쪽 두 칸이 상대 팀 — 적은 자리는 그대로 고정되고", False),
        ("  나머지는 알고리즘이 채웁니다", False),
        ("• 남자와 여자를 같은 경기에 적으면 그 경기는 자동으로 혼합복식이 됩니다", False),
        ("• 회색 칸은 그 시간에 운영하지 않는 코트라 입력할 수 없습니다", False),
        ("• 같은 사람을 같은 시간대에 두 곳에 적거나, 본인 IN~OUT 시간 밖에 적으면 오류로 표시되고", False),
        ("  대진표 생성이 중단됩니다", False),
        ("", False),
        ("■ 4. 자동 배정 규칙 (알고리즘 동작 원리)", True),
        ("", False),
        ("[A] 기본 단위", True),
        ("• 1게임 = 30분", False),
        ("• 한 사람은 같은 시간 슬롯에 두 개 코트 동시 출전 불가", False),
        ("• 각자 본인 IN~OUT 범위 안의 슬롯에만 배정됨", False),
        ("", False),
        ("[B] 복식 종류 우선순위", True),
        ("• 남자복식·여자복식이 원칙 — 만들 수 있으면 항상 이쪽이 먼저", False),
        ("• 다만 여자가 6명 이하면 혼합복식을 기본으로 1~2판 넣습니다", False),
        ("    (여자복식만 돌리면 늘 같은 사람끼리 붙게 되므로)", False),
        ("• 그 밖에 혼합복식이 늘어나는 경우 — 단성 복식을 더 짜도 '쓸 만한 대진'이", False),
        ("  안 나올 때입니다. 쓸 만하다 = ① 같은 짝이 처음이고 ② 같은 편 구성으로 다시", False),
        ("  붙는 게 아니고 ③ 두 팀 구력 합 차이가 허용치 이내.", False),
        ("      허용치는 출전 4명이 모두 구력 10년 미만이면 3, 10년 이상인 분이 끼면 4입니다.", False),
        ("    예) 방미라(3)+노남숙(3)=6 vs 서명숙(5)+정정희(5)=10 은 4 차이인데", False),
        ("        전원 10년 미만이라 허용치가 3 → 이 조합은 만들지 않습니다.", False),
        ("        쓸 만한 조합이 모자라면 그만큼 혼합복식으로 채웁니다.", False),
        ("• 남자 또는 여자가 4명이 안 되면 단성 복식 자체가 불가능 → 혼합복식만 나옵니다", False),
        ("• 혼합복식 규칙: 같은 팀의 남자 구력 ≥ 여자 구력", False),
        ("    (남자가 같은 팀 여자보다 경력이 같거나 더 많아야 함)", False),
        ("• 남자 게스트는 혼합복식보다 남자복식 위주로 배정합니다", False),
        ("    (혼합복식의 남자 자리는 가급적 정회원이 맡음 — 여자 게스트는 혼합복식 가능)", False),
        ("", False),
        ("[C] 코트별 우선 배정", True),
        ("• A코트: 여자복식 + 혼합복식 우선", False),
        ("• B코트: 남자복식 우선", False),
        ("• C코트: 무관 (균등 배분)", False),
        ("  ※ 단, 인원 부족 시 위 우선순위는 양보될 수 있음", False),
        ("", False),
        ("[D] 시간 제약", True),
        ("• 각자 IN~OUT 시간 안에서만 배정 (성별에 따른 시간대 제한은 없음)", False),
        ("", False),
        ("[E] 팀 매칭 (재미와 균형)", True),
        ("• 두 팀의 합산 구력이 비슷하도록 매칭 (예: 5+7=12 vs 6+6=12)", False),
        ("• 합산 구력 차이가 허용치(전원 10년 미만 3 / 10년 이상 포함 4)를 넘으면 강하게 회피 —", False),
        ("  그렇게 될 바에는 다른 조합이나 혼합복식으로 돌립니다", False),
        ("• 한 번 같은 팀이었던 페어는 가능한 한 다시 같은 팀이 안 되게", False),
        ("• 같은 4명이 다시 붙는 것을 강하게 회피 — 상대편까지 그대로든, 편만 바꾸든", False),
        ("  (그렇게 될 바에는 혼합복식으로 돌립니다)", False),
        ("    ※ 예외: 그 성별로 만들 4명 조합이 하나뿐일 때(예: 여자가 정확히 4명)는", False),
        ("       편을 바꾸면 짝도 상대도 매번 새로우므로 여자복식 3판까지 그대로 만듭니다.", False),
        ("       (여기서 막으면 여자복식이 1판에서 끝나고 나머지가 전부 혼합복식이 됩니다)", False),
        ("• 한 팀에 정회원+게스트 혼합 약하게 권장 (강제 아님)", False),
        ("", False),
        ("[F] 게임수 균형", True),
        ("• 가용 시간이 같은 사람끼리는 게임수 차이를 최대 1게임까지만 (2게임 차이 나면 재시도)", False),
        ("  (최소·최대게임수를 적은 사람은 의도적으로 다르므로 이 비교에서 제외)", False),
        ("• 남녀 평균 게임수가 한쪽만 낮아지지 않게 별도로 확인", False),
        ("  (여자 인원이 적으면 복식 특성상 '4명 통째로'만 늘어나 손해 보기 쉬움)", False),
        ("• 일찍 와서 늦게 가는 사람은 자연스럽게 더 많이 배정", False),
        ("• 늦게 오거나 일찍 가는 사람은 그만큼 적게 배정", False),
        ("• 전체 격차는 4 이하 권장 (가용 인원이 빠듯하면 더 클 수 있음)", False),
        ("", False),
        ("[G] 쉬는 텀 · 끝나는 시간", True),
        ("• 경기 사이에 1시간 이상 쉬었다 다시 나오는 일을 최우선으로 줄임", False),
        ("  (30분 쉬고 다음 경기가 기본 리듬. 2게임 연속은 정상)", False),
        ("• 한 사람이 3슬롯(1.5시간) 연속 출전은 금지 — 단, 개인별 연속게임 '허용' 지정자는 면제", False),
        ("• 도착(IN) 후 첫 경기까지 오래 기다리지 않게 배정", False),
        ("• 노는 시간이 한 사람에게 몰리지 않게 고르게 분산", False),
        ("  → 결과적으로 일찍 온 사람이 일찍 끝나고, 늦게 온 사람이 뒤쪽을 맡음", False),
        ("  (가용 인원이 코트수×4보다 훨씬 많으면 공백이 구조적으로 불가피)", False),
        ("", False),
        ("[H] 개인별 최소·최대 게임수", True),
        ("• 최소게임수 칸에 정수 입력 시 적어도 그 수 이상 배정 (어기지 않는 규칙)", False),
        ("    - 본인 가용 슬롯 수를 넘는 값은 슬롯 수까지만 보장", False),
        ("    - 여러 명의 최소게임수 합이 전체 자리보다 크면 다 지킬 수 없음 (경고 표시)", False),
        ("• 최대게임수 칸에 정수 입력 시 정확히 그 수 이하로 배정 (hard 제약)", False),
        ("• 빈 칸이면 IN~OUT 범위 안에서 균형 배정", False),
        ("• 혼복희망 칸에 정수 입력 시 그만큼 혼복 판을 열고 그 사람을 그 자리에 배정 (소프트)", False),
        ("    - 아무도 안 적으면 지금까지와 동일 — 남복/여복 우선 원칙 그대로", False),
        ("    - 남자 게스트도 혼복희망을 적으면 '남복 위주' 규칙에서 본인만 예외가 됩니다", False),
        ("    - 구력·공백·최소게임수 같은 더 강한 규칙에 밀리면 못 채울 수 있음 (검토 보고에 표시)", False),
        ("", False),
        ("[I] 교류전 (클럽 대항)", True),
        ("• '클럽' 칸에 둘 이상의 클럽명이 있으면 교류전 모드 자동 작동", False),
        ("• 한 팀(복식 2명)은 반드시 같은 클럽끼리 구성 (hard 제약)", False),
        ("• 모든 경기는 반드시 상대 클럽과 (같은 클럽끼리는 절대 안 붙음, hard)", False),
        ("• 게임수 균형은 '각 클럽 내부'에서만 — 두 클럽 인원이 다르면 클럽 간 평균은 다를 수 있음", False),
        ("• 클럽 칸이 모두 비었거나 한 종류면 평소대로 동작 (영향 없음)", False),
        ("• 주의: 상대 클럽 인원이 부족한 슬롯에선 남는 사람이 쉬고 그 코트는 공석이 될 수 있음", False),
        ("", False),
        ("■ 5. 알고리즘이 회피 못 하는 입력 구조 (참고)", True),
        ("• 여자 인원이 4명 미만 → 여자복식 불가, 혼복만 가능", False),
        ("• 남자 인원이 4명 미만 → 남자복식 불가, 혼복만 가능", False),
        ("• 여자 최고 구력 > 남자 최고 구력 → 혼복 시 일부 규칙 위반 불가피", False),
        ("    (예: 김승회 10년 vs 남자 최고 8년)", False),
        ("• 가용 인원 = 코트수×4 → 휴식 자리 없어 연속 출전 발생", False),
        ("• 특정 슬롯에 가용 인원 4명 미만 → 그 코트는 자동 공석", False),
        ("", False),
        ("■ 6. 결과 생성 방법", True),
        ("• Windows 명령창(또는 PowerShell)에서:", False),
        ('     python 대진표_생성.py --date "26.5.30"', False),
        ("• 또는 대진표_생성.bat 더블클릭", False),
        ("• 결과 파일: 출력/테니스_대진표_<날짜>.xlsx", False),
        ("• 채운 양식을 먼저 검사만 하려면 (대진표 생성 안 함):", False),
        ("     python 대진표_생성.py --check   (또는 대진표_생성.bat 실행 후 C 입력)", False),
        ("     - 오류·경고, 인원·자리 요약, 최소/최대게임수 지정 현황을 보여줍니다", False),
        ("", False),
        ("■ 7. 다시 생성하고 싶을 때", True),
        ("• 같은 입력으로 다른 패턴 원하면 --seed 99 (또는 다른 숫자)", False),
        ("• 입력 양식 수정 후 다시 실행하면 새로 생성", False),
        ("", False),
        ("■ 8. 파일 위치", True),
        ("• 입력 양식: 입력/테니스_입력양식.xlsx", False),
        ("• 결과 파일: 출력/테니스_대진표_<YYMMDD>.xlsx", False),
        ("• 샘플/참고: 샘플/  (이미지·예시 데이터)", False),
    ]
    for i, (text, bold) in enumerate(lines, start=1):
        cell = ws.cell(row=i, column=1, value=text)
        cell.alignment = LEFT
        cell.font = Font(name="맑은 고딕", size=12 if bold else 11, bold=bold)
        if bold and text:
            cell.fill = GUIDE_FILL


def build_member_settings(out_path) -> None:
    """멤버 설정 엑셀 생성: '멤버'(카톡아이디↔실제이름) + '부부'(부부 페어와 혼복 페어 희망).

    웹 '멤버 설정 다운로드' 버튼과 CLI --create-members가 사용.
    수정해서 웹에 업로드하거나 입력/클럽멤버_설정.xlsx 로 두면 대진 생성에 반영된다.
    """
    from parse_input import MEMBERS_DEFAULT, COUPLES_DEFAULT

    wb = Workbook()

    ws = wb.active
    ws.title = "멤버"
    for col_idx, (title, w) in enumerate(
            [("번호", 6), ("카톡아이디", 18), ("실제이름", 14), ("성별", 8), ("구력", 8), ("메모", 28)], start=1):
        cell = ws.cell(row=1, column=col_idx, value=title)
        _style_header(cell)
        ws.column_dimensions[get_column_letter(col_idx)].width = w
    for i, (kid, real, gender, exp, memo) in enumerate(MEMBERS_DEFAULT, 1):
        for c, v in ((1, i), (2, kid), (3, real), (4, gender), (5, exp), (6, memo)):
            cell = ws.cell(row=i + 1, column=c, value=v if v != "" else None)
            _style_body(cell)
        if memo:
            ws.cell(row=i + 1, column=5).fill = PREFILL_NOTE_FILL   # 확인 필요한 구력 강조
    dv_g = DataValidation(type="list", formula1='"남,여"', allow_blank=True)
    dv_g.add(f"D2:D{len(MEMBERS_DEFAULT) + 6}")
    ws.add_data_validation(dv_g)
    ws.freeze_panes = "A2"

    ws2 = wb.create_sheet("부부")
    for col_idx, (title, w) in enumerate(
            [("이름1", 14), ("이름2", 14), ("부부페어", 12), ("종료시간차", 12), ("메모", 30)], start=1):
        cell = ws2.cell(row=1, column=col_idx, value=title)
        _style_header(cell)
        ws2.column_dimensions[get_column_letter(col_idx)].width = w
    for i, (a, b, want, gap) in enumerate(COUPLES_DEFAULT, 2):
        vals = (a, b, "원함" if want else "피함", "30분차이" if gap else "", "")
        for c, v in enumerate(vals, 1):
            cell = ws2.cell(row=i, column=c, value=v if v != "" else None)
            _style_body(cell)
    # 추가 기입용 빈 행
    for r in range(len(COUPLES_DEFAULT) + 2, len(COUPLES_DEFAULT) + 8):
        for c in range(1, 6):
            _style_body(ws2.cell(row=r, column=c))
    dv = DataValidation(type="list", formula1='"원함,피함"', allow_blank=True)
    dv.add(f"C2:C{len(COUPLES_DEFAULT) + 7}")
    ws2.add_data_validation(dv)
    dv_gap = DataValidation(type="list", formula1='"같이,30분차이"', allow_blank=True)
    dv_gap.add(f"D2:D{len(COUPLES_DEFAULT) + 7}")
    ws2.add_data_validation(dv_gap)
    note_row = len(COUPLES_DEFAULT) + 9
    notes = [
        "■ 부부페어: '원함' = 혼복이 나올 때 부부가 같은 팀이 되게 우대 / '피함' = 같은 팀이 안 되게 회피",
        "■ 종료시간차: 빈칸/'같이' = 같이 끝나거나 30분 이내 차이 / '30분차이' = 마지막 경기 종료가 반드시 30분 차이 나게",
        "■ 이름은 입력 양식의 참가자 이름(실제이름)과 똑같이 적어야 반영됩니다",
        "■ 이 파일을 웹의 '멤버 설정' 칸에 올리거나, 입력/클럽멤버_설정.xlsx 로 두면 반영됩니다 (없으면 내장 기본값)",
    ]
    for i, t in enumerate(notes):
        c = ws2.cell(row=note_row + i, column=1, value=t)
        c.font = Font(name="맑은 고딕", size=10, color="666666")
        c.alignment = LEFT

    wb.save(out_path)


def build_template(out_path: str, prefill: str = "") -> None:
    wb = Workbook()
    ws_players = wb.active
    ws_players.title = "참가자"
    _build_players_sheet(ws_players, prefill=prefill)

    ws_courts = wb.create_sheet("코트")
    _build_courts_sheet(ws_courts)

    ws_seed = wb.create_sheet("씨드대진")
    _build_seed_sheet(ws_seed, COURTS_DEFAULT)

    ws_guide = wb.create_sheet("안내")
    _build_guide_sheet(ws_guide)

    # 시트 순서: 안내, 참가자, 코트, 씨드대진
    wb.move_sheet(ws_guide, offset=-3)

    wb.save(out_path)


def main():
    p = argparse.ArgumentParser()
    p.add_argument("--out", required=True, help="출력 .xlsx 경로")
    p.add_argument("--prefill", default="", choices=["", "image"],
                   help="image: 우리클럽 멤버 명단으로 사전 채움")
    p.add_argument("--member-settings", action="store_true",
                   help="입력 템플릿 대신 멤버 설정 파일(멤버·부부 시트)을 생성")
    args = p.parse_args()
    if args.member_settings:
        build_member_settings(args.out)
        print(f"[OK] 멤버 설정 파일 생성: {args.out}  (멤버 25명 + 부부 페어 시트)")
        return
    build_template(args.out, prefill=args.prefill)
    msg = f"[OK] 입력 템플릿 생성: {args.out}"
    if args.prefill == "image":
        msg += "  (우리클럽 멤버 사전 채움 — 클럽 칸은 교류전 때만 입력)"
    print(msg)


if __name__ == "__main__":
    main()
